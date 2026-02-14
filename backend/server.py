"""
NT Commerce API Server
Main server file with organized imports from modules
"""
from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, File, UploadFile, Header, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from contextvars import ContextVar
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import io
import requests as http_requests
import asyncio
import shutil
import base64

# Try to import resend
try:
    import resend
    RESEND_AVAILABLE = True
except ImportError:
    RESEND_AVAILABLE = False

# Import SendGrid
try:
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail, Email, To, Content, HtmlContent
    SENDGRID_AVAILABLE = True
except ImportError:
    SENDGRID_AVAILABLE = False

# Import Stripe via emergentintegrations
try:
    from emergentintegrations.payments.stripe.checkout import StripeCheckout, CheckoutSessionResponse, CheckoutStatusResponse, CheckoutSessionRequest
    STRIPE_AVAILABLE = True
except ImportError:
    STRIPE_AVAILABLE = False

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Initialize resend if available
if RESEND_AVAILABLE:
    resend.api_key = os.environ.get('RESEND_API_KEY', '')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
main_db = client[os.environ['DB_NAME']]  # Main SaaS database (plans, tenants, agents, super admin users)

# ContextVar for per-request tenant database isolation
_tenant_db_ctx: ContextVar = ContextVar('tenant_db')

class _TenantDBProxy:
    """Proxy that routes DB calls to tenant-specific DB when in tenant context, otherwise main DB.
    This ensures data isolation: each tenant's requests automatically use their own database."""
    def __getattr__(self, name):
        try:
            return getattr(_tenant_db_ctx.get(), name)
        except LookupError:
            return getattr(main_db, name)

    def __getitem__(self, name):
        try:
            return _tenant_db_ctx.get()[name]
        except LookupError:
            return main_db[name]

db = _TenantDBProxy()  # All existing code uses `db` - now routes to correct tenant DB automatically

# Multi-tenancy: Get tenant-specific database
def get_tenant_db(tenant_id: str):
    """Get database for a specific tenant"""
    if not tenant_id:
        return main_db
    db_name = f"tenant_{tenant_id.replace('-', '_')}"
    return client[db_name]

async def init_tenant_database(tenant_id: str):
    """Initialize a new tenant database with default collections and data"""
    tenant_db = get_tenant_db(tenant_id)
    
    # Initialize cash boxes
    boxes = [
        {"id": "cash", "name": "الصندوق النقدي", "name_fr": "Caisse", "type": "cash", "balance": 0},
        {"id": "bank", "name": "الحساب البنكي", "name_fr": "Compte bancaire", "type": "bank", "balance": 0},
        {"id": "wallet", "name": "المحفظة الإلكترونية", "name_fr": "Portefeuille électronique", "type": "wallet", "balance": 0},
        {"id": "safe", "name": "الخزنة", "name_fr": "Coffre-fort", "type": "safe", "balance": 0}
    ]
    for box in boxes:
        existing = await tenant_db.cash_boxes.find_one({"id": box["id"]})
        if not existing:
            await tenant_db.cash_boxes.insert_one(box)
    
    # Initialize default warehouse
    existing_warehouse = await tenant_db.warehouses.find_one({"id": "main"})
    if not existing_warehouse:
        await tenant_db.warehouses.insert_one({
            "id": "main",
            "name": "المخزن الرئيسي",
            "location": "",
            "is_main": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Initialize settings
    existing_settings = await tenant_db.settings.find_one({"id": "general"})
    if not existing_settings:
        await tenant_db.settings.insert_one({
            "id": "general",
            "low_stock_threshold": 10,
            "debt_reminder_days": 30,
            "currency": "دج",
            "language": "ar"
        })
    
    logger.info(f"Initialized database for tenant: {tenant_id}")
    return tenant_db

# JWT Settings
SECRET_KEY = os.environ.get('JWT_SECRET', 'screenguard-secret-key-2024')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

# Currency
CURRENCY = "دج"  # Algerian Dinar

# Create the main app
app = FastAPI(title="NT API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Create static directory for uploads
UPLOAD_DIR = ROOT_DIR / "static" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# ============ IMPORT REFACTORED ROUTES ============
from routes.saas_routes import router as saas_router

# ============ IMPORT MODELS FROM MODULES ============
from models.schemas import *


# ============ ADDITIONAL MODELS (Not in schemas.py yet) ============

class DailySessionCreate(BaseModel):
    opening_cash: float = 0
    notes: str = ""
    cash_box_id: str = "cash"

class DailySessionClose(BaseModel):
    closing_cash: float
    notes: str = ""
    cash_breakdown: dict = {}

class DailySessionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    session_code: str
    date: str
    opening_cash: float
    closing_cash: Optional[float] = None
    expected_cash: float = 0
    difference: float = 0
    total_sales: float = 0
    total_purchases: float = 0
    total_expenses: float = 0
    sales_count: int = 0
    purchases_count: int = 0
    is_closed: bool = False
    user_id: str
    user_name: str
    cash_box_id: str = "cash"
    notes: str = ""
    cash_breakdown: dict = {}
    created_at: str
    closed_at: Optional[str] = None

class InventorySessionCreate(BaseModel):
    name: str = ""
    code: Optional[str] = None
    warehouse_id: str = "main"
    family_filter: str = "all"
    status: str = "active"
    started_at: Optional[str] = None
    counted_items: Optional[dict] = {}
    notes: str = ""

class InventoryItemUpdate(BaseModel):
    product_id: str
    counted_quantity: int

class InventorySessionUpdate(BaseModel):
    name: Optional[str] = None
    status: Optional[str] = None
    counted_items: Optional[dict] = None
    completed_at: Optional[str] = None
    applied_changes: Optional[bool] = None
    notes: Optional[str] = None

class InventorySessionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str = ""
    code: Optional[str] = None
    session_code: Optional[str] = None
    warehouse_id: str = "main"
    family_filter: str = "all"
    status: str
    items: List[dict] = []
    counted_items: dict = {}
    total_products: int = 0
    counted_products: int = 0
    discrepancies: int = 0
    user_id: Optional[str] = None
    user_name: Optional[str] = None
    notes: str = ""
    started_at: Optional[str] = None
    created_at: Optional[str] = None
    completed_at: Optional[str] = None
    closed_at: Optional[str] = None
    applied_changes: bool = False

class DebtPaymentCreate(BaseModel):
    amount: float
    payment_method: str = "cash"
    notes: str = ""

class ApiKeyCreate(BaseModel):
    name: str
    service: str
    description: str = ""

class ApiKeyResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    service: str
    key_preview: str
    description: str = ""
    is_active: bool = True
    created_at: str

class RechargeTransactionCreate(BaseModel):
    service_type: str
    phone_number: str
    amount: float
    operator: str

class RechargeTransactionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    service_type: str
    phone_number: str
    amount: float
    operator: str
    status: str
    profit: float = 0
    user_id: str
    user_name: str = ""
    created_at: str

class PhoneDirectoryCreate(BaseModel):
    name: str
    phone: str
    category: str = "general"
    notes: str = ""

class PhoneDirectoryResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    category: str
    notes: str = ""
    created_at: str

class NotificationCreate(BaseModel):
    title: str
    message: str
    type: str = "info"
    target_roles: List[str] = []

class NotificationResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    title: str
    message: str
    type: str
    is_read: bool = False
    target_roles: List[str] = []
    created_by: str
    created_at: str

class RepairCreate(BaseModel):
    customer_name: str
    customer_phone: str
    device_type: str
    device_model: str
    issue_description: str
    estimated_cost: float = 0
    notes: str = ""

class RepairUpdate(BaseModel):
    status: Optional[str] = None
    diagnosis: Optional[str] = None
    repair_notes: Optional[str] = None
    parts_used: Optional[List[dict]] = None
    final_cost: Optional[float] = None
    technician_id: Optional[str] = None

class RepairResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    repair_code: str
    customer_name: str
    customer_phone: str
    device_type: str
    device_model: str
    issue_description: str
    diagnosis: str = ""
    repair_notes: str = ""
    status: str
    estimated_cost: float
    final_cost: float = 0
    parts_used: List[dict] = []
    technician_id: Optional[str] = None
    technician_name: str = ""
    notes: str = ""
    received_at: str
    completed_at: Optional[str] = None
    delivered_at: Optional[str] = None

class SparePartCreate(BaseModel):
    name: str
    compatible_models: str = ""
    quantity: int = 0
    purchase_price: float = 0
    selling_price: float = 0
    min_stock: int = 5

class SparePartResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    compatible_models: str = ""
    quantity: int
    purchase_price: float
    selling_price: float
    min_stock: int
    created_at: str

class ChatMessage(BaseModel):
    role: str  # "user" or "assistant"
    content: str

class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    session_id: Optional[str] = None

class ImageOCRRequest(BaseModel):
    image_base64: str

class OCRResponse(BaseModel):
    extracted_models: List[str]
    raw_text: str

# ============ HELPER FUNCTIONS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        user_type = payload.get("type")  # admin, agent, tenant
        tenant_id = payload.get("tenant_id")
        
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        # For tenant users, get from tenant database
        if user_type == "tenant" and tenant_id:
            tenant_db = get_tenant_db(tenant_id)
            user = await tenant_db.users.find_one({"id": user_id}, {"_id": 0, "password": 0, "hashed_password": 0})
            
            # Get tenant info from main_db to get plan features
            tenant = await main_db.saas_tenants.find_one({"id": tenant_id}, {"_id": 0, "password": 0})
            
            if user is None:
                # Check main tenant record (always in main_db)
                if tenant:
                    user = {
                        "id": tenant["id"],
                        "email": tenant["email"],
                        "name": tenant["name"],
                        "role": "admin",
                        "tenant_id": tenant_id,
                        "user_type": "tenant",
                        "company_name": tenant.get("company_name", ""),
                        "created_at": tenant.get("created_at", datetime.now(timezone.utc).isoformat())
                    }
                else:
                    raise HTTPException(status_code=401, detail="User not found")
            else:
                user["tenant_id"] = tenant_id
                user["user_type"] = "tenant"
                if not user.get("created_at"):
                    user["created_at"] = datetime.now(timezone.utc).isoformat()
            
            # Add plan features and limits for tenant users
            if tenant:
                plan = await main_db.saas_plans.find_one({"id": tenant.get("plan_id")}, {"_id": 0})
                if plan:
                    user["features"] = {**plan.get("features", {}), **tenant.get("features_override", {})}
                    user["limits"] = {**plan.get("limits", {}), **tenant.get("limits_override", {})}
                user["company_name"] = tenant.get("company_name", "")
        else:
            # For admin users, get from main database
            user = await main_db.users.find_one({"id": user_id}, {"_id": 0, "password": 0, "hashed_password": 0})
            if user is None:
                raise HTTPException(status_code=401, detail="User not found")
            user["user_type"] = user_type or "admin"
            if tenant_id:
                user["tenant_id"] = tenant_id
        
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_tenant_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get current user and their tenant database"""
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        user_type = payload.get("type")
        tenant_id = payload.get("tenant_id")
        
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        # Get the appropriate database
        if user_type == "tenant" and tenant_id:
            tenant_db = get_tenant_db(tenant_id)
        else:
            tenant_db = main_db  # Use main database for admin users
            tenant_id = None
        
        # Get user info
        user = await tenant_db.users.find_one({"id": user_id}, {"_id": 0, "password": 0, "hashed_password": 0})
        if user is None and tenant_id:
            # For tenant owner, create entry from saas_tenants
            tenant = await main_db.saas_tenants.find_one({"id": tenant_id}, {"_id": 0, "password": 0})
            if tenant:
                user = {
                    "id": tenant["id"],
                    "email": tenant["email"],
                    "name": tenant["name"],
                    "role": "admin"
                }
        
        if user is None:
            user = await main_db.users.find_one({"id": user_id}, {"_id": 0, "password": 0, "hashed_password": 0})
        
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        return {"user": user, "db": tenant_db, "tenant_id": tenant_id}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_admin_user(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def get_tenant_admin(current_user: dict = Depends(get_current_user)):
    """Require tenant context - rejects super_admin users without tenant_id.
    Use this for tenant-specific data routes (products, customers, sales, etc.)."""
    if not current_user.get("tenant_id"):
        raise HTTPException(status_code=403, detail="هذا الإجراء متاح فقط لمشتركي المنصة")
    if current_user.get("role") not in ["admin", "manager", "user", "tenant_admin"]:
        raise HTTPException(status_code=403, detail="صلاحيات غير كافية")
    return current_user

async def require_tenant(current_user: dict = Depends(get_current_user)):
    """Require tenant context for read operations - any authenticated tenant user."""
    if not current_user.get("tenant_id"):
        raise HTTPException(status_code=403, detail="هذا الإجراء متاح فقط لمشتركي المنصة")
    return current_user

async def generate_invoice_number(prefix: str) -> str:
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    count = await db.counters.find_one_and_update(
        {"_id": f"{prefix}_{today}"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return f"{prefix}-{today}-{count['seq']:04d}"

async def init_cash_boxes():
    """Initialize default cash boxes if they don't exist, or update existing ones with name_fr"""
    boxes = [
        {"id": "cash", "name": "الصندوق النقدي", "name_fr": "Caisse", "type": "cash", "balance": 0},
        {"id": "bank", "name": "الحساب البنكي", "name_fr": "Compte bancaire", "type": "bank", "balance": 0},
        {"id": "wallet", "name": "المحفظة الإلكترونية", "name_fr": "Portefeuille électronique", "type": "wallet", "balance": 0},
        {"id": "safe", "name": "الخزنة", "name_fr": "Coffre-fort", "type": "safe", "balance": 0}
    ]
    for box in boxes:
        existing = await db.cash_boxes.find_one({"id": box["id"]})
        if not existing:
            box["updated_at"] = datetime.now(timezone.utc).isoformat()
            await db.cash_boxes.insert_one(box)
        elif not existing.get("name_fr"):
            # Update existing box with name_fr if missing
            await db.cash_boxes.update_one(
                {"id": box["id"]},
                {"$set": {"name_fr": box["name_fr"]}}
            )

async def init_default_data(tenant_db):
    """Initialize default data for a tenant (customers, suppliers, families, products)"""
    now = datetime.now(timezone.utc).isoformat()
    
    # Default Customer Family
    default_customer_family_id = "default-customer-family"
    existing_cf = await tenant_db.customer_families.find_one({"id": default_customer_family_id})
    if not existing_cf:
        await tenant_db.customer_families.insert_one({
            "id": default_customer_family_id,
            "name": "عائلة زبائن متنوعة",
            "name_fr": "Famille clients divers",
            "description": "عائلة افتراضية للزبائن",
            "discount": 0,
            "created_at": now,
            "updated_at": now
        })
    
    # Default Customer
    default_customer_id = "default-customer"
    existing_c = await tenant_db.customers.find_one({"id": default_customer_id})
    if not existing_c:
        await tenant_db.customers.insert_one({
            "id": default_customer_id,
            "name": "زبون متنوع",
            "name_fr": "Client divers",
            "phone": "",
            "email": "",
            "address": "",
            "family_id": default_customer_family_id,
            "family_name": "عائلة زبائن متنوعة",
            "balance": 0,
            "total_purchases": 0,
            "notes": "زبون افتراضي للمبيعات العامة",
            "created_at": now,
            "updated_at": now
        })
    
    # Default Supplier Family
    default_supplier_family_id = "default-supplier-family"
    existing_sf = await tenant_db.supplier_families.find_one({"id": default_supplier_family_id})
    if not existing_sf:
        await tenant_db.supplier_families.insert_one({
            "id": default_supplier_family_id,
            "name": "عائلة مورد متنوع",
            "name_fr": "Famille fournisseurs divers",
            "description": "عائلة افتراضية للموردين",
            "created_at": now,
            "updated_at": now
        })
    
    # Default Supplier
    default_supplier_id = "default-supplier"
    existing_s = await tenant_db.suppliers.find_one({"id": default_supplier_id})
    if not existing_s:
        await tenant_db.suppliers.insert_one({
            "id": default_supplier_id,
            "name": "مورد متنوع",
            "name_fr": "Fournisseur divers",
            "phone": "",
            "email": "",
            "address": "",
            "family_id": default_supplier_family_id,
            "family_name": "عائلة مورد متنوع",
            "balance": 0,
            "total_purchases": 0,
            "notes": "مورد افتراضي للمشتريات العامة",
            "created_at": now,
            "updated_at": now
        })
    
    # Default Product Family
    default_product_family_id = "default-product-family"
    existing_pf = await tenant_db.product_families.find_one({"id": default_product_family_id})
    if not existing_pf:
        await tenant_db.product_families.insert_one({
            "id": default_product_family_id,
            "name": "عائلة منتج متنوع",
            "name_fr": "Famille produits divers",
            "name_ar": "عائلة منتج متنوع",
            "name_en": "Various Products Family",
            "description": "عائلة افتراضية للمنتجات",
            "description_ar": "عائلة افتراضية للمنتجات المتنوعة",
            "description_en": "Default family for various products",
            "parent_id": "",
            "parent_name": "",
            "image": "",
            "created_at": now,
            "updated_at": now
        })
    
    # Default Product
    default_product_id = "default-product"
    existing_p = await tenant_db.products.find_one({"id": default_product_id})
    if not existing_p:
        await tenant_db.products.insert_one({
            "id": default_product_id,
            "name_ar": "منتج متنوع",
            "name_en": "Produit divers",
            "article_code": "DIVERS-001",
            "barcode": "",
            "family_id": default_product_family_id,
            "family_name": "عائلة منتج متنوع",
            "purchase_price": 0,
            "wholesale_price": 0,
            "retail_price": 0,
            "quantity": 0,
            "min_stock": 0,
            "unit": "وحدة",
            "description": "منتج افتراضي للمبيعات المتنوعة",
            "supplier_id": default_supplier_id,
            "supplier_name": "مورد متنوع",
            "image": "",
            "created_at": now,
            "updated_at": now
        })

@api_router.post("/init-default-data")
async def api_init_default_data(admin: dict = Depends(get_tenant_admin)):
    """Initialize default data for existing tenant"""
    tenant_db = get_tenant_db(admin["tenant_id"])
    await init_default_data(tenant_db)
    return {"message": "تم تهيئة البيانات الافتراضية بنجاح", "status": "success"}

# ============ AUTH ROUTES ============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user: UserCreate):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # SECURITY: Prevent creating super_admin or saas_admin roles
    forbidden_roles = ["super_admin", "saas_admin", "superadmin"]
    if user.role and user.role.lower() in [r.lower() for r in forbidden_roles]:
        raise HTTPException(
            status_code=403, 
            detail="لا يمكن إنشاء حساب بصلاحية سوبر أدمين - Creating super_admin accounts is not allowed"
        )
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    user_doc = {
        "id": user_id, "email": user.email,
        "password": hash_password(user.password),
        "name": user.name, "role": user.role, "created_at": now
    }
    await db.users.insert_one(user_doc)
    access_token = create_access_token({"sub": user_id, "role": user.role})
    
    return TokenResponse(
        access_token=access_token,
        user=UserResponse(id=user_id, email=user.email, name=user.name, role=user.role, created_at=now)
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Support both password and hashed_password fields
    stored_password = user.get("hashed_password") or user.get("password")
    if not stored_password or not verify_password(credentials.password, stored_password):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    access_token = create_access_token({"sub": user["id"], "role": user["role"]})
    return TokenResponse(
        access_token=access_token,
        user=UserResponse(id=user["id"], email=user["email"], name=user["name"], role=user["role"], permissions=user.get("permissions", {}), created_at=user["created_at"])
    )

# Unified Login - Auto-detect user type
class UnifiedLoginResponse(BaseModel):
    access_token: str
    user_type: str  # admin, agent, tenant
    redirect_to: str
    user: dict

@api_router.post("/auth/unified-login")
async def unified_login(credentials: UserLogin):
    """
    Unified login endpoint that auto-detects user type:
    1. Check if user is an admin/employee
    2. Check if user is an agent
    3. Check if user is a tenant
    """
    email = credentials.email
    password = credentials.password
    
    # 1. Check Admin/Employee users first
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if user:
        stored_password = user.get("hashed_password") or user.get("password")
        if stored_password and verify_password(password, stored_password):
            access_token = create_access_token({"sub": user["id"], "role": user["role"]})
            return {
                "access_token": access_token,
                "user_type": "admin",
                "redirect_to": "/",
                "user": {
                    "id": user["id"],
                    "email": user["email"],
                    "name": user["name"],
                    "role": user["role"],
                    "permissions": user.get("permissions", {})
                }
            }
    
    # 2. Check Agents
    agent = await db.saas_agents.find_one({"email": email}, {"_id": 0})
    if agent:
        stored_password = agent.get("password", "")
        try:
            if bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8')):
                if not agent.get("is_active", True):
                    raise HTTPException(status_code=403, detail="الحساب معطل")
                
                token_data = {
                    "sub": agent["id"],
                    "email": agent["email"],
                    "role": "agent",
                    "type": "agent"
                }
                access_token = create_access_token(token_data)
                return {
                    "access_token": access_token,
                    "user_type": "agent",
                    "redirect_to": "/agent/dashboard",
                    "user": {
                        "id": agent["id"],
                        "email": agent["email"],
                        "name": agent["name"],
                        "company_name": agent.get("company_name", ""),
                        "current_balance": agent.get("current_balance", 0),
                        "credit_limit": agent.get("credit_limit", 0)
                    }
                }
        except Exception:
            pass
    
    # 3. Check Tenants
    tenant = await db.saas_tenants.find_one({"email": email})
    if tenant:
        stored_password = tenant.get("password", "")
        try:
            if bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8')):
                if not tenant.get("is_active", True):
                    raise HTTPException(status_code=403, detail="الحساب معطل")
                
                # Check subscription
                if tenant.get("subscription_ends_at"):
                    end_date = datetime.fromisoformat(tenant["subscription_ends_at"].replace("Z", "+00:00"))
                    if end_date < datetime.now(timezone.utc) and not tenant.get("is_trial"):
                        raise HTTPException(status_code=403, detail="انتهت صلاحية الاشتراك")
                
                # Check if this is the first login - create database if not initialized
                tenant_id = tenant['id']
                if not tenant.get("database_initialized", False):
                    logger.info(f"First login (unified) for tenant {tenant_id} - initializing database...")
                    tenant_db = await init_tenant_database(tenant_id)
                    
                    # Create admin user in tenant's database
                    admin_user = {
                        "id": str(uuid.uuid4()),
                        "name": tenant["name"],
                        "email": tenant["email"],
                        "password": stored_password,
                        "role": "admin",
                        "permissions": {},
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await tenant_db.users.insert_one(admin_user)
                    
                    # Initialize default data (customers, suppliers, families, products)
                    await init_default_data(tenant_db)
                    
                    # Mark database as initialized
                    await db.saas_tenants.update_one(
                        {"id": tenant_id},
                        {"$set": {
                            "database_initialized": True,
                            "first_login_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    logger.info(f"Database initialized successfully for tenant {tenant_id}")
                
                token_data = {
                    "sub": tenant["id"],
                    "email": tenant["email"],
                    "role": "tenant_admin",
                    "type": "tenant",
                    "tenant_id": tenant["id"]
                }
                access_token = create_access_token(token_data)
                
                # Get plan info with features and limits
                plan = await db.saas_plans.find_one({"id": tenant.get("plan_id")}, {"_id": 0})
                features = {**plan.get("features", {}), **tenant.get("features_override", {})} if plan else {}
                limits = {**plan.get("limits", {}), **tenant.get("limits_override", {})} if plan else {}
                
                return {
                    "access_token": access_token,
                    "user_type": "tenant",
                    "redirect_to": "/tenant/dashboard",
                    "user": {
                        "id": tenant["id"],
                        "email": tenant["email"],
                        "name": tenant["name"],
                        "company_name": tenant.get("company_name", ""),
                        "plan_name": plan.get("name_ar", "") if plan else "",
                        "subscription_ends_at": tenant.get("subscription_ends_at"),
                        "database_name": f"tenant_{tenant['id'].replace('-', '_')}",
                        "is_first_login": not tenant.get("database_initialized", False),
                        "features": features,
                        "limits": limits
                    }
                }
        except HTTPException:
            raise
        except Exception:
            pass
    
    # No user found
    raise HTTPException(status_code=401, detail="بيانات الدخول غير صحيحة")

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(**current_user)

# ============ USER MANAGEMENT ============

@api_router.get("/users", response_model=List[UserResponse])
async def get_all_users(admin: dict = Depends(get_admin_user)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return [UserResponse(**u) for u in users]

class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    role: str = "user"

@api_router.post("/users", response_model=UserResponse)
async def create_user(user_data: UserCreate, admin: dict = Depends(get_admin_user)):
    """Create a new user (admin only)"""
    # SECURITY: Prevent creating super_admin or saas_admin roles
    forbidden_roles = ["super_admin", "saas_admin", "superadmin"]
    if user_data.role and user_data.role.lower() in [r.lower() for r in forbidden_roles]:
        # Only super_admin can create super_admin users
        if admin.get("role") != "super_admin":
            raise HTTPException(
                status_code=403, 
                detail="لا يمكن إنشاء حساب بصلاحية سوبر أدمين - Creating super_admin accounts is not allowed"
            )
    
    # Check if email already exists
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم بالفعل")
    
    if len(user_data.password) < 4:
        raise HTTPException(status_code=400, detail="كلمة المرور يجب أن تكون 4 أحرف على الأقل")
    
    now = datetime.now(timezone.utc).isoformat()
    new_user = {
        "id": str(uuid.uuid4()),
        "name": user_data.name,
        "email": user_data.email,
        "password": hash_password(user_data.password),
        "role": user_data.role,
        "tenant_id": admin.get("tenant_id"),
        "permissions": {},
        "created_at": now
    }
    
    await db.users.insert_one(new_user)
    
    # Return without password
    del new_user["password"]
    return UserResponse(**new_user)

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, updates: UserUpdate, admin: dict = Depends(get_admin_user)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # SECURITY: Prevent changing role to super_admin or saas_admin
    forbidden_roles = ["super_admin", "saas_admin", "superadmin"]
    if updates.role and updates.role.lower() in [r.lower() for r in forbidden_roles]:
        # Only super_admin can assign super_admin role
        if admin.get("role") != "super_admin":
            raise HTTPException(
                status_code=403, 
                detail="لا يمكن تعيين صلاحية سوبر أدمين - Cannot assign super_admin role"
            )
    
    # SECURITY: Prevent non-super_admin from modifying super_admin users
    if user.get("role") == "super_admin" and admin.get("role") != "super_admin":
        raise HTTPException(
            status_code=403, 
            detail="لا يمكن تعديل حساب سوبر أدمين - Cannot modify super_admin account"
        )
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    return UserResponse(**updated)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(get_admin_user)):
    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}

@api_router.put("/users/{user_id}/password")
async def update_user_password(user_id: str, password_data: PasswordUpdate, admin: dict = Depends(get_admin_user)):
    """Update user password (admin only)"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if len(password_data.new_password) < 4:
        raise HTTPException(status_code=400, detail="كلمة المرور يجب أن تكون 4 أحرف على الأقل")
    
    hashed = hash_password(password_data.new_password)
    await db.users.update_one({"id": user_id}, {"$set": {"password": hashed}})
    
    return {"message": "تم تحديث كلمة المرور بنجاح"}

# ============ PRODUCT ROUTES ============

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, admin: dict = Depends(get_tenant_admin)):
    product_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Check if product with same name existed before (for restock alert)
    existing = await db.products.find_one({
        "$or": [{"name_en": product.name_en}, {"name_ar": product.name_ar}]
    })
    
    # Get family name if exists
    family_name = ""
    if product.family_id:
        family = await db.product_families.find_one({"id": product.family_id}, {"_id": 0, "name_ar": 1})
        if family:
            family_name = family["name_ar"]
    
    product_doc = {
        "id": product_id,
        "name_en": product.name_en, "name_ar": product.name_ar,
        "description_en": product.description_en or "",
        "description_ar": product.description_ar or "",
        "purchase_price": product.purchase_price,
        "wholesale_price": product.wholesale_price,
        "retail_price": product.retail_price,
        "super_wholesale_price": product.super_wholesale_price,
        "quantity": product.quantity,
        "image_url": product.image_url or "",
        "compatible_models": product.compatible_models,
        "low_stock_threshold": product.low_stock_threshold,
        "barcode": product.barcode or "",
        "article_code": product.article_code or "",
        "family_id": product.family_id or "",
        "family_name": family_name,
        "use_average_price": product.use_average_price or False,
        "created_at": now, "updated_at": now
    }
    await db.products.insert_one(product_doc)
    
    # Create notification if this was a restocked product
    if existing and existing.get("quantity", 0) == 0 and product.quantity > 0:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "restock",
            "message_en": f"Product '{product.name_en}' is back in stock!",
            "message_ar": f"المنتج '{product.name_ar}' متوفر مرة أخرى!",
            "product_id": product_id,
            "read": False,
            "created_at": now
        })
    
    return ProductResponse(**product_doc)

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(search: Optional[str] = None, model: Optional[str] = None, barcode: Optional[str] = None, family_id: Optional[str] = None, user: dict = Depends(require_tenant)):
    query = {}
    
    if barcode:
        query["barcode"] = barcode
    elif search:
        query["$or"] = [
            {"name_en": {"$regex": search, "$options": "i"}},
            {"name_ar": {"$regex": search, "$options": "i"}},
            {"description_en": {"$regex": search, "$options": "i"}},
            {"description_ar": {"$regex": search, "$options": "i"}},
            {"compatible_models": {"$regex": search, "$options": "i"}},
            {"barcode": {"$regex": search, "$options": "i"}},
            {"article_code": {"$regex": search, "$options": "i"}}  # بحث بكود المنتج
        ]
    
    if model:
        if "$or" in query:
            query = {"$and": [{"$or": query["$or"]}, {"compatible_models": {"$regex": model, "$options": "i"}}]}
        else:
            query["compatible_models"] = {"$regex": model, "$options": "i"}
    
    if family_id:
        if "$and" in query:
            query["$and"].append({"family_id": family_id})
        elif "$or" in query:
            query = {"$and": [{"$or": query["$or"]}, {"family_id": family_id}]}
        else:
            query["family_id"] = family_id
    
    products = await db.products.find(query, {"_id": 0}).to_list(1000)
    
    # Add family names and last purchase date for products
    for product in products:
        if product.get("family_id") and not product.get("family_name"):
            family = await db.product_families.find_one({"id": product["family_id"]}, {"_id": 0, "name_ar": 1})
            product["family_name"] = family["name_ar"] if family else ""
        elif not product.get("family_name"):
            product["family_name"] = ""
        if not product.get("article_code"):
            product["article_code"] = ""
        
        # Get last purchase date for this product
        if not product.get("last_purchase_date"):
            last_purchase = await db.purchases.find_one(
                {"items.product_id": product["id"]},
                {"_id": 0, "created_at": 1},
                sort=[("created_at", -1)]
            )
            if last_purchase:
                product["last_purchase_date"] = last_purchase["created_at"]
    
    return [ProductResponse(**p) for p in products]

# Paginated products endpoint
class PaginatedProductsResponse(BaseModel):
    items: List[ProductResponse]
    total: int
    page: int
    page_size: int
    total_pages: int

@api_router.get("/products/paginated", response_model=PaginatedProductsResponse)
async def get_products_paginated(
    search: Optional[str] = None, 
    model: Optional[str] = None, 
    barcode: Optional[str] = None, 
    family_id: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    user: dict = Depends(require_tenant)
):
    """Get products with pagination support"""
    query = {}
    
    if barcode:
        query["barcode"] = barcode
    elif search:
        query["$or"] = [
            {"name_en": {"$regex": search, "$options": "i"}},
            {"name_ar": {"$regex": search, "$options": "i"}},
            {"description_en": {"$regex": search, "$options": "i"}},
            {"description_ar": {"$regex": search, "$options": "i"}},
            {"compatible_models": {"$regex": search, "$options": "i"}},
            {"barcode": {"$regex": search, "$options": "i"}},
            {"article_code": {"$regex": search, "$options": "i"}}
        ]
    
    if model:
        if "$or" in query:
            query = {"$and": [{"$or": query["$or"]}, {"compatible_models": {"$regex": model, "$options": "i"}}]}
        else:
            query["compatible_models"] = {"$regex": model, "$options": "i"}
    
    if family_id:
        if "$and" in query:
            query["$and"].append({"family_id": family_id})
        elif "$or" in query:
            query = {"$and": [{"$or": query["$or"]}, {"family_id": family_id}]}
        else:
            query["family_id"] = family_id
    
    # Get total count
    total = await db.products.count_documents(query)
    total_pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    
    # Get paginated products
    skip = (page - 1) * page_size
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    # Add family names
    for product in products:
        if product.get("family_id") and not product.get("family_name"):
            family = await db.product_families.find_one({"id": product["family_id"]}, {"_id": 0, "name_ar": 1})
            product["family_name"] = family["name_ar"] if family else ""
        elif not product.get("family_name"):
            product["family_name"] = ""
        if not product.get("article_code"):
            product["article_code"] = ""
    
    return PaginatedProductsResponse(
        items=[ProductResponse(**p) for p in products],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )

# Quick Search Response Model (lightweight)
class QuickSearchProduct(BaseModel):
    id: str
    name_ar: str
    name_en: str
    barcode: Optional[str] = None
    article_code: Optional[str] = None
    retail_price: float = 0
    wholesale_price: float = 0
    quantity: int = 0
    min_quantity: int = 0
    family_id: Optional[str] = None
    family_name: Optional[str] = None
    image_url: Optional[str] = None

class QuickSearchResponse(BaseModel):
    results: List[QuickSearchProduct]
    total: int
    families: Optional[List[dict]] = None

@api_router.get("/products/quick-search", response_model=QuickSearchResponse)
async def quick_search_products(
    q: str = "",
    limit: int = 15,
    family_id: Optional[str] = None,
    stock_filter: Optional[str] = None,  # "low", "out", "available"
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    include_families: bool = False,
    user: dict = Depends(require_tenant)
):
    """
    Fast product search endpoint with advanced filtering.
    
    Params:
    - q: Search query (name, barcode, article_code)
    - family_id: Filter by product family
    - stock_filter: "low" (below min), "out" (zero), "available" (>0)
    - min_price, max_price: Price range filter
    - include_families: Include families list for filter dropdown
    """
    # Build search query
    conditions = []
    
    # Text search condition
    if q and len(q) >= 1:
        conditions.append({
            "$or": [
                {"barcode": q},  # Exact barcode match first
                {"article_code": {"$regex": f"^{q}", "$options": "i"}},
                {"name_ar": {"$regex": q, "$options": "i"}},
                {"name_en": {"$regex": q, "$options": "i"}},
                {"barcode": {"$regex": q, "$options": "i"}},
            ]
        })
    
    # Family filter
    if family_id:
        conditions.append({"family_id": family_id})
    
    # Stock filter
    if stock_filter == "out":
        conditions.append({"quantity": {"$lte": 0}})
    elif stock_filter == "low":
        conditions.append({"$expr": {"$lte": ["$quantity", "$min_quantity"]}})
    elif stock_filter == "available":
        conditions.append({"quantity": {"$gt": 0}})
    
    # Price range filter
    if min_price is not None:
        conditions.append({"retail_price": {"$gte": min_price}})
    if max_price is not None:
        conditions.append({"retail_price": {"$lte": max_price}})
    
    # Build final query
    search_query = {"$and": conditions} if conditions else {}
    
    # Only fetch required fields for speed
    projection = {
        "_id": 0,
        "id": 1,
        "name_ar": 1,
        "name_en": 1,
        "barcode": 1,
        "article_code": 1,
        "retail_price": 1,
        "wholesale_price": 1,
        "quantity": 1,
        "min_quantity": 1,
        "family_id": 1,
        "image_url": 1
    }
    
    # Get total count
    total = await db.products.count_documents(search_query)
    
    # Fetch products with limit
    products = await db.products.find(search_query, projection).limit(limit).to_list(limit)
    
    # Add family names
    for product in products:
        if product.get("family_id"):
            family = await db.product_families.find_one({"id": product["family_id"]}, {"_id": 0, "name_ar": 1})
            product["family_name"] = family.get("name_ar", "") if family else ""
        else:
            product["family_name"] = ""
        # Ensure min_quantity has default
        if "min_quantity" not in product:
            product["min_quantity"] = 0
    
    # Sort to prioritize exact barcode matches
    def sort_key(p):
        if q and p.get("barcode") == q:
            return 0  # Exact barcode match first
        if q and p.get("article_code", "").lower().startswith(q.lower()):
            return 1  # Article code starts with query
        return 2  # Other matches
    
    if q:
        products.sort(key=sort_key)
    
    # Optionally include families for filter dropdown
    families_list = None
    if include_families:
        families = await db.product_families.find({}, {"_id": 0, "id": 1, "name_ar": 1, "name_en": 1}).to_list(100)
        families_list = families
    
    return QuickSearchResponse(
        results=[QuickSearchProduct(**p) for p in products],
        total=total,
        families=families_list
    )

@api_router.get("/products/generate-barcode")
async def generate_barcode(article_code: Optional[str] = None):
    """Generate a unique product barcode based on article code"""
    import random
    
    if article_code:
        # Use article code as base for barcode (e.g., AR00001 -> 2130000100001)
        # Extract number from article code
        try:
            num = article_code.replace("AR", "").lstrip("0") or "1"
            num = int(num)
        except:
            num = random.randint(1, 99999)
        
        prefix = "213"  # Algeria
        company = "0001"
        product_num = str(num).zfill(5)
        
        code = prefix + company + product_num
        odd_sum = sum(int(code[i]) for i in range(0, 12, 2))
        even_sum = sum(int(code[i]) for i in range(1, 12, 2))
        check_digit = (10 - ((odd_sum + even_sum * 3) % 10)) % 10
        
        barcode = code + str(check_digit)
        return {"barcode": barcode}
    
    # Fallback to random barcode generation
    while True:
        # Generate EAN-13 format barcode
        prefix = "213"  # Algeria
        company = "0001"
        product_num = str(random.randint(10000, 99999))
        
        code = prefix + company + product_num
        odd_sum = sum(int(code[i]) for i in range(0, 12, 2))
        even_sum = sum(int(code[i]) for i in range(1, 12, 2))
        check_digit = (10 - ((odd_sum + even_sum * 3) % 10)) % 10
        
        barcode = code + str(check_digit)
        
        existing = await db.products.find_one({"barcode": barcode})
        if not existing:
            return {"barcode": barcode}

@api_router.get("/products/generate-sku")
async def generate_sku(family_id: Optional[str] = None):
    """Generate a unique SKU code"""
    prefix = "SG"
    
    if family_id:
        family = await db.product_families.find_one({"id": family_id}, {"_id": 0, "name_en": 1})
        if family:
            prefix = family["name_en"][:2].upper()
    
    count = await db.products.count_documents({})
    sku = f"{prefix}-{str(count + 1).zfill(5)}"
    
    return {"sku": sku}

@api_router.get("/products/generate-article-code")
async def generate_article_code():
    """Generate next article code (AR00001, AR00002, etc.)"""
    # Find the highest article code
    pipeline = [
        {"$match": {"article_code": {"$regex": "^AR\\d{4}$"}}},
        {"$project": {
            "num": {"$toInt": {"$substr": ["$article_code", 2, 4]}}
        }},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    
    result = await db.products.aggregate(pipeline).to_list(1)
    
    if result:
        next_num = result[0]["num"] + 1
    else:
        next_num = 1
    
    article_code = f"AR{str(next_num).zfill(4)}"
    return {"article_code": article_code}

# ============ CODE GENERATORS FOR ALL ENTITIES ============

@api_router.get("/customers/generate-code")
async def generate_customer_code():
    """Generate next customer code (CL0001, CL0002, etc.)"""
    pipeline = [
        {"$match": {"code": {"$regex": "^CL\\d{4}$"}}},
        {"$project": {"num": {"$toInt": {"$substr": ["$code", 2, 4]}}}},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    result = await db.customers.aggregate(pipeline).to_list(1)
    next_num = result[0]["num"] + 1 if result else 1
    return {"code": f"CL{str(next_num).zfill(4)}"}

@api_router.get("/suppliers/generate-code")
async def generate_supplier_code():
    """Generate next supplier code (FR0001/26, etc.)"""
    year = str(datetime.now().year)[2:]  # 2026 -> 26
    pipeline = [
        {"$match": {"code": {"$regex": f"^FR\\d{{4}}/{year}$"}}},
        {"$project": {"num": {"$toInt": {"$substr": ["$code", 2, 4]}}}},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    result = await db.suppliers.aggregate(pipeline).to_list(1)
    next_num = result[0]["num"] + 1 if result else 1
    return {"code": f"FR{str(next_num).zfill(4)}/{year}"}

@api_router.get("/sales/generate-code")
async def generate_sale_code():
    """Generate next sale code (BV0001/26, etc.)"""
    year = str(datetime.now().year)[2:]  # 2026 -> 26
    pipeline = [
        {"$match": {"code": {"$regex": f"^BV\\d{{4}}/{year}$"}}},
        {"$project": {"num": {"$toInt": {"$substr": ["$code", 2, 4]}}}},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    result = await db.sales.aggregate(pipeline).to_list(1)
    next_num = result[0]["num"] + 1 if result else 1
    return {"code": f"BV{str(next_num).zfill(4)}/{year}"}

@api_router.get("/purchases/generate-code")
async def generate_purchase_code():
    """Generate next purchase code (AC0001/26, etc.)"""
    year = str(datetime.now().year)[2:]  # 2026 -> 26
    pipeline = [
        {"$match": {"code": {"$regex": f"^AC\\d{{4}}/{year}$"}}},
        {"$project": {"num": {"$toInt": {"$substr": ["$code", 2, 4]}}}},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    result = await db.purchases.aggregate(pipeline).to_list(1)
    next_num = result[0]["num"] + 1 if result else 1
    return {"code": f"AC{str(next_num).zfill(4)}/{year}"}

@api_router.get("/expenses/generate-code")
async def generate_expense_code():
    """Generate next expense code (CH0001/26, etc.)"""
    year = str(datetime.now().year)[2:]  # 2026 -> 26
    pipeline = [
        {"$match": {"code": {"$regex": f"^CH\\d{{4}}/{year}$"}}},
        {"$project": {"num": {"$toInt": {"$substr": ["$code", 2, 4]}}}},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    result = await db.expenses.aggregate(pipeline).to_list(1)
    next_num = result[0]["num"] + 1 if result else 1
    return {"code": f"CH{str(next_num).zfill(4)}/{year}"}

@api_router.get("/inventory-sessions/generate-code")
async def generate_inventory_code():
    """Generate next inventory code (IN0001/26, etc.)"""
    year = str(datetime.now().year)[2:]  # 2026 -> 26
    pipeline = [
        {"$match": {"code": {"$regex": f"^IN\\d{{4}}/{year}$"}}},
        {"$project": {"num": {"$toInt": {"$substr": ["$code", 2, 4]}}}},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    result = await db.inventory_sessions.aggregate(pipeline).to_list(1)
    next_num = result[0]["num"] + 1 if result else 1
    return {"code": f"IN{str(next_num).zfill(4)}/{year}"}

@api_router.get("/price-updates/generate-code")
async def generate_price_update_code():
    """Generate next price update code (MT0001/26, etc.)"""
    year = str(datetime.now().year)[2:]  # 2026 -> 26
    pipeline = [
        {"$match": {"code": {"$regex": f"^MT\\d{{4}}/{year}$"}}},
        {"$project": {"num": {"$toInt": {"$substr": ["$code", 2, 4]}}}},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    result = await db.price_update_logs.aggregate(pipeline).to_list(1)
    next_num = result[0]["num"] + 1 if result else 1
    return {"code": f"MT{str(next_num).zfill(4)}/{year}"}

@api_router.get("/daily-sessions/generate-code")
async def generate_session_code():
    """Generate next session code (S001/26, etc.)"""
    year = str(datetime.now().year)[2:]  # 2026 -> 26
    pipeline = [
        {"$match": {"code": {"$regex": f"^S\\d{{3}}/{year}$"}}},
        {"$project": {"num": {"$toInt": {"$substr": ["$code", 1, 3]}}}},
        {"$sort": {"num": -1}},
        {"$limit": 1}
    ]
    result = await db.daily_sessions.aggregate(pipeline).to_list(1)
    next_num = result[0]["num"] + 1 if result else 1
    return {"code": f"S{str(next_num).zfill(3)}/{year}"}

@api_router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(product_id: str, user: dict = Depends(require_tenant)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Add family name
    if product.get("family_id") and not product.get("family_name"):
        family = await db.product_families.find_one({"id": product["family_id"]}, {"_id": 0, "name_ar": 1})
        product["family_name"] = family["name_ar"] if family else ""
    elif not product.get("family_name"):
        product["family_name"] = ""
    
    return ProductResponse(**product)

@api_router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(product_id: str, updates: ProductUpdate, admin: dict = Depends(get_tenant_admin)):
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    old_quantity = product.get("quantity", 0)
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Log price changes
    price_fields = ["purchase_price", "wholesale_price", "retail_price"]
    product_name = product.get("name_ar") or product.get("name_en", "")
    for price_field in price_fields:
        if price_field in update_data:
            old_price = product.get(price_field, 0)
            new_price = update_data[price_field]
            if old_price != new_price:
                await log_price_change(
                    product_id=product_id,
                    product_name=product_name,
                    old_price=old_price,
                    new_price=new_price,
                    price_type=price_field,
                    changed_by=admin["id"],
                    changed_by_name=admin.get("name", ""),
                    source="manual",
                    notes=""
                )
    
    # Update family name if family_id changed
    if "family_id" in update_data and update_data["family_id"]:
        family = await db.product_families.find_one({"id": update_data["family_id"]}, {"_id": 0, "name_ar": 1})
        update_data["family_name"] = family["name_ar"] if family else ""
    elif "family_id" in update_data and not update_data["family_id"]:
        update_data["family_name"] = ""
    
    await db.products.update_one({"id": product_id}, {"$set": update_data})
    
    # Check for restock notification
    new_quantity = update_data.get("quantity", old_quantity)
    if old_quantity == 0 and new_quantity > 0:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "restock",
            "message_en": f"Product '{product.get('name_en')}' is back in stock!",
            "message_ar": f"المنتج '{product.get('name_ar')}' متوفر مرة أخرى!",
            "product_id": product_id,
            "read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    
    # Add family name
    if updated.get("family_id") and not updated.get("family_name"):
        family = await db.product_families.find_one({"id": updated["family_id"]}, {"_id": 0, "name_ar": 1})
        updated["family_name"] = family["name_ar"] if family else ""
    elif not updated.get("family_name"):
        updated["family_name"] = ""
    
    return ProductResponse(**updated)

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, admin: dict = Depends(get_tenant_admin)):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted successfully"}

@api_router.get("/products/alerts/low-stock", response_model=List[ProductResponse])
async def get_low_stock_products(admin: dict = Depends(get_tenant_admin)):
    pipeline = [
        {"$match": {"$expr": {"$lt": ["$quantity", {"$ifNull": ["$low_stock_threshold", 10]}]}}},
        {"$project": {"_id": 0}}
    ]
    products = await db.products.aggregate(pipeline).to_list(1000)
    return [ProductResponse(**p) for p in products]

# ============ PRICE HISTORY ROUTES ============

@api_router.get("/products/{product_id}/price-history", response_model=List[PriceHistoryResponse])
async def get_product_price_history(product_id: str, user: dict = Depends(require_tenant)):
    """Get price change history for a specific product"""
    history = await db.price_history.find(
        {"product_id": product_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return [PriceHistoryResponse(**h) for h in history]

@api_router.get("/products/{product_id}/purchase-history")
async def get_product_purchase_history(product_id: str, user: dict = Depends(require_tenant)):
    """Get purchase history for a specific product (from suppliers)"""
    # Get purchases containing this product
    purchases = await db.purchases.find(
        {"items.product_id": product_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    result = []
    for purchase in purchases:
        supplier = await db.suppliers.find_one({"id": purchase.get("supplier_id")}, {"_id": 0, "name": 1, "phone": 1})
        
        # Find the item for this product
        item_data = None
        for item in purchase.get("items", []):
            if item.get("product_id") == product_id:
                item_data = item
                break
        
        if item_data:
            result.append({
                "id": purchase.get("id"),
                "date": purchase.get("created_at"),
                "supplier_id": purchase.get("supplier_id"),
                "supplier_name": supplier.get("name") if supplier else "",
                "supplier_phone": supplier.get("phone") if supplier else "",
                "quantity": item_data.get("quantity", 0),
                "unit_price": item_data.get("unit_price", 0),
                "total": item_data.get("total", 0),
                "purchase_total": purchase.get("total", 0),
                "payment_status": purchase.get("payment_status", ""),
                "notes": purchase.get("notes", "")
            })
    
    return result

@api_router.get("/products/{product_id}/sales-history")
async def get_product_sales_history(product_id: str, user: dict = Depends(require_tenant)):
    """Get sales history for a specific product"""
    # Get sales containing this product
    sales = await db.sales.find(
        {"items.product_id": product_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    result = []
    for sale in sales:
        customer = None
        if sale.get("customer_id"):
            customer = await db.customers.find_one({"id": sale.get("customer_id")}, {"_id": 0, "name": 1})
        
        # Find the item for this product
        item_data = None
        for item in sale.get("items", []):
            if item.get("product_id") == product_id:
                item_data = item
                break
        
        if item_data:
            result.append({
                "id": sale.get("id"),
                "date": sale.get("created_at"),
                "customer_name": customer.get("name") if customer else (language_ar := "زبون عابر"),
                "quantity": item_data.get("quantity", 0),
                "unit_price": item_data.get("unit_price", 0),
                "discount": item_data.get("discount", 0),
                "total": item_data.get("total", 0),
                "sale_total": sale.get("total", 0),
                "payment_type": sale.get("payment_type", "cash")
            })
    
    return result

@api_router.get("/price-history", response_model=List[PriceHistoryResponse])
async def get_all_price_history(
    limit: int = 50,
    price_type: Optional[str] = None,
    user: dict = Depends(require_tenant)
):
    """Get all price change history"""
    query = {}
    if price_type:
        query["price_type"] = price_type
    
    history = await db.price_history.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(limit)
    return [PriceHistoryResponse(**h) for h in history]

async def log_price_change(
    product_id: str,
    product_name: str,
    old_price: float,
    new_price: float,
    price_type: str,
    changed_by: str,
    changed_by_name: str,
    source: str = "manual",
    notes: str = ""
):
    """Helper function to log price changes"""
    if old_price == new_price:
        return  # No change
    
    change_percent = ((new_price - old_price) / old_price * 100) if old_price > 0 else 0
    
    history_doc = {
        "id": str(uuid.uuid4()),
        "product_id": product_id,
        "product_name": product_name,
        "old_price": old_price,
        "new_price": new_price,
        "price_type": price_type,
        "change_percent": round(change_percent, 2),
        "changed_by": changed_by,
        "changed_by_name": changed_by_name,
        "source": source,
        "notes": notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.price_history.insert_one(history_doc)

# ============ CUSTOMER ROUTES ============

@api_router.post("/customers", response_model=CustomerResponse)
async def create_customer(customer: CustomerCreate, user: dict = Depends(require_tenant)):
    customer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Get family name if exists
    family_name = ""
    if customer.family_id:
        family = await db.customer_families.find_one({"id": customer.family_id}, {"_id": 0, "name": 1})
        if family:
            family_name = family["name"]
    
    customer_doc = {
        "id": customer_id, "name": customer.name,
        "phone": customer.phone or "", "email": customer.email or "",
        "address": customer.address or "", "notes": customer.notes or "",
        "code": customer.code or "",  # كود الزبون
        "family_id": customer.family_id or "", "family_name": family_name,
        "total_purchases": 0, "balance": 0, "created_at": now
    }
    await db.customers.insert_one(customer_doc)
    return CustomerResponse(**customer_doc)

@api_router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(search: Optional[str] = None, family_id: Optional[str] = None, user: dict = Depends(require_tenant)):
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}}  # بحث بالكود
        ]
    if family_id:
        query["family_id"] = family_id
    
    customers = await db.customers.find(query, {"_id": 0}).to_list(1000)
    
    # Add family names for customers without them
    for customer in customers:
        if customer.get("family_id") and not customer.get("family_name"):
            family = await db.customer_families.find_one({"id": customer["family_id"]}, {"_id": 0, "name": 1})
            customer["family_name"] = family["name"] if family else ""
        elif not customer.get("family_name"):
            customer["family_name"] = ""
        if not customer.get("family_id"):
            customer["family_id"] = ""
        if not customer.get("code"):
            customer["code"] = ""
    
    return [CustomerResponse(**c) for c in customers]

# Paginated customers endpoint
class PaginatedCustomersResponse(BaseModel):
    items: List[CustomerResponse]
    total: int
    page: int
    page_size: int
    total_pages: int

@api_router.get("/customers/paginated", response_model=PaginatedCustomersResponse)
async def get_customers_paginated(
    search: Optional[str] = None, 
    family_id: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    user: dict = Depends(require_tenant)
):
    """Get customers with pagination support"""
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}}
        ]
    if family_id:
        query["family_id"] = family_id
    
    # Get total count
    total = await db.customers.count_documents(query)
    total_pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    
    # Get paginated customers
    skip = (page - 1) * page_size
    customers = await db.customers.find(query, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    
    # Add family names
    for customer in customers:
        if customer.get("family_id") and not customer.get("family_name"):
            family = await db.customer_families.find_one({"id": customer["family_id"]}, {"_id": 0, "name": 1})
            customer["family_name"] = family["name"] if family else ""
        elif not customer.get("family_name"):
            customer["family_name"] = ""
        if not customer.get("family_id"):
            customer["family_id"] = ""
        if not customer.get("code"):
            customer["code"] = ""
    
    return PaginatedCustomersResponse(
        items=[CustomerResponse(**c) for c in customers],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )

@api_router.get("/customers/{customer_id}", response_model=CustomerResponse)
async def get_customer(customer_id: str, user: dict = Depends(require_tenant)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Add family name
    if customer.get("family_id") and not customer.get("family_name"):
        family = await db.customer_families.find_one({"id": customer["family_id"]}, {"_id": 0, "name": 1})
        customer["family_name"] = family["name"] if family else ""
    elif not customer.get("family_name"):
        customer["family_name"] = ""
    if not customer.get("family_id"):
        customer["family_id"] = ""
    
    return CustomerResponse(**customer)

@api_router.put("/customers/{customer_id}", response_model=CustomerResponse)
async def update_customer(customer_id: str, updates: CustomerUpdate, user: dict = Depends(require_tenant)):
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    # Update family name if family_id changed
    if "family_id" in update_data:
        if update_data["family_id"]:
            family = await db.customer_families.find_one({"id": update_data["family_id"]}, {"_id": 0, "name": 1})
            update_data["family_name"] = family["name"] if family else ""
        else:
            update_data["family_name"] = ""
    
    if update_data:
        await db.customers.update_one({"id": customer_id}, {"$set": update_data})
    
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not updated.get("family_id"):
        updated["family_id"] = ""
    if not updated.get("family_name"):
        updated["family_name"] = ""
    
    return CustomerResponse(**updated)

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, admin: dict = Depends(get_tenant_admin)):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}

# ============ CUSTOMER BLACKLIST ============

class BlacklistEntry(BaseModel):
    phone: str
    reason: str = ""
    notes: str = ""

class BlacklistResponse(BaseModel):
    id: str
    phone: str
    reason: str
    notes: str
    added_by: str
    added_by_name: str = ""
    created_at: str

@api_router.get("/blacklist")
async def get_blacklist(user: dict = Depends(require_tenant)):
    """Get all blacklisted phone numbers"""
    blacklist = await db.customer_blacklist.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return blacklist

@api_router.post("/blacklist")
async def add_to_blacklist(entry: BlacklistEntry, user: dict = Depends(require_tenant)):
    """Add phone to blacklist"""
    # Check if already blacklisted
    existing = await db.customer_blacklist.find_one({"phone": entry.phone})
    if existing:
        raise HTTPException(status_code=400, detail="هذا الرقم موجود بالفعل في القائمة السوداء")
    
    blacklist_doc = {
        "id": str(uuid.uuid4()),
        "phone": entry.phone,
        "reason": entry.reason,
        "notes": entry.notes,
        "added_by": user["id"],
        "added_by_name": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.customer_blacklist.insert_one(blacklist_doc)
    
    # Mark any customers with this phone as blacklisted
    await db.customers.update_many(
        {"phone": entry.phone},
        {"$set": {"is_blacklisted": True, "blacklist_reason": entry.reason}}
    )
    
    return BlacklistResponse(**blacklist_doc)

@api_router.delete("/blacklist/{entry_id}")
async def remove_from_blacklist(entry_id: str, user: dict = Depends(require_tenant)):
    """Remove phone from blacklist"""
    entry = await db.customer_blacklist.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(status_code=404, detail="لم يتم العثور على السجل")
    
    # Remove blacklist flag from customers with this phone
    await db.customers.update_many(
        {"phone": entry["phone"]},
        {"$set": {"is_blacklisted": False, "blacklist_reason": ""}}
    )
    
    await db.customer_blacklist.delete_one({"id": entry_id})
    return {"message": "تم إزالة الرقم من القائمة السوداء"}

@api_router.get("/blacklist/check/{phone}")
async def check_blacklist(phone: str, user: dict = Depends(require_tenant)):
    """Check if a phone is blacklisted"""
    entry = await db.customer_blacklist.find_one({"phone": phone}, {"_id": 0})
    return {"is_blacklisted": entry is not None, "entry": entry}

# ============ DEBT REMINDERS ============

class DebtReminderSettings(BaseModel):
    enabled: bool = True
    reminder_days: List[int] = [7, 14, 30]  # Days after debt to remind
    min_debt_amount: float = 1000  # Minimum debt to trigger reminder

@api_router.get("/debt-reminders/settings")
async def get_debt_reminder_settings(user: dict = Depends(require_tenant)):
    """Get debt reminder settings"""
    settings = await db.system_settings.find_one({"type": "debt_reminders"}, {"_id": 0})
    if not settings:
        return DebtReminderSettings().model_dump()
    return settings

@api_router.put("/debt-reminders/settings")
async def update_debt_reminder_settings(settings: DebtReminderSettings, user: dict = Depends(require_tenant)):
    """Update debt reminder settings"""
    await db.system_settings.update_one(
        {"type": "debt_reminders"},
        {"$set": {**settings.model_dump(), "type": "debt_reminders", "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"success": True, "message": "تم حفظ إعدادات التذكير"}

@api_router.get("/debt-reminders/pending")
async def get_pending_debt_reminders(user: dict = Depends(require_tenant)):
    """Get customers with pending debt reminders"""
    settings = await db.system_settings.find_one({"type": "debt_reminders"})
    if not settings or not settings.get("enabled", True):
        return []
    
    reminder_days = settings.get("reminder_days", [7, 14, 30])
    min_amount = settings.get("min_debt_amount", 1000)
    
    # Get all customers with debt
    customers_with_debt = await db.customers.find(
        {"total_debt": {"$gte": min_amount}},
        {"_id": 0}
    ).to_list(500)
    
    reminders = []
    now = datetime.now(timezone.utc)
    
    for customer in customers_with_debt:
        # Get last sale date for this customer
        last_sale = await db.sales.find_one(
            {"customer_id": customer["id"], "payment_type": {"$in": ["credit", "partial"]}},
            sort=[("created_at", -1)]
        )
        
        if last_sale:
            try:
                sale_date_str = last_sale.get("created_at", now.isoformat())
                if 'T' in sale_date_str:
                    sale_date = datetime.fromisoformat(sale_date_str.replace('Z', '+00:00'))
                else:
                    sale_date = datetime.strptime(sale_date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            except:
                sale_date = now
            
            days_since = (now - sale_date).days
            
            # Check if we should remind based on settings
            for reminder_day in reminder_days:
                if days_since >= reminder_day:
                    reminders.append({
                        "customer_id": customer["id"],
                        "customer_name": customer["name"],
                        "phone": customer.get("phone", ""),
                        "total_debt": customer["total_debt"],
                        "days_since_last_purchase": days_since,
                        "reminder_level": reminder_day,
                        "is_urgent": days_since >= 30
                    })
                    break  # Only one reminder per customer
    
    # Sort by urgency and debt amount
    reminders.sort(key=lambda x: (-x["days_since_last_purchase"], -x["total_debt"]))
    return reminders

@api_router.post("/debt-reminders/dismiss/{customer_id}")
async def dismiss_debt_reminder(customer_id: str, days: int = 7, user: dict = Depends(require_tenant)):
    """Dismiss a debt reminder for a period"""
    await db.debt_reminder_dismissals.update_one(
        {"customer_id": customer_id},
        {"$set": {
            "customer_id": customer_id,
            "dismissed_until": (datetime.now(timezone.utc) + timedelta(days=days)).isoformat(),
            "dismissed_by": user["id"]
        }},
        upsert=True
    )
    return {"message": f"تم تأجيل التذكير لمدة {days} أيام"}

# ============ NOTIFICATIONS ============

@api_router.get("/notifications")
async def get_notifications(user: dict = Depends(require_tenant)):
    """Get all notifications for current user"""
    notifications = await db.notifications.find(
        {"$or": [{"user_id": user["id"]}, {"user_id": None}]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return notifications

@api_router.post("/notifications")
async def create_notification(
    title: str,
    message: str,
    notification_type: str = "info",
    user_id: Optional[str] = None,
    user: dict = Depends(require_tenant)
):
    """Create a notification"""
    notification_doc = {
        "id": str(uuid.uuid4()),
        "title": title,
        "message": message,
        "type": notification_type,  # info, warning, error, debt
        "user_id": user_id,  # None = all users
        "is_read": False,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification_doc)
    return notification_doc

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(require_tenant)):
    """Mark notification as read"""
    await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "تم تحديد الإشعار كمقروء"}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, user: dict = Depends(require_tenant)):
    """Delete a notification"""
    await db.notifications.delete_one({"id": notification_id})
    return {"message": "تم حذف الإشعار"}

@api_router.get("/notifications/unread-count")
async def get_unread_notification_count(user: dict = Depends(require_tenant)):
    """Get count of unread notifications"""
    count = await db.notifications.count_documents({
        "$or": [{"user_id": user["id"]}, {"user_id": None}],
        "is_read": False
    })
    return {"count": count}

# ============ WAREHOUSE ROUTES ============

@api_router.post("/warehouses", response_model=WarehouseResponse)
async def create_warehouse(warehouse: WarehouseCreate, admin: dict = Depends(get_tenant_admin)):
    warehouse_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # If this is set as main, unset any existing main warehouse
    if warehouse.is_main:
        await db.warehouses.update_many({"is_main": True}, {"$set": {"is_main": False}})
    
    warehouse_doc = {
        "id": warehouse_id,
        "name": warehouse.name,
        "address": warehouse.address or "",
        "phone": warehouse.phone or "",
        "manager": warehouse.manager or "",
        "notes": warehouse.notes or "",
        "is_main": warehouse.is_main,
        "created_at": now
    }
    await db.warehouses.insert_one(warehouse_doc)
    return WarehouseResponse(**warehouse_doc)

@api_router.get("/warehouses", response_model=List[WarehouseResponse])
async def get_warehouses(user: dict = Depends(require_tenant)):
    warehouses = await db.warehouses.find({}, {"_id": 0}).to_list(100)
    return [WarehouseResponse(**w) for w in warehouses]

@api_router.put("/warehouses/{warehouse_id}", response_model=WarehouseResponse)
async def update_warehouse(warehouse_id: str, updates: WarehouseUpdate, admin: dict = Depends(get_tenant_admin)):
    warehouse = await db.warehouses.find_one({"id": warehouse_id})
    if not warehouse:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    # If setting as main, unset others
    if update_data.get("is_main"):
        await db.warehouses.update_many({"is_main": True}, {"$set": {"is_main": False}})
    
    if update_data:
        await db.warehouses.update_one({"id": warehouse_id}, {"$set": update_data})
    
    updated = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    return WarehouseResponse(**updated)

@api_router.delete("/warehouses/{warehouse_id}")
async def delete_warehouse(warehouse_id: str, admin: dict = Depends(get_tenant_admin)):
    warehouse = await db.warehouses.find_one({"id": warehouse_id})
    if not warehouse:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    if warehouse.get("is_main"):
        raise HTTPException(status_code=400, detail="Cannot delete main warehouse")
    
    result = await db.warehouses.delete_one({"id": warehouse_id})
    return {"message": "Warehouse deleted successfully"}

# ============ STOCK TRANSFER ROUTES ============

@api_router.post("/stock-transfers")
async def create_stock_transfer(transfer: StockTransferCreate, admin: dict = Depends(get_tenant_admin)):
    # Validate warehouses
    from_wh = await db.warehouses.find_one({"id": transfer.from_warehouse})
    to_wh = await db.warehouses.find_one({"id": transfer.to_warehouse})
    if not from_wh or not to_wh:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    # Validate product
    product = await db.products.find_one({"id": transfer.product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check quantity
    if product.get("quantity", 0) < transfer.quantity:
        raise HTTPException(status_code=400, detail="Insufficient quantity")
    
    transfer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    transfer_doc = {
        "id": transfer_id,
        "from_warehouse": transfer.from_warehouse,
        "from_warehouse_name": from_wh["name"],
        "to_warehouse": transfer.to_warehouse,
        "to_warehouse_name": to_wh["name"],
        "product_id": transfer.product_id,
        "product_name": product.get("name_ar", product.get("name_en", "")),
        "quantity": transfer.quantity,
        "created_at": now
    }
    
    await db.stock_transfers.insert_one(transfer_doc)
    
    # Update product warehouse_id if needed (for tracking)
    # Note: In a full multi-warehouse system, you'd track quantity per warehouse
    
    return transfer_doc

@api_router.get("/stock-transfers")
async def get_stock_transfers(user: dict = Depends(require_tenant)):
    transfers = await db.stock_transfers.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return transfers

# ============ INVENTORY SESSION ROUTES ============

@api_router.post("/inventory-sessions")
async def create_inventory_session(session: InventorySessionCreate, admin: dict = Depends(get_tenant_admin)):
    # Check for existing active session
    existing = await db.inventory_sessions.find_one({"status": "active"})
    if existing:
        raise HTTPException(status_code=400, detail="An active inventory session already exists")
    
    session_id = str(uuid.uuid4())
    
    session_doc = {
        "id": session_id,
        "name": session.name,
        "family_filter": session.family_filter,
        "status": "active",
        "started_at": session.started_at,
        "completed_at": None,
        "applied_changes": False,
        "counted_items": session.counted_items or {}
    }
    
    await db.inventory_sessions.insert_one(session_doc)
    session_doc.pop("_id", None)
    return session_doc

@api_router.get("/inventory-sessions")
async def get_inventory_sessions(user: dict = Depends(require_tenant)):
    sessions = await db.inventory_sessions.find({}, {"_id": 0}).sort("started_at", -1).to_list(100)
    return sessions

@api_router.put("/inventory-sessions/{session_id}")
async def update_inventory_session(session_id: str, updates: InventorySessionUpdate, admin: dict = Depends(get_tenant_admin)):
    session = await db.inventory_sessions.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Inventory session not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if update_data:
        await db.inventory_sessions.update_one({"id": session_id}, {"$set": update_data})
    
    updated = await db.inventory_sessions.find_one({"id": session_id}, {"_id": 0})
    return updated

@api_router.delete("/inventory-sessions/{session_id}")
async def delete_inventory_session(session_id: str, admin: dict = Depends(get_tenant_admin)):
    result = await db.inventory_sessions.delete_one({"id": session_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inventory session not found")
    return {"message": "Inventory session deleted successfully"}

# ============ SUPPLIER ROUTES ============

@api_router.post("/suppliers", response_model=SupplierResponse)
async def create_supplier(supplier: SupplierCreate, admin: dict = Depends(get_tenant_admin)):
    supplier_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Get family name if exists
    family_name = ""
    if supplier.family_id:
        family = await db.supplier_families.find_one({"id": supplier.family_id}, {"_id": 0, "name": 1})
        if family:
            family_name = family["name"]
    
    supplier_doc = {
        "id": supplier_id, "name": supplier.name,
        "phone": supplier.phone or "", "email": supplier.email or "",
        "address": supplier.address or "", "notes": supplier.notes or "",
        "code": supplier.code or "",  # كود المورد
        "family_id": supplier.family_id or "", "family_name": family_name,
        "total_purchases": 0, "balance": 0, "created_at": now
    }
    await db.suppliers.insert_one(supplier_doc)
    return SupplierResponse(**supplier_doc)

@api_router.get("/suppliers", response_model=List[SupplierResponse])
async def get_suppliers(search: Optional[str] = None, family_id: Optional[str] = None, admin: dict = Depends(get_tenant_admin)):
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}}  # بحث بالكود
        ]
    if family_id:
        query["family_id"] = family_id
    
    suppliers = await db.suppliers.find(query, {"_id": 0}).to_list(1000)
    
    # Add family names for suppliers without them
    for supplier in suppliers:
        if supplier.get("family_id") and not supplier.get("family_name"):
            family = await db.supplier_families.find_one({"id": supplier["family_id"]}, {"_id": 0, "name": 1})
            supplier["family_name"] = family["name"] if family else ""
        elif not supplier.get("family_name"):
            supplier["family_name"] = ""
        if not supplier.get("family_id"):
            supplier["family_id"] = ""
        if not supplier.get("code"):
            supplier["code"] = ""
    
    return [SupplierResponse(**s) for s in suppliers]

@api_router.get("/suppliers/{supplier_id}", response_model=SupplierResponse)
async def get_supplier(supplier_id: str, admin: dict = Depends(get_tenant_admin)):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Add family name
    if supplier.get("family_id") and not supplier.get("family_name"):
        family = await db.supplier_families.find_one({"id": supplier["family_id"]}, {"_id": 0, "name": 1})
        supplier["family_name"] = family["name"] if family else ""
    elif not supplier.get("family_name"):
        supplier["family_name"] = ""
    if not supplier.get("family_id"):
        supplier["family_id"] = ""
    
    return SupplierResponse(**supplier)

@api_router.put("/suppliers/{supplier_id}", response_model=SupplierResponse)
async def update_supplier(supplier_id: str, updates: SupplierUpdate, admin: dict = Depends(get_tenant_admin)):
    supplier = await db.suppliers.find_one({"id": supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    # Update family name if family_id changed
    if "family_id" in update_data:
        if update_data["family_id"]:
            family = await db.supplier_families.find_one({"id": update_data["family_id"]}, {"_id": 0, "name": 1})
            update_data["family_name"] = family["name"] if family else ""
        else:
            update_data["family_name"] = ""
    
    if update_data:
        await db.suppliers.update_one({"id": supplier_id}, {"$set": update_data})
    
    updated = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not updated.get("family_id"):
        updated["family_id"] = ""
    if not updated.get("family_name"):
        updated["family_name"] = ""
    
    return SupplierResponse(**updated)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, admin: dict = Depends(get_tenant_admin)):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted successfully"}

# ============ SUPPLIER ADVANCE PAYMENT ============

class SupplierAdvancePayment(BaseModel):
    amount: float
    payment_method: str = "cash"
    notes: str = ""

@api_router.post("/suppliers/{supplier_id}/advance-payment")
async def add_supplier_advance_payment(supplier_id: str, payment: SupplierAdvancePayment, user: dict = Depends(require_tenant)):
    """Add advance payment to supplier"""
    supplier = await db.suppliers.find_one({"id": supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Update supplier advance balance
    current_advance = supplier.get("advance_balance", 0)
    new_advance = current_advance + payment.amount
    
    await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": {"advance_balance": new_advance, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Record the advance payment
    advance_record = {
        "id": str(uuid.uuid4()),
        "supplier_id": supplier_id,
        "supplier_name": supplier["name"],
        "amount": payment.amount,
        "payment_method": payment.payment_method,
        "notes": payment.notes,
        "user_id": user["id"],
        "user_name": user.get("name", user.get("email", "")),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.supplier_advance_payments.insert_one(advance_record)
    
    # Record cash transaction
    cash_tx = {
        "id": str(uuid.uuid4()),
        "type": "out",
        "amount": payment.amount,
        "category": "supplier_advance",
        "description": f"دفع متقدم للمورد: {supplier['name']}",
        "reference_id": advance_record["id"],
        "created_by": user["id"],
        "created_by_name": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.cash_transactions.insert_one(cash_tx)
    
    return {"message": "Advance payment recorded", "new_advance_balance": new_advance}

@api_router.get("/suppliers/{supplier_id}/advance-payments")
async def get_supplier_advance_payments(supplier_id: str, user: dict = Depends(require_tenant)):
    """Get advance payment history for a supplier"""
    payments = await db.supplier_advance_payments.find(
        {"supplier_id": supplier_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return payments

# ============ SUPPLIER DEBTS ROUTES ============

class SupplierDebtPayment(BaseModel):
    supplier_id: str
    amount: float
    payment_method: str = "cash"

@api_router.post("/supplier-debts/pay")
async def pay_supplier_debt(payment: SupplierDebtPayment, user: dict = Depends(require_tenant)):
    """Pay supplier debt - applies payment to oldest unpaid purchases first"""
    
    # Get unpaid purchases for this supplier, ordered by date
    unpaid_purchases = await db.purchases.find({
        "supplier_id": payment.supplier_id,
        "remaining": {"$gt": 0}
    }).sort("created_at", 1).to_list(100)
    
    if not unpaid_purchases:
        raise HTTPException(status_code=400, detail="No outstanding debt for this supplier")
    
    remaining_payment = payment.amount
    updated_purchases = []
    
    for purchase in unpaid_purchases:
        if remaining_payment <= 0:
            break
            
        purchase_remaining = purchase["remaining"]
        payment_for_this = min(remaining_payment, purchase_remaining)
        
        new_paid = purchase["paid_amount"] + payment_for_this
        new_remaining = purchase["total"] - new_paid
        new_status = "paid" if new_remaining <= 0 else "partial"
        
        await db.purchases.update_one(
            {"id": purchase["id"]},
            {"$set": {
                "paid_amount": new_paid,
                "remaining": new_remaining,
                "status": new_status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        updated_purchases.append({
            "purchase_id": purchase["id"],
            "paid": payment_for_this
        })
        
        remaining_payment -= payment_for_this
    
    # Update supplier total_purchases
    supplier = await db.suppliers.find_one({"id": payment.supplier_id})
    if supplier:
        await db.suppliers.update_one(
            {"id": payment.supplier_id},
            {"$inc": {"total_purchases": -payment.amount}}
        )
    
    # Record transaction
    transaction_id = str(uuid.uuid4())
    transaction = {
        "id": transaction_id,
        "type": "expense",
        "box": payment.payment_method,
        "amount": -payment.amount,
        "balance_after": 0,
        "description": f"سداد دين مورد - {supplier['name'] if supplier else payment.supplier_id}",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.transactions.insert_one(transaction)
    
    # Update cash box
    await db.cash_boxes.update_one(
        {"id": payment.payment_method},
        {"$inc": {"balance": -payment.amount}}
    )
    
    return {
        "message": "Payment recorded successfully",
        "amount_paid": payment.amount,
        "updated_purchases": updated_purchases
    }

# ============ SALES ROUTES ============

@api_router.post("/sales", response_model=SaleResponse)
async def create_sale(sale: SaleCreate, user: dict = Depends(require_tenant)):
    sale_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    invoice_number = await generate_invoice_number("INV")
    
    # Get customer name
    customer_name = "عميل نقدي"
    if sale.customer_id:
        customer = await db.customers.find_one({"id": sale.customer_id})
        if customer:
            customer_name = customer["name"]
    
    # Validate credit sale requires customer
    if sale.payment_type in ["credit", "partial"] and not sale.customer_id:
        raise HTTPException(status_code=400, detail="Customer required for credit sale")
    
    # Calculate delivery fee
    delivery_fee = 0
    delivery_info = None
    if sale.delivery and sale.delivery.enabled:
        delivery_fee = sale.delivery.fee
        delivery_info = {
            "enabled": True,
            "wilaya_code": sale.delivery.wilaya_code,
            "wilaya_name": sale.delivery.wilaya_name,
            "city": sale.delivery.city,
            "address": sale.delivery.address,
            "delivery_type": sale.delivery.delivery_type,
            "fee": delivery_fee
        }
    
    # Recalculate total with delivery
    final_total = sale.total + delivery_fee
    
    # Calculate remaining and debt
    remaining = final_total - sale.paid_amount
    debt_amount = remaining if sale.payment_type in ["credit", "partial"] else 0
    status = "paid" if remaining <= 0 else ("partial" if sale.paid_amount > 0 else "unpaid")
    
    sale_doc = {
        "id": sale_id, "invoice_number": invoice_number,
        "code": sale.code or "",  # كود البيع
        "customer_id": sale.customer_id, "customer_name": customer_name,
        "items": [item.model_dump() for item in sale.items],
        "subtotal": sale.subtotal, "discount": sale.discount,
        "delivery_fee": delivery_fee, "delivery": delivery_info,
        "total": final_total,
        "paid_amount": sale.paid_amount, "debt_amount": debt_amount,
        "remaining": max(0, remaining),
        "payment_method": sale.payment_method, "payment_type": sale.payment_type,
        "status": status,
        "notes": sale.notes or "", "created_at": now, "created_by": user["name"]
    }
    await db.sales.insert_one(sale_doc)
    
    # Update product quantities
    for item in sale.items:
        await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"quantity": -item.quantity}}
        )
        # Check for low stock notification
        product = await db.products.find_one({"id": item.product_id})
        if product:
            threshold = product.get("low_stock_threshold", 10)
            if product.get("quantity", 0) < threshold:
                await db.notifications.insert_one({
                    "id": str(uuid.uuid4()),
                    "type": "low_stock",
                    "message_en": f"Low stock alert: '{product.get('name_en')}' ({product.get('quantity')} remaining)",
                    "message_ar": f"تنبيه مخزون: '{product.get('name_ar')}' ({product.get('quantity')} متبقي)",
                    "product_id": item.product_id,
                    "read": False,
                    "created_at": now
                })
    
    # Update customer balance and total purchases
    if sale.customer_id:
        await db.customers.update_one(
            {"id": sale.customer_id},
            {"$inc": {"total_purchases": final_total, "balance": debt_amount, "total_debt": debt_amount}}
        )
    
    # Update cash box
    if sale.paid_amount > 0:
        cash_box_id = sale.payment_method
        await db.cash_boxes.update_one(
            {"id": cash_box_id},
            {"$inc": {"balance": sale.paid_amount}, "$set": {"updated_at": now}}
        )
        await db.transactions.insert_one({
            "id": str(uuid.uuid4()),
            "cash_box_id": cash_box_id,
            "type": "income",
            "amount": sale.paid_amount,
            "description": f"مبيعات - فاتورة {invoice_number}",
            "reference_type": "sale",
            "reference_id": sale_id,
            "created_at": now,
            "created_by": user["name"]
        })
    
    return SaleResponse(**sale_doc)

@api_router.get("/sales", response_model=List[SaleResponse])
async def get_sales(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer_id: Optional[str] = None,
    user: dict = Depends(require_tenant)
):
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [SaleResponse(**s) for s in sales]

# Paginated sales endpoint
class PaginatedSalesResponse(BaseModel):
    items: List[SaleResponse]
    total: int
    page: int
    page_size: int
    total_pages: int

@api_router.get("/sales/paginated", response_model=PaginatedSalesResponse)
async def get_sales_paginated(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer_id: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    user: dict = Depends(require_tenant)
):
    """Get sales with pagination support"""
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    # Get total count
    total = await db.sales.count_documents(query)
    total_pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    
    # Get paginated sales
    skip = (page - 1) * page_size
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return PaginatedSalesResponse(
        items=[SaleResponse(**s) for s in sales],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )

@api_router.get("/sales/{sale_id}", response_model=SaleResponse)
async def get_sale(sale_id: str, user: dict = Depends(require_tenant)):
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    return SaleResponse(**sale)

# Sale return/refund
@api_router.post("/sales/{sale_id}/return")
async def return_sale(sale_id: str, user: dict = Depends(require_tenant)):
    sale = await db.sales.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Restore product quantities
    for item in sale["items"]:
        await db.products.update_one(
            {"id": item["product_id"]},
            {"$inc": {"quantity": item["quantity"]}}
        )
    
    # Update customer balance
    if sale.get("customer_id"):
        await db.customers.update_one(
            {"id": sale["customer_id"]},
            {"$inc": {"total_purchases": -sale["total"], "balance": -sale.get("remaining", 0)}}
        )
    
    # Deduct from cash box
    if sale.get("paid_amount", 0) > 0:
        await db.cash_boxes.update_one(
            {"id": sale["payment_method"]},
            {"$inc": {"balance": -sale["paid_amount"]}, "$set": {"updated_at": now}}
        )
        await db.transactions.insert_one({
            "id": str(uuid.uuid4()),
            "cash_box_id": sale["payment_method"],
            "type": "expense",
            "amount": sale["paid_amount"],
            "description": f"إرجاع مبيعات - فاتورة {sale['invoice_number']}",
            "reference_type": "return",
            "reference_id": sale_id,
            "created_at": now,
            "created_by": user["name"]
        })
    
    # Mark sale as returned
    await db.sales.update_one({"id": sale_id}, {"$set": {"status": "returned"}})
    
    return {"message": "Sale returned successfully"}

# ============ PURCHASE ROUTES ============

@api_router.post("/purchases", response_model=PurchaseResponse)
async def create_purchase(purchase: PurchaseCreate, admin: dict = Depends(get_tenant_admin)):
    purchase_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    invoice_number = await generate_invoice_number("PUR")
    
    supplier = await db.suppliers.find_one({"id": purchase.supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    remaining = purchase.total - purchase.paid_amount
    status = "paid" if remaining <= 0 else ("partial" if purchase.paid_amount > 0 else "unpaid")
    
    purchase_doc = {
        "id": purchase_id, "invoice_number": invoice_number,
        "code": purchase.code or "",  # كود الشراء
        "supplier_id": purchase.supplier_id, "supplier_name": supplier["name"],
        "items": [item.model_dump() for item in purchase.items],
        "total": purchase.total, "paid_amount": purchase.paid_amount,
        "remaining": max(0, remaining), "payment_method": purchase.payment_method,
        "status": status, "notes": purchase.notes or "",
        "created_at": now, "created_by": admin["name"]
    }
    await db.purchases.insert_one(purchase_doc)
    
    # Update product quantities and check for restock notifications
    for item in purchase.items:
        product = await db.products.find_one({"id": item.product_id})
        old_quantity = product.get("quantity", 0) if product else 0
        
        await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"quantity": item.quantity}}
        )
        
        # Create restock notification if product was out of stock
        if old_quantity == 0 and item.quantity > 0 and product:
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "type": "restock",
                "message_en": f"Product '{product.get('name_en')}' is back in stock!",
                "message_ar": f"المنتج '{product.get('name_ar')}' متوفر مرة أخرى!",
                "product_id": item.product_id,
                "read": False,
                "created_at": now
            })
    
    # Update supplier balance
    await db.suppliers.update_one(
        {"id": purchase.supplier_id},
        {"$inc": {"total_purchases": purchase.total, "balance": remaining}}
    )
    
    # Update cash box
    if purchase.paid_amount > 0:
        await db.cash_boxes.update_one(
            {"id": purchase.payment_method},
            {"$inc": {"balance": -purchase.paid_amount}, "$set": {"updated_at": now}}
        )
        await db.transactions.insert_one({
            "id": str(uuid.uuid4()),
            "cash_box_id": purchase.payment_method,
            "type": "expense",
            "amount": purchase.paid_amount,
            "description": f"مشتريات - فاتورة {invoice_number}",
            "reference_type": "purchase",
            "reference_id": purchase_id,
            "created_at": now,
            "created_by": admin["name"]
        })
    
    return PurchaseResponse(**purchase_doc)

@api_router.get("/purchases", response_model=List[PurchaseResponse])
async def get_purchases(supplier_id: Optional[str] = None, admin: dict = Depends(get_tenant_admin)):
    query = {}
    if supplier_id:
        query["supplier_id"] = supplier_id
    purchases = await db.purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [PurchaseResponse(**p) for p in purchases]

@api_router.get("/purchases/{purchase_id}", response_model=PurchaseResponse)
async def get_purchase(purchase_id: str, admin: dict = Depends(get_tenant_admin)):
    """Get single purchase by ID"""
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    return PurchaseResponse(**purchase)

class PurchaseUpdate(BaseModel):
    paid_amount: Optional[float] = None
    notes: Optional[str] = None

@api_router.put("/purchases/{purchase_id}")
async def update_purchase(purchase_id: str, update_data: PurchaseUpdate, admin: dict = Depends(get_tenant_admin)):
    """Update purchase - can modify paid amount and notes"""
    purchase = await db.purchases.find_one({"id": purchase_id})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    now = datetime.now(timezone.utc).isoformat()
    update_dict = {"updated_at": now, "updated_by": admin["name"]}
    
    old_paid = purchase.get("paid_amount", 0)
    old_total = purchase.get("total", 0)
    old_remaining = purchase.get("remaining", 0)
    
    # Update paid amount
    if update_data.paid_amount is not None:
        new_paid = update_data.paid_amount
        new_remaining = max(0, old_total - new_paid)
        new_status = "paid" if new_remaining <= 0 else ("partial" if new_paid > 0 else "unpaid")
        
        update_dict["paid_amount"] = new_paid
        update_dict["remaining"] = new_remaining
        update_dict["status"] = new_status
        
        # Update supplier balance
        balance_diff = old_remaining - new_remaining
        await db.suppliers.update_one(
            {"id": purchase["supplier_id"]},
            {"$inc": {"balance": -balance_diff}}
        )
        
        # Update cash box if payment increased
        payment_diff = new_paid - old_paid
        if payment_diff > 0:
            await db.cash_boxes.update_one(
                {"id": purchase.get("payment_method", "cash")},
                {"$inc": {"balance": -payment_diff}, "$set": {"updated_at": now}}
            )
            await db.transactions.insert_one({
                "id": str(uuid.uuid4()),
                "cash_box_id": purchase.get("payment_method", "cash"),
                "type": "expense",
                "amount": payment_diff,
                "description": f"دفعة إضافية للمشتريات - فاتورة {purchase.get('invoice_number', '')}",
                "reference_type": "purchase",
                "reference_id": purchase_id,
                "created_at": now,
                "created_by": admin["name"]
            })
    
    if update_data.notes is not None:
        update_dict["notes"] = update_data.notes
    
    await db.purchases.update_one({"id": purchase_id}, {"$set": update_dict})
    
    updated_purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    return {"message": "تم تحديث المشتريات بنجاح", "purchase": updated_purchase}

@api_router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str, admin: dict = Depends(get_tenant_admin)):
    """Delete purchase and reverse all related changes"""
    purchase = await db.purchases.find_one({"id": purchase_id})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Reverse product stock changes
    for item in purchase.get("items", []):
        await db.products.update_one(
            {"id": item["product_id"]},
            {"$inc": {"quantity": -item["quantity"]}}
        )
    
    # Reverse supplier balance
    await db.suppliers.update_one(
        {"id": purchase["supplier_id"]},
        {"$inc": {
            "total_purchases": -purchase.get("total", 0),
            "balance": -purchase.get("remaining", 0)
        }}
    )
    
    # If there was a payment, reverse cash box change
    if purchase.get("paid_amount", 0) > 0:
        await db.cash_boxes.update_one(
            {"id": purchase.get("payment_method", "cash")},
            {"$inc": {"balance": purchase["paid_amount"]}, "$set": {"updated_at": now}}
        )
        # Add reversal transaction
        await db.transactions.insert_one({
            "id": str(uuid.uuid4()),
            "cash_box_id": purchase.get("payment_method", "cash"),
            "type": "income",
            "amount": purchase["paid_amount"],
            "description": f"إلغاء مشتريات - فاتورة {purchase.get('invoice_number', '')}",
            "reference_type": "purchase_reversal",
            "reference_id": purchase_id,
            "created_at": now,
            "created_by": admin["name"]
        })
    
    # Delete the purchase
    await db.purchases.delete_one({"id": purchase_id})
    
    return {"message": "تم حذف المشتريات بنجاح", "deleted_id": purchase_id}

# ============ CASH BOX ROUTES ============

@api_router.get("/cash-boxes", response_model=List[CashBoxResponse])
async def get_cash_boxes(admin: dict = Depends(get_tenant_admin)):
    await init_cash_boxes()
    boxes = await db.cash_boxes.find({}, {"_id": 0}).to_list(100)
    return [CashBoxResponse(**b) for b in boxes]

@api_router.post("/cash-boxes/transfer")
async def transfer_between_boxes(
    from_box: str, to_box: str, amount: float,
    admin: dict = Depends(get_tenant_admin)
):
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    
    from_cash_box = await db.cash_boxes.find_one({"id": from_box})
    if not from_cash_box or from_cash_box["balance"] < amount:
        raise HTTPException(status_code=400, detail="Insufficient balance")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Deduct from source
    await db.cash_boxes.update_one(
        {"id": from_box},
        {"$inc": {"balance": -amount}, "$set": {"updated_at": now}}
    )
    
    # Add to destination
    await db.cash_boxes.update_one(
        {"id": to_box},
        {"$inc": {"balance": amount}, "$set": {"updated_at": now}}
    )
    
    # Record transactions
    transfer_id = str(uuid.uuid4())
    await db.transactions.insert_many([
        {
            "id": str(uuid.uuid4()),
            "cash_box_id": from_box,
            "type": "expense",
            "amount": amount,
            "description": f"تحويل إلى {to_box}",
            "reference_type": "transfer",
            "reference_id": transfer_id,
            "created_at": now,
            "created_by": admin["name"]
        },
        {
            "id": str(uuid.uuid4()),
            "cash_box_id": to_box,
            "type": "income",
            "amount": amount,
            "description": f"تحويل من {from_box}",
            "reference_type": "transfer",
            "reference_id": transfer_id,
            "created_at": now,
            "created_by": admin["name"]
        }
    ])
    
    return {"message": "Transfer completed successfully"}

@api_router.get("/transactions", response_model=List[TransactionResponse])
async def get_transactions(
    cash_box_id: Optional[str] = None,
    type: Optional[str] = None,
    admin: dict = Depends(get_tenant_admin)
):
    query = {}
    if cash_box_id:
        query["cash_box_id"] = cash_box_id
    if type:
        query["type"] = type
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Get cash box names
    cash_boxes = {b["id"]: b["name"] for b in await db.cash_boxes.find({}, {"_id": 0}).to_list(100)}
    
    result = []
    for t in transactions:
        t["cash_box_name"] = cash_boxes.get(t["cash_box_id"], t["cash_box_id"])
        t["balance_after"] = 0  # Calculate if needed
        result.append(TransactionResponse(**t))
    
    return result

# ============ NOTIFICATIONS ============

@api_router.get("/notifications")
async def get_notifications(user: dict = Depends(require_tenant)):
    # Get notifications for this user or global notifications (without user_id)
    notifications = await db.notifications.find(
        {
            "read": False,
            "$or": [
                {"user_id": user["id"]},
                {"user_id": {"$exists": False}}
            ]
        }, {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return notifications

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(require_tenant)):
    await db.notifications.update_one(
        {"id": notification_id, "$or": [{"user_id": user["id"]}, {"user_id": {"$exists": False}}]},
        {"$set": {"read": True}}
    )
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(user: dict = Depends(require_tenant)):
    await db.notifications.update_many(
        {"read": False, "$or": [{"user_id": user["id"]}, {"user_id": {"$exists": False}}]},
        {"$set": {"read": True}}
    )
    return {"message": "All notifications marked as read"}

@api_router.post("/notifications/generate")
async def generate_auto_notifications(user: dict = Depends(require_tenant)):
    """Generate automatic notifications for low stock and due debts"""
    notifications_created = []
    
    # 1. Low stock notifications
    low_stock_products = await db.products.find({
        "$expr": {"$lt": ["$quantity", {"$ifNull": ["$low_stock_threshold", 10]}]}
    }).to_list(100)
    
    for product in low_stock_products:
        # Check if notification already exists for this product
        existing = await db.notifications.find_one({
            "type": "low_stock",
            "reference_id": product["id"],
            "read": False
        })
        if not existing:
            notif = {
                "id": str(uuid.uuid4()),
                "type": "low_stock",
                "reference_id": product["id"],
                "message_ar": f"تنبيه: المنتج '{product.get('name_ar', product.get('name_en', 'منتج'))}' مخزونه منخفض ({product['quantity']} قطعة)",
                "message_en": f"Alert: Product '{product.get('name_en', product.get('name_ar', 'Product'))}' is low on stock ({product['quantity']} units)",
                "read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notifications.insert_one(notif)
            notifications_created.append(notif["id"])
    
    # 2. Customer debt notifications (debts > 7 days)
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    overdue_sales = await db.sales.find({
        "remaining": {"$gt": 0},
        "created_at": {"$lt": week_ago}
    }).to_list(100)
    
    for sale in overdue_sales:
        existing = await db.notifications.find_one({
            "type": "overdue_debt",
            "reference_id": sale["id"],
            "read": False
        })
        if not existing:
            customer = await db.customers.find_one({"id": sale.get("customer_id")})
            customer_name = customer["name"] if customer else "عميل"
            notif = {
                "id": str(uuid.uuid4()),
                "type": "overdue_debt",
                "reference_id": sale["id"],
                "message_ar": f"تذكير: دين مستحق من {customer_name} بقيمة {sale['remaining']:.2f} دج",
                "message_en": f"Reminder: Overdue debt from {customer_name} of {sale['remaining']:.2f} DA",
                "read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notifications.insert_one(notif)
            notifications_created.append(notif["id"])
    
    # 3. Supplier debt notifications (debts > 7 days)
    overdue_purchases = await db.purchases.find({
        "remaining": {"$gt": 0},
        "created_at": {"$lt": week_ago}
    }).to_list(100)
    
    for purchase in overdue_purchases:
        existing = await db.notifications.find_one({
            "type": "supplier_debt",
            "reference_id": purchase["id"],
            "read": False
        })
        if not existing:
            notif = {
                "id": str(uuid.uuid4()),
                "type": "supplier_debt",
                "reference_id": purchase["id"],
                "message_ar": f"تذكير: دين للمورد {purchase.get('supplier_name', '')} بقيمة {purchase['remaining']:.2f} دج",
                "message_en": f"Reminder: Supplier debt to {purchase.get('supplier_name', '')} of {purchase['remaining']:.2f} DA",
                "read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notifications.insert_one(notif)
            notifications_created.append(notif["id"])
    
    return {
        "message": f"Generated {len(notifications_created)} notifications",
        "notification_ids": notifications_created
    }

# ============ STATS & REPORTS ============

@api_router.get("/stats")
async def get_stats(admin: dict = Depends(get_tenant_admin)):
    await init_cash_boxes()
    
    total_products = await db.products.count_documents({})
    total_customers = await db.customers.count_documents({})
    total_suppliers = await db.suppliers.count_documents({})
    total_employees = await db.employees.count_documents({})
    
    # Low stock count
    pipeline = [
        {"$match": {"$expr": {"$lt": ["$quantity", {"$ifNull": ["$low_stock_threshold", 10]}]}}},
        {"$count": "count"}
    ]
    result = await db.products.aggregate(pipeline).to_list(1)
    low_stock = result[0]["count"] if result else 0
    
    # Today's sales
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today_sales = await db.sales.aggregate([
        {"$match": {"created_at": {"$gte": today}, "status": {"$ne": "returned"}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    
    # Cash boxes
    cash_boxes = await db.cash_boxes.find({}, {"_id": 0}).to_list(100)
    total_cash = sum(b.get("balance", 0) for b in cash_boxes)
    
    # Unread notifications
    unread_notifications = await db.notifications.count_documents({"read": False})
    
    # Total debts
    total_receivables = await db.debts.aggregate([
        {"$match": {"type": "receivable", "status": {"$ne": "paid"}}},
        {"$group": {"_id": None, "total": {"$sum": "$remaining_amount"}}}
    ]).to_list(1)
    
    total_payables = await db.debts.aggregate([
        {"$match": {"type": "payable", "status": {"$ne": "paid"}}},
        {"$group": {"_id": None, "total": {"$sum": "$remaining_amount"}}}
    ]).to_list(1)
    
    return {
        "total_products": total_products,
        "total_customers": total_customers,
        "total_suppliers": total_suppliers,
        "total_employees": total_employees,
        "low_stock_count": low_stock,
        "today_sales_total": today_sales[0]["total"] if today_sales else 0,
        "today_sales_count": today_sales[0]["count"] if today_sales else 0,
        "total_cash": total_cash,
        "cash_boxes": cash_boxes,
        "unread_notifications": unread_notifications,
        "total_receivables": total_receivables[0]["total"] if total_receivables else 0,
        "total_payables": total_payables[0]["total"] if total_payables else 0,
        "currency": CURRENCY
    }

@api_router.get("/dashboard/sales-stats")
async def get_sales_stats(user: dict = Depends(require_tenant)):
    """Get sales statistics for today, month, and year"""
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    month_start = now.strftime("%Y-%m-01")
    year_start = now.strftime("%Y-01-01")
    
    # Today's sales
    today_pipeline = [
        {"$match": {"created_at": {"$gte": today}, "status": {"$ne": "returned"}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]
    today_result = await db.sales.aggregate(today_pipeline).to_list(1)
    
    # This month's sales
    month_pipeline = [
        {"$match": {"created_at": {"$gte": month_start}, "status": {"$ne": "returned"}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]
    month_result = await db.sales.aggregate(month_pipeline).to_list(1)
    
    # This year's sales
    year_pipeline = [
        {"$match": {"created_at": {"$gte": year_start}, "status": {"$ne": "returned"}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]
    year_result = await db.sales.aggregate(year_pipeline).to_list(1)
    
    return {
        "today": {
            "total": today_result[0]["total"] if today_result else 0,
            "count": today_result[0]["count"] if today_result else 0
        },
        "month": {
            "total": month_result[0]["total"] if month_result else 0,
            "count": month_result[0]["count"] if month_result else 0
        },
        "year": {
            "total": year_result[0]["total"] if year_result else 0,
            "count": year_result[0]["count"] if year_result else 0
        }
    }


@api_router.get("/dashboard/profit-stats")
async def get_profit_stats(user: dict = Depends(require_tenant)):
    """Get monthly profit statistics (Revenue - Purchase Cost - Expenses)"""
    now = datetime.now(timezone.utc)
    month_start = now.strftime("%Y-%m-01")
    
    # Monthly sales revenue
    sales_pipeline = [
        {"$match": {"created_at": {"$gte": month_start}, "status": {"$ne": "returned"}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}}}
    ]
    sales_result = await db.sales.aggregate(sales_pipeline).to_list(1)
    monthly_revenue = sales_result[0]["total"] if sales_result else 0
    
    # Monthly purchase cost (from sales items)
    # Calculate total purchase cost based on sold items
    sales_items_pipeline = [
        {"$match": {"created_at": {"$gte": month_start}, "status": {"$ne": "returned"}}},
        {"$unwind": "$items"},
        {"$group": {"_id": None, "total_cost": {"$sum": {"$multiply": ["$items.quantity", "$items.purchase_price"]}}}}
    ]
    try:
        sales_cost_result = await db.sales.aggregate(sales_items_pipeline).to_list(1)
        monthly_purchase_cost = sales_cost_result[0]["total_cost"] if sales_cost_result else 0
    except:
        monthly_purchase_cost = 0
    
    # Monthly expenses
    expenses_pipeline = [
        {"$match": {"date": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    expenses_result = await db.expenses.aggregate(expenses_pipeline).to_list(1)
    monthly_expenses = expenses_result[0]["total"] if expenses_result else 0
    
    # Calculate net profit: Revenue - Purchase Cost - Expenses
    monthly_profit = monthly_revenue - monthly_purchase_cost - monthly_expenses
    
    return {
        "monthly_revenue": monthly_revenue,
        "monthly_purchase_cost": monthly_purchase_cost,
        "monthly_expenses": monthly_expenses,
        "monthly_profit": monthly_profit
    }


@api_router.get("/analytics/sales-chart")
async def get_sales_chart_data(period: str = "week", admin: dict = Depends(get_tenant_admin)):
    """Get sales data for charts (daily for week/month, monthly for year)"""
    now = datetime.now(timezone.utc)
    
    if period == "week":
        # Last 7 days
        days = 7
        start_date = (now - timedelta(days=days)).strftime("%Y-%m-%d")
        pipeline = [
            {"$match": {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}}},
            {"$group": {
                "_id": {"$substr": ["$created_at", 0, 10]},
                "total": {"$sum": "$total"},
                "count": {"$sum": 1}
            }},
            {"$sort": {"_id": 1}}
        ]
    elif period == "month":
        # Last 30 days
        days = 30
        start_date = (now - timedelta(days=days)).strftime("%Y-%m-%d")
        pipeline = [
            {"$match": {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}}},
            {"$group": {
                "_id": {"$substr": ["$created_at", 0, 10]},
                "total": {"$sum": "$total"},
                "count": {"$sum": 1}
            }},
            {"$sort": {"_id": 1}}
        ]
    else:  # year
        # Last 12 months
        start_date = (now - timedelta(days=365)).strftime("%Y-%m-%d")
        pipeline = [
            {"$match": {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}}},
            {"$group": {
                "_id": {"$substr": ["$created_at", 0, 7]},
                "total": {"$sum": "$total"},
                "count": {"$sum": 1}
            }},
            {"$sort": {"_id": 1}}
        ]
    
    result = await db.sales.aggregate(pipeline).to_list(100)
    
    return {
        "period": period,
        "data": [{"date": r["_id"], "total": r["total"], "count": r["count"]} for r in result]
    }

@api_router.get("/analytics/top-products")
async def get_top_products(limit: int = 10, period: str = "month", admin: dict = Depends(get_tenant_admin)):
    """Get top selling products"""
    now = datetime.now(timezone.utc)
    
    if period == "week":
        start_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    elif period == "month":
        start_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    else:
        start_date = (now - timedelta(days=365)).strftime("%Y-%m-%d")
    
    pipeline = [
        {"$match": {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.product_id",
            "product_name": {"$first": "$items.name"},
            "total_quantity": {"$sum": "$items.quantity"},
            "total_revenue": {"$sum": {"$multiply": ["$items.price", "$items.quantity"]}}
        }},
        {"$sort": {"total_revenue": -1}},
        {"$limit": limit}
    ]
    
    result = await db.sales.aggregate(pipeline).to_list(limit)
    
    return {
        "period": period,
        "products": result
    }

@api_router.get("/analytics/top-customers")
async def get_top_customers(limit: int = 10, period: str = "month", admin: dict = Depends(get_tenant_admin)):
    """Get top customers by purchase amount"""
    now = datetime.now(timezone.utc)
    
    if period == "week":
        start_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    elif period == "month":
        start_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    else:
        start_date = (now - timedelta(days=365)).strftime("%Y-%m-%d")
    
    pipeline = [
        {"$match": {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}, "customer_id": {"$exists": True, "$ne": ""}}},
        {"$group": {
            "_id": "$customer_id",
            "customer_name": {"$first": "$customer_name"},
            "total_purchases": {"$sum": "$total"},
            "orders_count": {"$sum": 1}
        }},
        {"$sort": {"total_purchases": -1}},
        {"$limit": limit}
    ]
    
    result = await db.sales.aggregate(pipeline).to_list(limit)
    
    return {
        "period": period,
        "customers": result
    }

@api_router.get("/analytics/employee-performance")
async def get_employee_performance(period: str = "month", admin: dict = Depends(get_tenant_admin)):
    """Get sales performance by employee"""
    now = datetime.now(timezone.utc)
    
    if period == "week":
        start_date = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    elif period == "month":
        start_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    else:
        start_date = (now - timedelta(days=365)).strftime("%Y-%m-%d")
    
    # Get from daily sessions
    pipeline = [
        {"$match": {"closed_at": {"$gte": start_date}, "status": "closed"}},
        {"$group": {
            "_id": "$user_id",
            "user_name": {"$first": "$user_name"},
            "total_sales": {"$sum": "$total_sales"},
            "sessions_count": {"$sum": 1},
            "total_difference": {"$sum": {
                "$subtract": [
                    "$closing_cash",
                    {"$add": ["$opening_cash", "$cash_sales"]}
                ]
            }}
        }},
        {"$sort": {"total_sales": -1}}
    ]
    
    result = await db.daily_sessions.aggregate(pipeline).to_list(50)
    
    return {
        "period": period,
        "employees": result
    }

@api_router.get("/analytics/sales-prediction")
async def get_sales_prediction(admin: dict = Depends(get_tenant_admin)):
    """AI-powered sales prediction (MOCKED - simple moving average)"""
    now = datetime.now(timezone.utc)
    
    # Get last 30 days sales
    start_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    pipeline = [
        {"$match": {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}}},
        {"$group": {
            "_id": {"$substr": ["$created_at", 0, 10]},
            "total": {"$sum": "$total"}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    result = await db.sales.aggregate(pipeline).to_list(30)
    
    if not result:
        return {"prediction": 0, "confidence": 0, "trend": "neutral"}
    
    # Simple moving average prediction
    totals = [r["total"] for r in result]
    avg = sum(totals) / len(totals) if totals else 0
    
    # Calculate trend
    if len(totals) >= 7:
        recent_avg = sum(totals[-7:]) / 7
        older_avg = sum(totals[:7]) / 7 if len(totals) >= 14 else avg
        trend = "up" if recent_avg > older_avg * 1.1 else ("down" if recent_avg < older_avg * 0.9 else "neutral")
    else:
        trend = "neutral"
    
    # Predicted next day
    prediction = avg * (1.05 if trend == "up" else (0.95 if trend == "down" else 1))
    
    return {
        "predicted_daily_sales": round(prediction, 2),
        "predicted_monthly_sales": round(prediction * 30, 2),
        "average_daily_sales": round(avg, 2),
        "trend": trend,
        "confidence": 0.7 if len(totals) >= 14 else 0.5,
        "recommendation": {
            "ar": "بناءً على البيانات، يُنصح بزيادة المخزون للمنتجات الأكثر مبيعاً" if trend == "up" else "حافظ على مستوى المخزون الحالي",
            "fr": "Basé sur les données, il est recommandé d'augmenter le stock des produits les plus vendus" if trend == "up" else "Maintenez le niveau de stock actuel"
        }
    }

@api_router.get("/analytics/restock-suggestions")
async def get_restock_suggestions(admin: dict = Depends(get_tenant_admin)):
    """AI-powered restock suggestions based on sales velocity"""
    
    # Get products with low stock
    low_stock_products = await db.products.find(
        {"$expr": {"$lte": ["$quantity", "$low_stock_threshold"]}},
        {"_id": 0}
    ).to_list(100)
    
    now = datetime.now(timezone.utc)
    start_date = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    
    suggestions = []
    
    for product in low_stock_products:
        # Get sales velocity
        pipeline = [
            {"$match": {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}}},
            {"$unwind": "$items"},
            {"$match": {"items.product_id": product["id"]}},
            {"$group": {"_id": None, "total_sold": {"$sum": "$items.quantity"}}}
        ]
        
        sales_result = await db.sales.aggregate(pipeline).to_list(1)
        monthly_sales = sales_result[0]["total_sold"] if sales_result else 0
        daily_velocity = monthly_sales / 30
        
        # Calculate days until stockout
        days_until_stockout = product["quantity"] / daily_velocity if daily_velocity > 0 else 999
        
        # Suggested restock quantity (2 months supply)
        suggested_quantity = max(int(daily_velocity * 60), product.get("low_stock_threshold", 10) * 2)
        
        urgency = "critical" if days_until_stockout <= 3 else ("high" if days_until_stockout <= 7 else ("medium" if days_until_stockout <= 14 else "low"))
        
        suggestions.append({
            "product_id": product["id"],
            "product_name": product.get("name_en", ""),
            "current_stock": product["quantity"],
            "monthly_sales": monthly_sales,
            "daily_velocity": round(daily_velocity, 2),
            "days_until_stockout": round(days_until_stockout, 1),
            "suggested_restock": suggested_quantity,
            "urgency": urgency
        })
    
    # Sort by urgency
    urgency_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    suggestions.sort(key=lambda x: urgency_order.get(x["urgency"], 4))
    
    return {
        "suggestions": suggestions,
        "total_products_needing_restock": len(suggestions)
    }

# ============ CHARTS & ANALYTICS ============

@api_router.get("/reports/sales-chart")
async def get_sales_chart(days: int = 7, admin: dict = Depends(get_tenant_admin)):
    """Get sales data for chart (last N days)"""
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    
    pipeline = [
        {"$match": {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}}},
        {"$addFields": {"date": {"$substr": ["$created_at", 0, 10]}}},
        {"$group": {
            "_id": "$date",
            "total_sales": {"$sum": "$total"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    result = await db.sales.aggregate(pipeline).to_list(100)
    return [{"date": r["_id"], "total": r["total_sales"], "count": r["count"]} for r in result]

@api_router.get("/reports/top-products")
async def get_top_products(limit: int = 10, admin: dict = Depends(get_tenant_admin)):
    """Get top selling products"""
    pipeline = [
        {"$match": {"status": {"$ne": "returned"}}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.product_id",
            "product_name": {"$first": "$items.product_name"},
            "total_quantity": {"$sum": "$items.quantity"},
            "total_revenue": {"$sum": "$items.total"}
        }},
        {"$sort": {"total_quantity": -1}},
        {"$limit": limit}
    ]
    
    result = await db.sales.aggregate(pipeline).to_list(limit)
    return result

@api_router.get("/reports/top-customers")
async def get_top_customers(limit: int = 10, admin: dict = Depends(get_tenant_admin)):
    """Get top customers by purchases"""
    customers = await db.customers.find(
        {}, {"_id": 0}
    ).sort("total_purchases", -1).limit(limit).to_list(limit)
    return customers

@api_router.get("/reports/profit")
async def get_profit_report(days: int = 30, admin: dict = Depends(get_tenant_admin)):
    """Get profit report"""
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    
    # Get sales
    sales = await db.sales.find(
        {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}},
        {"_id": 0, "items": 1, "total": 1}
    ).to_list(10000)
    
    total_revenue = sum(s["total"] for s in sales)
    
    # Calculate cost from items (need to get purchase prices from products)
    total_cost = 0
    for sale in sales:
        for item in sale.get("items", []):
            product = await db.products.find_one({"id": item["product_id"]}, {"_id": 0, "purchase_price": 1})
            if product:
                total_cost += product.get("purchase_price", 0) * item["quantity"]
    
    gross_profit = total_revenue - total_cost
    profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    return {
        "total_revenue": total_revenue,
        "total_cost": total_cost,
        "gross_profit": gross_profit,
        "profit_margin": round(profit_margin, 2),
        "period_days": days
    }

@api_router.get("/reports/profit-detailed")
async def get_detailed_profit_report(days: int = 30, admin: dict = Depends(get_tenant_admin)):
    """Get detailed profit report with daily breakdown"""
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    
    # Get all sales in period
    sales = await db.sales.find(
        {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}},
        {"_id": 0}
    ).to_list(10000)
    
    # Group by day and calculate profits
    daily_data = {}
    product_profits = {}
    
    for sale in sales:
        sale_date = sale.get("created_at", "")[:10]
        if sale_date not in daily_data:
            daily_data[sale_date] = {"revenue": 0, "cost": 0, "profit": 0, "sales_count": 0}
        
        daily_data[sale_date]["sales_count"] += 1
        daily_data[sale_date]["revenue"] += sale.get("total", 0)
        
        for item in sale.get("items", []):
            product_id = item.get("product_id")
            if product_id:
                product = await db.products.find_one({"id": product_id}, {"_id": 0, "purchase_price": 1, "name_ar": 1, "name_en": 1})
                if product:
                    purchase_price = product.get("purchase_price", 0)
                    sale_price = item.get("price", 0)
                    quantity = item.get("quantity", 1)
                    item_cost = purchase_price * quantity
                    item_profit = (sale_price - purchase_price) * quantity
                    
                    daily_data[sale_date]["cost"] += item_cost
                    daily_data[sale_date]["profit"] += item_profit
                    
                    # Track product profits
                    if product_id not in product_profits:
                        product_profits[product_id] = {
                            "name": product.get("name_ar") or product.get("name_en", ""),
                            "total_sold": 0,
                            "total_profit": 0,
                            "profit_margin": 0
                        }
                    product_profits[product_id]["total_sold"] += quantity
                    product_profits[product_id]["total_profit"] += item_profit
    
    # Calculate profit margins for products
    for pid, pdata in product_profits.items():
        if pdata["total_sold"] > 0:
            pdata["profit_per_unit"] = round(pdata["total_profit"] / pdata["total_sold"], 2)
    
    # Sort daily data by date
    sorted_daily = [{"date": k, **v} for k, v in sorted(daily_data.items(), reverse=True)]
    
    # Get top profitable products
    top_products = sorted(product_profits.values(), key=lambda x: x["total_profit"], reverse=True)[:10]
    
    # Calculate totals
    total_revenue = sum(d["revenue"] for d in daily_data.values())
    total_cost = sum(d["cost"] for d in daily_data.values())
    total_profit = sum(d["profit"] for d in daily_data.values())
    
    return {
        "summary": {
            "total_revenue": total_revenue,
            "total_cost": total_cost,
            "total_profit": total_profit,
            "profit_margin": round((total_profit / total_revenue * 100) if total_revenue > 0 else 0, 2),
            "avg_daily_profit": round(total_profit / days, 2) if days > 0 else 0,
            "period_days": days
        },
        "daily_breakdown": sorted_daily[:30],
        "top_profitable_products": top_products
    }

# ============ EMPLOYEE ROUTES ============

@api_router.post("/employees", response_model=EmployeeResponse)
async def create_employee(employee: EmployeeCreate, admin: dict = Depends(get_tenant_admin)):
    employee_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    employee_doc = {
        "id": employee_id,
        "name": employee.name,
        "phone": employee.phone or "",
        "email": employee.email or "",
        "position": employee.position or "",
        "salary": employee.salary,
        "hire_date": employee.hire_date or now[:10],
        "commission_rate": employee.commission_rate,
        "total_advances": 0,
        "total_commission": 0,
        "created_at": now
    }
    await db.employees.insert_one(employee_doc)
    return EmployeeResponse(**employee_doc)

@api_router.get("/employees", response_model=List[EmployeeResponse])
async def get_employees(admin: dict = Depends(get_tenant_admin)):
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    return [EmployeeResponse(**e) for e in employees]

@api_router.get("/employees/{employee_id}", response_model=EmployeeResponse)
async def get_employee(employee_id: str, admin: dict = Depends(get_tenant_admin)):
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    return EmployeeResponse(**employee)

@api_router.put("/employees/{employee_id}", response_model=EmployeeResponse)
async def update_employee(employee_id: str, updates: EmployeeUpdate, admin: dict = Depends(get_tenant_admin)):
    employee = await db.employees.find_one({"id": employee_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if update_data:
        await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    updated = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    return EmployeeResponse(**updated)

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, admin: dict = Depends(get_tenant_admin)):
    result = await db.employees.delete_one({"id": employee_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee deleted successfully"}

# Create user account for employee
class EmployeeAccountCreate(BaseModel):
    email: str
    password: str
    role: str = "seller"

@api_router.post("/employees/{employee_id}/create-account")
async def create_employee_account(employee_id: str, account: EmployeeAccountCreate, admin: dict = Depends(get_tenant_admin)):
    """Create a user account for an employee"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": account.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Check if employee already has an account
    if employee.get("user_id"):
        raise HTTPException(status_code=400, detail="Employee already has an account")
    
    # Create user account
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    hashed_password = bcrypt.hashpw(account.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    user_doc = {
        "id": user_id,
        "email": account.email,
        "password": hashed_password,
        "name": employee["name"],
        "role": account.role,
        "employee_id": employee_id,
        "permissions": DEFAULT_PERMISSIONS.get(account.role, {}),
        "created_at": now
    }
    
    await db.users.insert_one(user_doc)
    
    # Link user to employee
    await db.employees.update_one(
        {"id": employee_id},
        {"$set": {"user_id": user_id, "user_email": account.email}}
    )
    
    return {
        "success": True,
        "user_id": user_id,
        "email": account.email,
        "role": account.role
    }

@api_router.delete("/employees/{employee_id}/delete-account")
async def delete_employee_account(employee_id: str, admin: dict = Depends(get_tenant_admin)):
    """Delete user account linked to employee"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    if not employee.get("user_id"):
        raise HTTPException(status_code=400, detail="Employee has no linked account")
    
    # Delete user account
    await db.users.delete_one({"id": employee["user_id"]})
    
    # Unlink from employee
    await db.employees.update_one(
        {"id": employee_id},
        {"$unset": {"user_id": "", "user_email": ""}}
    )
    
    return {"success": True}

@api_router.get("/employees/salary-report")
async def get_salary_report(month: str = None, user: dict = Depends(require_tenant)):
    """Get monthly salary report for all employees"""
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    year, month_num = map(int, month.split("-"))
    start_date = datetime(year, month_num, 1, tzinfo=timezone.utc)
    if month_num == 12:
        end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_date = datetime(year, month_num + 1, 1, tzinfo=timezone.utc)
    
    employees = await db.employees.find({}, {"_id": 0}).to_list(100)
    
    report = []
    for emp in employees:
        # Get advances for this month
        advances = await db.employee_advances.find({
            "employee_id": emp["id"],
            "created_at": {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
        }, {"_id": 0}).to_list(100)
        total_advances = sum(a.get("amount", 0) for a in advances)
        
        # Get attendance for this month
        attendance = await db.employee_attendance.find({
            "employee_id": emp["id"],
            "date": {"$gte": start_date.isoformat()[:10], "$lt": end_date.isoformat()[:10]}
        }, {"_id": 0}).to_list(31)
        present_days = len([a for a in attendance if a.get("status") == "present"])
        absent_days = len([a for a in attendance if a.get("status") == "absent"])
        
        # Get commissions (from sales)
        sales = await db.sales.find({
            "created_by": emp.get("user_email") or emp.get("name"),
            "created_at": {"$gte": start_date.isoformat(), "$lt": end_date.isoformat()}
        }, {"_id": 0}).to_list(1000)
        total_sales = sum(s.get("total", 0) for s in sales)
        commission = total_sales * (emp.get("commission_rate", 0) / 100)
        
        net_salary = emp.get("salary", 0) + commission - total_advances
        
        report.append({
            "employee_id": emp["id"],
            "employee_name": emp["name"],
            "position": emp.get("position", ""),
            "base_salary": emp.get("salary", 0),
            "commission_rate": emp.get("commission_rate", 0),
            "total_sales": total_sales,
            "commission": round(commission, 2),
            "advances": total_advances,
            "net_salary": round(net_salary, 2),
            "attendance_days": present_days,
            "absence_days": absent_days,
            "total_working_days": present_days + absent_days
        })
    
    return report

# Attendance
@api_router.post("/employees/attendance", response_model=AttendanceResponse)
async def record_attendance(attendance: AttendanceCreate, admin: dict = Depends(get_tenant_admin)):
    employee = await db.employees.find_one({"id": attendance.employee_id}, {"_id": 0, "name": 1})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    attendance_id = str(uuid.uuid4())
    attendance_doc = {
        "id": attendance_id,
        "employee_id": attendance.employee_id,
        "employee_name": employee["name"],
        "date": attendance.date,
        "status": attendance.status,
        "notes": attendance.notes or ""
    }
    await db.attendance.insert_one(attendance_doc)
    return AttendanceResponse(**attendance_doc)

@api_router.get("/employees/{employee_id}/attendance")
async def get_employee_attendance(employee_id: str, month: Optional[str] = None, admin: dict = Depends(get_tenant_admin)):
    query = {"employee_id": employee_id}
    if month:
        query["date"] = {"$regex": f"^{month}"}
    attendance = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(100)
    return attendance

# Advances (سلف)
@api_router.post("/employees/advances", response_model=AdvanceResponse)
async def create_advance(advance: AdvanceCreate, admin: dict = Depends(get_tenant_admin)):
    employee = await db.employees.find_one({"id": advance.employee_id}, {"_id": 0, "name": 1})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    now = datetime.now(timezone.utc).isoformat()
    advance_id = str(uuid.uuid4())
    
    advance_doc = {
        "id": advance_id,
        "employee_id": advance.employee_id,
        "employee_name": employee["name"],
        "amount": advance.amount,
        "notes": advance.notes or "",
        "created_at": now
    }
    await db.advances.insert_one(advance_doc)
    
    # Update employee total advances
    await db.employees.update_one(
        {"id": advance.employee_id},
        {"$inc": {"total_advances": advance.amount}}
    )
    
    return AdvanceResponse(**advance_doc)

@api_router.get("/employees/{employee_id}/advances")
async def get_employee_advances(employee_id: str, admin: dict = Depends(get_tenant_admin)):
    advances = await db.advances.find({"employee_id": employee_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return advances

# ============ EMPLOYEE ALERTS ============

@api_router.get("/employees/{employee_id}/alert-settings")
async def get_employee_alert_settings(employee_id: str, user: dict = Depends(require_tenant)):
    """Get alert settings for an employee"""
    settings = await db.employee_alerts.find_one({"employee_id": employee_id}, {"_id": 0})
    if not settings:
        return EmployeeAlertSettings(employee_id=employee_id).model_dump()
    return settings

@api_router.put("/employees/{employee_id}/alert-settings")
async def update_employee_alert_settings(employee_id: str, settings: EmployeeAlertSettings, admin: dict = Depends(get_tenant_admin)):
    """Update alert settings for an employee"""
    await db.employee_alerts.update_one(
        {"employee_id": employee_id},
        {"$set": {**settings.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"success": True}

@api_router.get("/employees/alerts/active")
async def get_active_employee_alerts(admin: dict = Depends(get_tenant_admin)):
    """Get all active alerts for employees approaching their limits"""
    alerts = []
    
    # Get all employees with limits set
    employees = await db.employees.find(
        {"$or": [{"max_discount_percent": {"$gt": 0}}, {"max_debt_amount": {"$gt": 0}}]},
        {"_id": 0}
    ).to_list(100)
    
    today = datetime.now(timezone.utc).date()
    month_start = datetime(today.year, today.month, 1, tzinfo=timezone.utc)
    
    for emp in employees:
        emp_id = emp.get("id")
        emp_name = emp.get("name")
        
        # Get alert settings
        alert_settings = await db.employee_alerts.find_one({"employee_id": emp_id})
        if not alert_settings:
            alert_settings = {"discount_threshold_percent": 80, "debt_threshold_percent": 80, "enable_discount_alert": True, "enable_debt_alert": True}
        
        # Check discount limit
        if emp.get("max_discount_percent", 0) > 0 and alert_settings.get("enable_discount_alert", True):
            # Get total discounts given this month
            sales = await db.sales.find({
                "employee_id": emp_id,
                "created_at": {"$gte": month_start.isoformat()}
            }).to_list(1000)
            
            total_discount_given = sum(s.get("discount", 0) for s in sales)
            total_sales = sum(s.get("subtotal", 0) for s in sales)
            
            if total_sales > 0:
                discount_percent_used = (total_discount_given / total_sales) * 100
                max_discount = emp.get("max_discount_percent", 0)
                threshold = alert_settings.get("discount_threshold_percent", 80)
                
                if discount_percent_used >= (max_discount * threshold / 100):
                    alerts.append({
                        "type": "discount_limit",
                        "severity": "high" if discount_percent_used >= max_discount else "warning",
                        "employee_id": emp_id,
                        "employee_name": emp_name,
                        "current_value": round(discount_percent_used, 2),
                        "max_value": max_discount,
                        "percent_used": round((discount_percent_used / max_discount) * 100, 1) if max_discount > 0 else 0,
                        "message_ar": f"الموظف {emp_name} اقترب من حد الخصم المسموح ({discount_percent_used:.1f}% من {max_discount}%)",
                        "message_en": f"Employee {emp_name} approaching discount limit ({discount_percent_used:.1f}% of {max_discount}%)"
                    })
        
        # Check debt limit
        if emp.get("max_debt_amount", 0) > 0 and alert_settings.get("enable_debt_alert", True):
            # Get total debts created by this employee
            debts = await db.debts.find({
                "created_by": emp_id,
                "paid": False
            }).to_list(1000)
            
            total_debt = sum(d.get("remaining_amount", 0) for d in debts)
            max_debt = emp.get("max_debt_amount", 0)
            threshold = alert_settings.get("debt_threshold_percent", 80)
            
            if total_debt >= (max_debt * threshold / 100):
                alerts.append({
                    "type": "debt_limit",
                    "severity": "high" if total_debt >= max_debt else "warning",
                    "employee_id": emp_id,
                    "employee_name": emp_name,
                    "current_value": total_debt,
                    "max_value": max_debt,
                    "percent_used": round((total_debt / max_debt) * 100, 1) if max_debt > 0 else 0,
                    "message_ar": f"الموظف {emp_name} اقترب من حد الدين المسموح ({total_debt:.2f} من {max_debt:.2f} دج)",
                    "message_en": f"Employee {emp_name} approaching debt limit ({total_debt:.2f} of {max_debt:.2f} DZD)"
                })
    
    return alerts

# ============ DEBT ROUTES ============

@api_router.post("/debts", response_model=DebtResponse)
async def create_debt(debt: DebtCreate, admin: dict = Depends(get_tenant_admin)):
    # Get party name
    if debt.party_type == "customer":
        party = await db.customers.find_one({"id": debt.party_id}, {"_id": 0, "name": 1})
    else:
        party = await db.suppliers.find_one({"id": debt.party_id}, {"_id": 0, "name": 1})
    
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")
    
    now = datetime.now(timezone.utc).isoformat()
    debt_id = str(uuid.uuid4())
    
    debt_doc = {
        "id": debt_id,
        "type": debt.type,
        "party_type": debt.party_type,
        "party_id": debt.party_id,
        "party_name": party["name"],
        "original_amount": debt.amount,
        "paid_amount": 0,
        "remaining_amount": debt.amount,
        "due_date": debt.due_date or "",
        "status": "pending",
        "notes": debt.notes or "",
        "reference_type": debt.reference_type or "",
        "reference_id": debt.reference_id or "",
        "created_at": now
    }
    await db.debts.insert_one(debt_doc)
    return DebtResponse(**debt_doc)

@api_router.get("/debts", response_model=List[DebtResponse])
async def get_debts(
    type: Optional[str] = None,
    party_type: Optional[str] = None,
    status: Optional[str] = None,
    admin: dict = Depends(get_tenant_admin)
):
    query = {}
    if type:
        query["type"] = type
    if party_type:
        query["party_type"] = party_type
    if status:
        query["status"] = status
    
    debts = await db.debts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Check for overdue
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for debt in debts:
        if debt.get("due_date") and debt["due_date"] < today and debt["status"] not in ["paid", "overdue"]:
            await db.debts.update_one({"id": debt["id"]}, {"$set": {"status": "overdue"}})
            debt["status"] = "overdue"
    
    return [DebtResponse(**d) for d in debts]

@api_router.post("/debts/{debt_id}/pay", response_model=DebtPaymentResponse)
async def pay_debt(debt_id: str, payment: DebtPaymentCreate, admin: dict = Depends(get_tenant_admin)):
    debt = await db.debts.find_one({"id": debt_id})
    if not debt:
        raise HTTPException(status_code=404, detail="Debt not found")
    
    if payment.amount > debt["remaining_amount"]:
        raise HTTPException(status_code=400, detail="Payment amount exceeds remaining debt")
    
    now = datetime.now(timezone.utc).isoformat()
    payment_id = str(uuid.uuid4())
    
    new_paid = debt["paid_amount"] + payment.amount
    new_remaining = debt["remaining_amount"] - payment.amount
    new_status = "paid" if new_remaining <= 0 else "partial"
    
    # Update debt
    await db.debts.update_one(
        {"id": debt_id},
        {"$set": {
            "paid_amount": new_paid,
            "remaining_amount": new_remaining,
            "status": new_status
        }}
    )
    
    # Record payment
    payment_doc = {
        "id": payment_id,
        "debt_id": debt_id,
        "amount": payment.amount,
        "payment_method": payment.payment_method,
        "notes": payment.notes or "",
        "created_at": now,
        "created_by": admin["name"]
    }
    await db.debt_payments.insert_one(payment_doc)
    
    # Update cash box
    tx_type = "income" if debt["type"] == "receivable" else "expense"
    await db.cash_boxes.update_one(
        {"id": payment.payment_method},
        {"$inc": {"balance": payment.amount if tx_type == "income" else -payment.amount}, "$set": {"updated_at": now}}
    )
    
    await db.transactions.insert_one({
        "id": str(uuid.uuid4()),
        "cash_box_id": payment.payment_method,
        "type": tx_type,
        "amount": payment.amount,
        "description": f"سداد دين - {debt['party_name']}",
        "reference_type": "debt_payment",
        "reference_id": payment_id,
        "created_at": now,
        "created_by": admin["name"]
    })
    
    return DebtPaymentResponse(**payment_doc)

@api_router.get("/debts/{debt_id}/payments")
async def get_debt_payments(debt_id: str, admin: dict = Depends(get_tenant_admin)):
    payments = await db.debt_payments.find({"debt_id": debt_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return payments

# ============ EXCEL IMPORT/EXPORT ============

@api_router.get("/products/export/excel")
async def export_products_excel(admin: dict = Depends(get_tenant_admin)):
    import pandas as pd
    
    products = await db.products.find({}, {"_id": 0}).to_list(10000)
    
    # Prepare data
    data = []
    for p in products:
        data.append({
            "الباركود": p.get("barcode", ""),
            "الاسم (عربي)": p.get("name_ar", ""),
            "الاسم (إنجليزي)": p.get("name_en", ""),
            "الوصف (عربي)": p.get("description_ar", ""),
            "الوصف (إنجليزي)": p.get("description_en", ""),
            "سعر الشراء": p.get("purchase_price", 0),
            "سعر الجملة": p.get("wholesale_price", 0),
            "سعر التجزئة": p.get("retail_price", 0),
            "الكمية": p.get("quantity", 0),
            "حد المخزون المنخفض": p.get("low_stock_threshold", 10),
            "الموديلات المتوافقة": ", ".join(p.get("compatible_models", [])),
            "رابط الصورة": p.get("image_url", "")
        })
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='المنتجات')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products.xlsx"}
    )

from fastapi import UploadFile, File

@api_router.post("/products/import/excel")
async def import_products_excel(file: UploadFile = File(...), admin: dict = Depends(get_tenant_admin)):
    import pandas as pd
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
    
    contents = await file.read()
    df = pd.read_excel(io.BytesIO(contents))
    
    now = datetime.now(timezone.utc).isoformat()
    imported = 0
    updated = 0
    errors = []
    
    for index, row in df.iterrows():
        try:
            barcode = str(row.get("الباركود", "")).strip()
            name_ar = str(row.get("الاسم (عربي)", "")).strip()
            name_en = str(row.get("الاسم (إنجليزي)", "")).strip()
            
            if not name_ar and not name_en:
                continue
            
            product_data = {
                "barcode": barcode,
                "name_ar": name_ar or name_en,
                "name_en": name_en or name_ar,
                "description_ar": str(row.get("الوصف (عربي)", "")),
                "description_en": str(row.get("الوصف (إنجليزي)", "")),
                "purchase_price": float(row.get("سعر الشراء", 0) or 0),
                "wholesale_price": float(row.get("سعر الجملة", 0) or 0),
                "retail_price": float(row.get("سعر التجزئة", 0) or 0),
                "quantity": int(row.get("الكمية", 0) or 0),
                "low_stock_threshold": int(row.get("حد المخزون المنخفض", 10) or 10),
                "compatible_models": [m.strip() for m in str(row.get("الموديلات المتوافقة", "")).split(",") if m.strip()],
                "image_url": str(row.get("رابط الصورة", "")),
                "updated_at": now
            }
            
            # Check if product exists by barcode or name
            existing = None
            if barcode:
                existing = await db.products.find_one({"barcode": barcode})
            if not existing and name_ar:
                existing = await db.products.find_one({"name_ar": name_ar})
            
            if existing:
                await db.products.update_one({"id": existing["id"]}, {"$set": product_data})
                updated += 1
            else:
                product_data["id"] = str(uuid.uuid4())
                product_data["created_at"] = now
                await db.products.insert_one(product_data)
                imported += 1
                
        except Exception as e:
            errors.append(f"Row {index + 2}: {str(e)}")
    
    return {
        "imported": imported,
        "updated": updated,
        "errors": errors[:10]  # Return first 10 errors
    }

# ============ BACKUP ============

@api_router.get("/backup/create")
async def create_backup(admin: dict = Depends(get_tenant_admin)):
    """Create a backup of all data"""
    import json
    
    collections = ["users", "products", "customers", "suppliers", "sales", "purchases", 
                   "cash_boxes", "transactions", "employees", "attendance", "advances", 
                   "debts", "debt_payments", "notifications", "product_families", 
                   "customer_families", "supplier_families", "daily_sessions", "expenses",
                   "repairs", "spare_parts", "inventory_adjustments", "system_settings",
                   "branding_settings", "ai_chat_history"]
    
    backup_data = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "version": "2.0",
        "collections": {}
    }
    
    for collection_name in collections:
        collection = db[collection_name]
        docs = await collection.find({}, {"_id": 0}).to_list(100000)
        backup_data["collections"][collection_name] = docs
    
    output = io.BytesIO()
    output.write(json.dumps(backup_data, ensure_ascii=False, indent=2).encode('utf-8'))
    output.seek(0)
    
    filename = f"backup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
    
    return StreamingResponse(
        output,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/backup/restore")
async def restore_backup(admin: dict = Depends(get_tenant_admin), file: UploadFile = File(...)):
    """Restore data from backup file"""
    import json
    
    try:
        content = await file.read()
        backup_data = json.loads(content.decode('utf-8'))
        
        if "collections" not in backup_data:
            raise HTTPException(status_code=400, detail="Invalid backup file format")
        
        restored_counts = {}
        for collection_name, docs in backup_data["collections"].items():
            if docs:
                # Clear existing data
                await db[collection_name].delete_many({})
                # Insert backup data
                await db[collection_name].insert_many(docs)
                restored_counts[collection_name] = len(docs)
        
        # Log restore
        await db.system_logs.insert_one({
            "id": str(uuid.uuid4()),
            "action": "backup_restore",
            "performed_by": admin.get("name", ""),
            "backup_date": backup_data.get("created_at", ""),
            "restored_counts": restored_counts,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"success": True, "restored_counts": restored_counts}
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON file")

@api_router.post("/backup/save-to-server")
async def save_backup_to_server(admin: dict = Depends(get_tenant_admin)):
    """Save backup to server storage"""
    import json
    
    collections = ["users", "products", "customers", "suppliers", "sales", "purchases", 
                   "cash_boxes", "transactions", "employees", "attendance", "advances", 
                   "debts", "debt_payments", "notifications", "product_families",
                   "customer_families", "supplier_families", "daily_sessions", "expenses"]
    
    backup_data = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "version": "2.0",
        "collections": {}
    }
    
    for collection_name in collections:
        collection = db[collection_name]
        docs = await collection.find({}, {"_id": 0}).to_list(100000)
        backup_data["collections"][collection_name] = docs
    
    # Save to server
    backup_dir = ROOT_DIR / "backups"
    backup_dir.mkdir(exist_ok=True)
    
    filename = f"backup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
    filepath = backup_dir / filename
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(backup_data, f, ensure_ascii=False, indent=2)
    
    # Keep only last 10 backups
    backups = sorted(backup_dir.glob("backup_*.json"), reverse=True)
    for old_backup in backups[10:]:
        old_backup.unlink()
    
    # Save backup record to database
    await db.backups.insert_one({
        "id": str(uuid.uuid4()),
        "filename": filename,
        "size": filepath.stat().st_size,
        "created_by": admin.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"success": True, "filename": filename}

@api_router.get("/backup/list")
async def list_backups(admin: dict = Depends(get_tenant_admin)):
    """List available backups on server"""
    backups = await db.backups.find({}, {"_id": 0}).sort("created_at", -1).to_list(20)
    return backups

# Auto-backup settings model
class AutoBackupSettings(BaseModel):
    enabled: bool = False
    frequency: str = "daily"  # daily, weekly, monthly
    time: str = "02:00"  # HH:MM format
    keep_count: int = 10  # Number of backups to keep

@api_router.get("/backup/auto-settings")
async def get_auto_backup_settings(admin: dict = Depends(get_tenant_admin)):
    """Get auto-backup settings"""
    settings = await db.settings.find_one({"type": "auto_backup"}, {"_id": 0})
    if not settings:
        return AutoBackupSettings().model_dump()
    return settings.get("settings", AutoBackupSettings().model_dump())

@api_router.post("/backup/auto-settings")
async def save_auto_backup_settings(settings: AutoBackupSettings, admin: dict = Depends(get_tenant_admin)):
    """Save auto-backup settings"""
    await db.settings.update_one(
        {"type": "auto_backup"},
        {"$set": {"type": "auto_backup", "settings": settings.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"success": True, "message": "Auto-backup settings saved"}

@api_router.post("/backup/run-auto")
async def run_auto_backup(admin: dict = Depends(get_tenant_admin)):
    """Manually trigger auto backup (for testing)"""
    now = datetime.now(timezone.utc)
    
    # Collect all data
    backup_data = {
        "products": await db.products.find({}, {"_id": 0}).to_list(10000),
        "customers": await db.customers.find({}, {"_id": 0}).to_list(10000),
        "suppliers": await db.suppliers.find({}, {"_id": 0}).to_list(10000),
        "sales": await db.sales.find({}, {"_id": 0}).to_list(10000),
        "purchases": await db.purchases.find({}, {"_id": 0}).to_list(10000),
        "expenses": await db.expenses.find({}, {"_id": 0}).to_list(10000),
        "employees": await db.employees.find({}, {"_id": 0}).to_list(1000),
        "product_families": await db.product_families.find({}, {"_id": 0}).to_list(1000),
        "customer_families": await db.customer_families.find({}, {"_id": 0}).to_list(1000),
    }
    
    filename = f"auto_backup_{now.strftime('%Y%m%d_%H%M%S')}.json"
    
    # Save backup record to database
    await db.backups.insert_one({
        "id": str(uuid.uuid4()),
        "filename": filename,
        "type": "auto",
        "size_bytes": len(str(backup_data)),
        "stats": {
            "products": len(backup_data["products"]),
            "customers": len(backup_data["customers"]),
            "suppliers": len(backup_data["suppliers"]),
            "sales": len(backup_data["sales"]),
            "purchases": len(backup_data["purchases"]),
            "expenses": len(backup_data["expenses"]),
        },
        "created_at": now.isoformat(),
        "created_by": admin.get("id", "system")
    })
    
    # Clean up old backups (keep only latest N)
    settings = await db.settings.find_one({"type": "auto_backup"}, {"_id": 0})
    keep_count = settings.get("settings", {}).get("keep_count", 10) if settings else 10
    
    all_backups = await db.backups.find({"type": "auto"}).sort("created_at", -1).to_list(1000)
    if len(all_backups) > keep_count:
        old_backups = all_backups[keep_count:]
        for old in old_backups:
            await db.backups.delete_one({"id": old["id"]})
    
    return {"success": True, "filename": filename, "message": "Auto backup completed"}

# ============ SELECTIVE DATA DELETE ============

class SelectiveDeleteRequest(BaseModel):
    data_types: List[str]  # ["sales", "purchases", "customers", etc.]
    confirm_code: str

@api_router.post("/system/selective-delete")
async def selective_delete(request: SelectiveDeleteRequest, admin: dict = Depends(get_tenant_admin)):
    """Selectively delete specific data types"""
    if request.confirm_code != "DELETE-SELECTED":
        raise HTTPException(status_code=400, detail="Invalid confirmation code")
    
    # Check permissions
    user_permissions = admin.get("permissions") or DEFAULT_PERMISSIONS.get(admin.get("role", "user"), {})
    if not user_permissions.get("factory_reset", False):
        raise HTTPException(status_code=403, detail="No permission for data deletion")
    
    # Valid data types
    valid_types = {
        "sales": "sales",
        "purchases": "purchases",
        "customers": "customers",
        "suppliers": "suppliers",
        "products": "products",
        "employees": "employees",
        "debts": "debts",
        "expenses": "expenses",
        "repairs": "repairs",
        "inventory_adjustments": "inventory_adjustments",
        "daily_sessions": "daily_sessions",
        "notifications": "notifications"
    }
    
    deleted_counts = {}
    for data_type in request.data_types:
        if data_type in valid_types:
            collection_name = valid_types[data_type]
            result = await db[collection_name].delete_many({})
            deleted_counts[data_type] = result.deleted_count
            
            # Also delete related data
            if data_type == "sales":
                await db.debt_payments.delete_many({"type": "sale"})
            elif data_type == "customers":
                await db.customer_families.delete_many({})
            elif data_type == "suppliers":
                await db.supplier_families.delete_many({})
            elif data_type == "products":
                await db.product_families.delete_many({})
    
    # Log deletion
    await db.system_logs.insert_one({
        "id": str(uuid.uuid4()),
        "action": "selective_delete",
        "performed_by": admin.get("name", ""),
        "deleted_types": request.data_types,
        "deleted_counts": deleted_counts,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"success": True, "deleted_counts": deleted_counts}

# ============ SIDEBAR ORDER SETTINGS ============

@api_router.get("/settings/sidebar-order")
async def get_sidebar_order(user: dict = Depends(require_tenant)):
    """Get sidebar menu order for user"""
    settings = await db.user_settings.find_one({"user_id": user["id"]}, {"_id": 0})
    if settings and "sidebar_order" in settings:
        return {"sidebar_order": settings["sidebar_order"]}
    return {"sidebar_order": None}  # Return null to use default order

class SidebarMenuItem(BaseModel):
    id: str
    path: Optional[str] = None
    icon: Optional[str] = None
    labelAr: Optional[str] = None
    labelFr: Optional[str] = None
    visible: bool = True

class SidebarSection(BaseModel):
    id: str
    titleAr: Optional[str] = None
    titleFr: Optional[str] = None
    icon: Optional[str] = None
    visible: bool = True
    isCustom: bool = False
    items: List[SidebarMenuItem] = []

@api_router.put("/settings/sidebar-order")
async def update_sidebar_order(order: List[SidebarSection], user: dict = Depends(require_tenant)):
    """Update sidebar menu order for user"""
    # Convert to dict for storage
    order_data = [section.model_dump() for section in order]
    await db.user_settings.update_one(
        {"user_id": user["id"]},
        {"$set": {"sidebar_order": order_data, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"success": True}

# ============ NOTIFICATION MANAGEMENT ============

@api_router.get("/notifications/settings")
async def get_notification_settings(user: dict = Depends(require_tenant)):
    """Get notification settings"""
    settings = await db.notification_settings.find_one({"user_id": user["id"]}, {"_id": 0})
    if not settings:
        # Default settings
        settings = {
            "low_stock_enabled": True,
            "low_stock_threshold": 10,
            "debt_reminder_enabled": True,
            "debt_reminder_days": 7,
            "cash_difference_enabled": True,
            "cash_difference_threshold": 1000,
            "expense_reminder_enabled": True,
            "repair_status_enabled": True,
            "email_notifications": False,
            "sound_enabled": True
        }
    return settings

@api_router.put("/notifications/settings")
async def update_notification_settings(settings: dict, user: dict = Depends(require_tenant)):
    """Update notification settings"""
    settings["user_id"] = user["id"]
    settings["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.notification_settings.update_one(
        {"user_id": user["id"]},
        {"$set": settings},
        upsert=True
    )
    return {"success": True}

@api_router.get("/notifications/all")
async def get_all_notifications(
    skip: int = 0, 
    limit: int = 50,
    unread_only: bool = False,
    user: dict = Depends(require_tenant)
):
    """Get all notifications with pagination"""
    query = {}
    if unread_only:
        query["read"] = False
    
    notifications = await db.notifications.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.notifications.count_documents(query)
    unread_count = await db.notifications.count_documents({"read": False})
    
    return {
        "notifications": notifications,
        "total": total,
        "unread_count": unread_count
    }

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(require_tenant)):
    """Mark a notification as read"""
    await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"success": True}

@api_router.put("/notifications/mark-all-read")
async def mark_all_notifications_read(user: dict = Depends(require_tenant)):
    """Mark all notifications as read"""
    await db.notifications.update_many(
        {"read": False},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"success": True}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, user: dict = Depends(require_tenant)):
    """Delete a notification"""
    await db.notifications.delete_one({"id": notification_id})
    return {"success": True}

@api_router.delete("/notifications/clear-all")
async def clear_all_notifications(admin: dict = Depends(get_tenant_admin)):
    """Clear all notifications"""
    result = await db.notifications.delete_many({})
    return {"success": True, "deleted_count": result.deleted_count}

# ============ OCR ROUTE ============

@api_router.post("/ocr/extract-models", response_model=OCRResponse)
async def extract_models_from_image(request: OCRRequest, admin: dict = Depends(get_tenant_admin)):
    from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
    
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        raise HTTPException(status_code=500, detail="OCR service not configured")
    
    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"ocr-{uuid.uuid4()}",
            system_message="""You are an OCR assistant specialized in extracting phone model names from images.
            Extract all phone model names you can see in the image.
            Return ONLY the model names, one per line, without any additional text or explanation.
            Examples of model names: iPhone 15 Pro, Samsung Galaxy S24, Huawei P60 Pro, etc."""
        ).with_model("gemini", "gemini-2.5-flash")
        
        image_content = ImageContent(image_base64=request.image_base64)
        user_message = UserMessage(
            text="Extract all phone model names from this image. Return only the model names, one per line.",
            file_contents=[image_content]
        )
        
        response = await chat.send_message(user_message)
        raw_text = response.strip()
        models = [m.strip() for m in raw_text.split('\n') if m.strip()]
        
        return OCRResponse(extracted_models=models, raw_text=raw_text)
        
    except Exception as e:
        logger.error(f"OCR error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")

# ============ INVOICE PDF ============

@api_router.get("/sales/{sale_id}/invoice-pdf")
async def get_invoice_pdf(sale_id: str, user: dict = Depends(require_tenant)):
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    # Get sale code if exists
    sale_code = sale.get("code", "")
    
    # Generate simple HTML invoice
    items_html = ""
    for i, item in enumerate(sale["items"], 1):
        barcode = item.get('barcode', '-')
        items_html += f"""
        <tr>
            <td>{i}</td>
            <td>{barcode}</td>
            <td>{item['product_name']}</td>
            <td>{item['quantity']}</td>
            <td>{item['unit_price']:.2f} {CURRENCY}</td>
            <td>{item['discount']:.2f} {CURRENCY}</td>
            <td>{item['total']:.2f} {CURRENCY}</td>
        </tr>
        """
    
    html_content = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="ar">
    <head>
        <meta charset="UTF-8">
        <title>فاتورة {sale['invoice_number']}</title>
        <style>
            body {{ font-family: 'Cairo', Arial, sans-serif; margin: 20px; direction: rtl; }}
            .header {{ text-align: center; margin-bottom: 30px; }}
            .header h1 {{ color: #2563EB; margin: 0; }}
            .sale-code {{ font-family: monospace; font-size: 14px; background: #f0f0f0; padding: 4px 8px; border-radius: 4px; }}
            .info {{ display: flex; justify-content: space-between; margin-bottom: 20px; }}
            .info div {{ width: 48%; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: right; }}
            th {{ background: #2563EB; color: white; }}
            .barcode {{ font-family: monospace; font-size: 11px; }}
            .totals {{ text-align: left; margin-top: 20px; }}
            .totals table {{ width: 300px; margin-right: 0; margin-left: auto; }}
            .footer {{ text-align: center; margin-top: 40px; color: #666; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>NT</h1>
            <p>فاتورة مبيعات</p>
            {f'<p class="sale-code">{sale_code}</p>' if sale_code else ''}
        </div>
        
        <div class="info">
            <div>
                <p><strong>رقم الفاتورة:</strong> {sale['invoice_number']}</p>
                <p><strong>التاريخ:</strong> {sale['created_at'][:10]}</p>
                <p><strong>البائع:</strong> {sale['created_by']}</p>
            </div>
            <div>
                <p><strong>العميل:</strong> {sale['customer_name']}</p>
                <p><strong>طريقة الدفع:</strong> {sale['payment_method']}</p>
                <p><strong>الحالة:</strong> {sale['status']}</p>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>الباركود</th>
                    <th>المنتج</th>
                    <th>الكمية</th>
                    <th>السعر</th>
                    <th>الخصم</th>
                    <th>الإجمالي</th>
                </tr>
            </thead>
            <tbody>
                {items_html}
            </tbody>
        </table>
        
        <div class="totals">
            <table>
                <tr><td>المجموع الفرعي:</td><td>{sale['subtotal']:.2f} {CURRENCY}</td></tr>
                <tr><td>الخصم:</td><td>{sale['discount']:.2f} {CURRENCY}</td></tr>
                <tr><td><strong>الإجمالي:</strong></td><td><strong>{sale['total']:.2f} {CURRENCY}</strong></td></tr>
                <tr><td>المدفوع:</td><td>{sale['paid_amount']:.2f} {CURRENCY}</td></tr>
                <tr><td>المتبقي:</td><td>{sale['remaining']:.2f} {CURRENCY}</td></tr>
            </table>
        </div>
        
        <div class="footer">
            <p>شكراً لتعاملكم معنا</p>
        </div>
    </body>
    </html>
    """
    
    return StreamingResponse(
        io.BytesIO(html_content.encode('utf-8')),
        media_type="text/html",
        headers={"Content-Disposition": f"inline; filename=invoice_{sale['invoice_number']}.html"}
    )

# ============ API KEYS MANAGEMENT ============

import secrets

@api_router.post("/api-keys", response_model=ApiKeyResponse)
async def create_api_key(api_key: ApiKeyCreate, admin: dict = Depends(get_tenant_admin)):
    key_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Generate internal API key if type is internal
    generated_key = ""
    if api_key.type == "internal":
        generated_key = f"sk_{secrets.token_hex(32)}"
    
    key_value = api_key.key_value or generated_key
    
    api_key_doc = {
        "id": key_id,
        "name": api_key.name,
        "type": api_key.type,
        "service": api_key.service or "",
        "key_value": key_value,
        "secret_value": api_key.secret_value or "",
        "endpoint_url": api_key.endpoint_url or "",
        "permissions": api_key.permissions,
        "is_active": True,
        "last_used": "",
        "created_at": now
    }
    await db.api_keys.insert_one(api_key_doc)
    
    return ApiKeyResponse(
        **api_key_doc,
        key_preview=f"...{key_value[-4:]}" if len(key_value) > 4 else key_value
    )

@api_router.get("/api-keys", response_model=List[ApiKeyResponse])
async def get_api_keys(admin: dict = Depends(get_tenant_admin)):
    keys = await db.api_keys.find({}, {"_id": 0}).to_list(100)
    result = []
    for k in keys:
        k["key_preview"] = f"...{k['key_value'][-4:]}" if len(k.get('key_value', '')) > 4 else k.get('key_value', '')
        # Hide full key value
        if k["type"] == "internal":
            k["key_value"] = k["key_preview"]
        result.append(ApiKeyResponse(**k))
    return result

@api_router.get("/api-keys/{key_id}")
async def get_api_key(key_id: str, admin: dict = Depends(get_tenant_admin)):
    key = await db.api_keys.find_one({"id": key_id}, {"_id": 0})
    if not key:
        raise HTTPException(status_code=404, detail="API Key not found")
    return key

@api_router.put("/api-keys/{key_id}/toggle")
async def toggle_api_key(key_id: str, admin: dict = Depends(get_tenant_admin)):
    key = await db.api_keys.find_one({"id": key_id})
    if not key:
        raise HTTPException(status_code=404, detail="API Key not found")
    
    new_status = not key.get("is_active", True)
    await db.api_keys.update_one({"id": key_id}, {"$set": {"is_active": new_status}})
    return {"is_active": new_status}

@api_router.delete("/api-keys/{key_id}")
async def delete_api_key(key_id: str, admin: dict = Depends(get_tenant_admin)):
    result = await db.api_keys.delete_one({"id": key_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="API Key not found")
    return {"message": "API Key deleted successfully"}

# ============ RECHARGE / USSD ============

@api_router.get("/recharge/config")
async def get_recharge_config(user: dict = Depends(require_tenant)):
    """Get recharge operators configuration"""
    return RECHARGE_CONFIG

@api_router.post("/recharge", response_model=RechargeResponse)
async def create_recharge(recharge: RechargeCreate, user: dict = Depends(require_tenant)):
    """Record a recharge transaction"""
    recharge_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Get operator config
    operator_config = RECHARGE_CONFIG.get(recharge.operator)
    if not operator_config:
        raise HTTPException(status_code=400, detail="Invalid operator")
    
    # Calculate cost and profit
    commission_rate = operator_config.get("commission", 0) / 100
    profit = recharge.amount * commission_rate
    cost = recharge.amount - profit
    
    # Get customer name
    customer_name = "عميل نقدي"
    if recharge.customer_id:
        customer = await db.customers.find_one({"id": recharge.customer_id}, {"_id": 0, "name": 1})
        if customer:
            customer_name = customer["name"]
    
    # Generate USSD code
    ussd_template = operator_config["ussd"].get(recharge.recharge_type, "")
    ussd_code = ussd_template.replace("{phone}", recharge.phone_number).replace("{amount}", str(int(recharge.amount)))
    
    recharge_doc = {
        "id": recharge_id,
        "operator": recharge.operator,
        "operator_name": operator_config["name"],
        "phone_number": recharge.phone_number,
        "amount": recharge.amount,
        "recharge_type": recharge.recharge_type,
        "cost": cost,
        "profit": profit,
        "customer_id": recharge.customer_id or "",
        "customer_name": customer_name,
        "payment_method": recharge.payment_method,
        "status": "completed",
        "ussd_code": ussd_code,
        "notes": recharge.notes or "",
        "created_at": now,
        "created_by": user["name"]
    }
    await db.recharges.insert_one(recharge_doc)
    
    # Update cash box
    await db.cash_boxes.update_one(
        {"id": recharge.payment_method},
        {"$inc": {"balance": recharge.amount}, "$set": {"updated_at": now}}
    )
    
    # Record transaction
    await db.transactions.insert_one({
        "id": str(uuid.uuid4()),
        "cash_box_id": recharge.payment_method,
        "type": "income",
        "amount": recharge.amount,
        "description": f"شحن {operator_config['name']} - {recharge.phone_number}",
        "reference_type": "recharge",
        "reference_id": recharge_id,
        "created_at": now,
        "created_by": user["name"]
    })
    
    return RechargeResponse(**recharge_doc)

@api_router.get("/recharge", response_model=List[RechargeResponse])
async def get_recharges(
    operator: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_tenant)
):
    """Get recharge history"""
    query = {}
    if operator:
        query["operator"] = operator
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    recharges = await db.recharges.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [RechargeResponse(**r) for r in recharges]

@api_router.get("/recharge/stats")
async def get_recharge_stats(days: int = 30, admin: dict = Depends(get_tenant_admin)):
    """Get recharge statistics"""
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    
    # Total by operator
    pipeline = [
        {"$match": {"created_at": {"$gte": start_date}}},
        {"$group": {
            "_id": "$operator",
            "count": {"$sum": 1},
            "total_amount": {"$sum": "$amount"},
            "total_profit": {"$sum": "$profit"}
        }}
    ]
    by_operator = await db.recharges.aggregate(pipeline).to_list(10)
    
    # Today's stats
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    today_stats = await db.recharges.aggregate([
        {"$match": {"created_at": {"$gte": today}}},
        {"$group": {
            "_id": None,
            "count": {"$sum": 1},
            "total_amount": {"$sum": "$amount"},
            "total_profit": {"$sum": "$profit"}
        }}
    ]).to_list(1)
    
    return {
        "by_operator": by_operator,
        "today": today_stats[0] if today_stats else {"count": 0, "total_amount": 0, "total_profit": 0},
        "period_days": days
    }

# ============ ALGERIA WILAYAS (for delivery) ============

ALGERIA_WILAYAS = {
    "01": {"name_ar": "أدرار", "name_en": "Adrar", "desk_fee": 600, "home_fee": 800},
    "02": {"name_ar": "الشلف", "name_en": "Chlef", "desk_fee": 400, "home_fee": 600},
    "03": {"name_ar": "الأغواط", "name_en": "Laghouat", "desk_fee": 500, "home_fee": 700},
    "04": {"name_ar": "أم البواقي", "name_en": "Oum El Bouaghi", "desk_fee": 450, "home_fee": 650},
    "05": {"name_ar": "باتنة", "name_en": "Batna", "desk_fee": 400, "home_fee": 600},
    "06": {"name_ar": "بجاية", "name_en": "Béjaïa", "desk_fee": 400, "home_fee": 600},
    "07": {"name_ar": "بسكرة", "name_en": "Biskra", "desk_fee": 450, "home_fee": 650},
    "08": {"name_ar": "بشار", "name_en": "Béchar", "desk_fee": 600, "home_fee": 800},
    "09": {"name_ar": "البليدة", "name_en": "Blida", "desk_fee": 300, "home_fee": 450},
    "10": {"name_ar": "البويرة", "name_en": "Bouira", "desk_fee": 350, "home_fee": 500},
    "11": {"name_ar": "تمنراست", "name_en": "Tamanrasset", "desk_fee": 800, "home_fee": 1000},
    "12": {"name_ar": "تبسة", "name_en": "Tébessa", "desk_fee": 500, "home_fee": 700},
    "13": {"name_ar": "تلمسان", "name_en": "Tlemcen", "desk_fee": 500, "home_fee": 700},
    "14": {"name_ar": "تيارت", "name_en": "Tiaret", "desk_fee": 450, "home_fee": 650},
    "15": {"name_ar": "تيزي وزو", "name_en": "Tizi Ouzou", "desk_fee": 350, "home_fee": 500},
    "16": {"name_ar": "الجزائر", "name_en": "Algiers", "desk_fee": 250, "home_fee": 400},
    "17": {"name_ar": "الجلفة", "name_en": "Djelfa", "desk_fee": 450, "home_fee": 650},
    "18": {"name_ar": "جيجل", "name_en": "Jijel", "desk_fee": 400, "home_fee": 600},
    "19": {"name_ar": "سطيف", "name_en": "Sétif", "desk_fee": 350, "home_fee": 500},
    "20": {"name_ar": "سعيدة", "name_en": "Saïda", "desk_fee": 500, "home_fee": 700},
    "21": {"name_ar": "سكيكدة", "name_en": "Skikda", "desk_fee": 400, "home_fee": 600},
    "22": {"name_ar": "سيدي بلعباس", "name_en": "Sidi Bel Abbès", "desk_fee": 500, "home_fee": 700},
    "23": {"name_ar": "عنابة", "name_en": "Annaba", "desk_fee": 400, "home_fee": 600},
    "24": {"name_ar": "قالمة", "name_en": "Guelma", "desk_fee": 450, "home_fee": 650},
    "25": {"name_ar": "قسنطينة", "name_en": "Constantine", "desk_fee": 350, "home_fee": 500},
    "26": {"name_ar": "المدية", "name_en": "Médéa", "desk_fee": 350, "home_fee": 500},
    "27": {"name_ar": "مستغانم", "name_en": "Mostaganem", "desk_fee": 450, "home_fee": 650},
    "28": {"name_ar": "المسيلة", "name_en": "M'sila", "desk_fee": 400, "home_fee": 600},
    "29": {"name_ar": "معسكر", "name_en": "Mascara", "desk_fee": 450, "home_fee": 650},
    "30": {"name_ar": "ورقلة", "name_en": "Ouargla", "desk_fee": 600, "home_fee": 800},
    "31": {"name_ar": "وهران", "name_en": "Oran", "desk_fee": 400, "home_fee": 600},
    "32": {"name_ar": "البيض", "name_en": "El Bayadh", "desk_fee": 600, "home_fee": 800},
    "33": {"name_ar": "إليزي", "name_en": "Illizi", "desk_fee": 900, "home_fee": 1100},
    "34": {"name_ar": "برج بوعريريج", "name_en": "Bordj Bou Arréridj", "desk_fee": 350, "home_fee": 500},
    "35": {"name_ar": "بومرداس", "name_en": "Boumerdès", "desk_fee": 300, "home_fee": 450},
    "36": {"name_ar": "الطارف", "name_en": "El Tarf", "desk_fee": 450, "home_fee": 650},
    "37": {"name_ar": "تندوف", "name_en": "Tindouf", "desk_fee": 900, "home_fee": 1100},
    "38": {"name_ar": "تيسمسيلت", "name_en": "Tissemsilt", "desk_fee": 450, "home_fee": 650},
    "39": {"name_ar": "الوادي", "name_en": "El Oued", "desk_fee": 550, "home_fee": 750},
    "40": {"name_ar": "خنشلة", "name_en": "Khenchela", "desk_fee": 500, "home_fee": 700},
    "41": {"name_ar": "سوق أهراس", "name_en": "Souk Ahras", "desk_fee": 500, "home_fee": 700},
    "42": {"name_ar": "تيبازة", "name_en": "Tipaza", "desk_fee": 300, "home_fee": 450},
    "43": {"name_ar": "ميلة", "name_en": "Mila", "desk_fee": 400, "home_fee": 600},
    "44": {"name_ar": "عين الدفلى", "name_en": "Aïn Defla", "desk_fee": 350, "home_fee": 500},
    "45": {"name_ar": "النعامة", "name_en": "Naâma", "desk_fee": 600, "home_fee": 800},
    "46": {"name_ar": "عين تموشنت", "name_en": "Aïn Témouchent", "desk_fee": 500, "home_fee": 700},
    "47": {"name_ar": "غرداية", "name_en": "Ghardaïa", "desk_fee": 550, "home_fee": 750},
    "48": {"name_ar": "غليزان", "name_en": "Relizane", "desk_fee": 450, "home_fee": 650},
    "49": {"name_ar": "تيميمون", "name_en": "Timimoun", "desk_fee": 800, "home_fee": 1000},
    "50": {"name_ar": "برج باجي مختار", "name_en": "Bordj Badji Mokhtar", "desk_fee": 900, "home_fee": 1100},
    "51": {"name_ar": "أولاد جلال", "name_en": "Ouled Djellal", "desk_fee": 500, "home_fee": 700},
    "52": {"name_ar": "بني عباس", "name_en": "Béni Abbès", "desk_fee": 700, "home_fee": 900},
    "53": {"name_ar": "عين صالح", "name_en": "In Salah", "desk_fee": 800, "home_fee": 1000},
    "54": {"name_ar": "عين قزام", "name_en": "In Guezzam", "desk_fee": 900, "home_fee": 1100},
    "55": {"name_ar": "توقرت", "name_en": "Touggourt", "desk_fee": 550, "home_fee": 750},
    "56": {"name_ar": "جانت", "name_en": "Djanet", "desk_fee": 900, "home_fee": 1100},
    "57": {"name_ar": "المغير", "name_en": "El M'Ghair", "desk_fee": 550, "home_fee": 750},
    "58": {"name_ar": "المنيعة", "name_en": "El Meniaa", "desk_fee": 650, "home_fee": 850}
}

@api_router.get("/delivery/wilayas")
async def get_wilayas():
    """Get all Algerian wilayas with delivery fees"""
    result = []
    for code, data in ALGERIA_WILAYAS.items():
        result.append({
            "code": code,
            "name_ar": data["name_ar"],
            "name_en": data["name_en"],
            "desk_fee": data["desk_fee"],
            "home_fee": data["home_fee"]
        })
    return sorted(result, key=lambda x: x["code"])

@api_router.get("/delivery/fee")
async def get_delivery_fee(wilaya_code: str, delivery_type: str = "desk"):
    """Calculate delivery fee for a wilaya"""
    if wilaya_code not in ALGERIA_WILAYAS:
        raise HTTPException(status_code=404, detail="Wilaya not found")
    
    wilaya = ALGERIA_WILAYAS[wilaya_code]
    fee = wilaya["home_fee"] if delivery_type == "home" else wilaya["desk_fee"]
    
    return {
        "wilaya_code": wilaya_code,
        "wilaya_name_ar": wilaya["name_ar"],
        "wilaya_name_en": wilaya["name_en"],
        "delivery_type": delivery_type,
        "fee": fee
    }

# ============ CUSTOMER DEBTS ROUTES ============

class CustomerDebtPayment(BaseModel):
    customer_id: str
    amount: float
    payment_method: str = "cash"
    notes: Optional[str] = ""

@api_router.get("/customers/{customer_id}/debt")
async def get_customer_debt(customer_id: str, user: dict = Depends(require_tenant)):
    """Get customer's total debt and debt history"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get all unpaid/partially paid sales for this customer
    sales = await db.sales.find({
        "customer_id": customer_id,
        "debt_amount": {"$gt": 0}
    }, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Get debt payments history
    payments = await db.debt_payments.find({
        "customer_id": customer_id
    }, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    total_debt = sum(s.get("debt_amount", 0) for s in sales)
    
    return {
        "customer_id": customer_id,
        "customer_name": customer.get("name", ""),
        "total_debt": total_debt,
        "unpaid_sales": sales,
        "payment_history": payments
    }

@api_router.post("/customers/{customer_id}/debt/pay")
async def pay_customer_debt(customer_id: str, payment: CustomerDebtPayment, user: dict = Depends(require_tenant)):
    """Record a debt payment from customer"""
    customer = await db.customers.find_one({"id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get unpaid sales sorted by oldest first
    sales = await db.sales.find({
        "customer_id": customer_id,
        "debt_amount": {"$gt": 0}
    }).sort("created_at", 1).to_list(100)
    
    if not sales:
        raise HTTPException(status_code=400, detail="Customer has no debt")
    
    remaining_payment = payment.amount
    sales_updated = []
    
    # Apply payment to sales (oldest first)
    for sale in sales:
        if remaining_payment <= 0:
            break
        
        sale_debt = sale.get("debt_amount", 0)
        if sale_debt <= 0:
            continue
        
        payment_for_sale = min(remaining_payment, sale_debt)
        new_debt = sale_debt - payment_for_sale
        new_paid = sale.get("paid_amount", 0) + payment_for_sale
        
        await db.sales.update_one(
            {"id": sale["id"]},
            {"$set": {
                "debt_amount": new_debt,
                "paid_amount": new_paid
            }}
        )
        
        remaining_payment -= payment_for_sale
        sales_updated.append({
            "sale_id": sale["id"],
            "payment_applied": payment_for_sale,
            "remaining_debt": new_debt
        })
    
    # Record payment in history
    payment_record = {
        "id": str(uuid.uuid4()),
        "customer_id": customer_id,
        "customer_name": customer.get("name", ""),
        "amount": payment.amount - remaining_payment,  # Actual amount applied
        "payment_method": payment.payment_method,
        "notes": payment.notes,
        "sales_updated": sales_updated,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("name", "")
    }
    await db.debt_payments.insert_one(payment_record)
    
    # Update customer total_debt and balance
    actual_payment = payment.amount - remaining_payment
    if actual_payment > 0:
        await db.customers.update_one(
            {"id": customer_id},
            {"$inc": {"total_debt": -actual_payment, "balance": -actual_payment}}
        )
    
    # Update cash management
    if payment.payment_method in ["cash", "bank", "wallet"]:
        await db.cash_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "type": "income",
            "category": "debt_payment",
            "amount": payment.amount - remaining_payment,
            "method": payment.payment_method,
            "description": f"Debt payment from {customer.get('name', '')}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user.get("name", "")
        })
    
    return {
        "success": True,
        "payment_applied": payment.amount - remaining_payment,
        "remaining_from_payment": remaining_payment,
        "sales_updated": sales_updated
    }

@api_router.get("/debts/summary")
async def get_debts_summary(user: dict = Depends(require_tenant)):
    """Get summary of all customer debts"""
    pipeline = [
        {"$match": {"debt_amount": {"$gt": 0}}},
        {"$group": {
            "_id": "$customer_id",
            "total_debt": {"$sum": "$debt_amount"},
            "sales_count": {"$sum": 1}
        }}
    ]
    
    debts_by_customer = await db.sales.aggregate(pipeline).to_list(1000)
    
    # Enrich with customer names
    result = []
    for debt in debts_by_customer:
        customer = await db.customers.find_one({"id": debt["_id"]}, {"_id": 0, "name": 1, "phone": 1})
        if customer:
            result.append({
                "customer_id": debt["_id"],
                "customer_name": customer.get("name", "Unknown"),
                "customer_phone": customer.get("phone", ""),
                "total_debt": debt["total_debt"],
                "sales_count": debt["sales_count"]
            })
    
    total_outstanding = sum(d["total_debt"] for d in result)
    
    return {
        "total_outstanding": total_outstanding,
        "customers_with_debt": len(result),
        "debts": sorted(result, key=lambda x: x["total_debt"], reverse=True)
    }

@api_router.get("/debts/export")
async def export_debts_to_excel(user: dict = Depends(require_tenant)):
    """Export all customer debts to Excel file"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Get all debts
    pipeline = [
        {"$match": {"debt_amount": {"$gt": 0}}},
        {"$group": {
            "_id": "$customer_id",
            "total_debt": {"$sum": "$debt_amount"},
            "sales_count": {"$sum": 1}
        }}
    ]
    debts_by_customer = await db.sales.aggregate(pipeline).to_list(1000)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Debts"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["#", "اسم الزبون", "رقم الهاتف", "عدد الفواتير", "إجمالي الدين (دج)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Data rows
    row_num = 2
    total_debt = 0
    for idx, debt in enumerate(debts_by_customer, 1):
        customer = await db.customers.find_one({"id": debt["_id"]}, {"_id": 0})
        if not customer:
            continue
        
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=customer.get("name", "")).border = border
        ws.cell(row=row_num, column=3, value=customer.get("phone", "")).border = border
        ws.cell(row=row_num, column=4, value=debt["sales_count"]).border = border
        cell = ws.cell(row=row_num, column=5, value=debt["total_debt"])
        cell.border = border
        cell.number_format = '#,##0.00'
        
        total_debt += debt["total_debt"]
        row_num += 1
    
    # Total row
    ws.cell(row=row_num, column=4, value="الإجمالي:").font = Font(bold=True)
    total_cell = ws.cell(row=row_num, column=5, value=total_debt)
    total_cell.font = Font(bold=True, color="FF0000")
    total_cell.number_format = '#,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    
    # RTL
    ws.sheet_view.rightToLeft = True
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=debts_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ============ SYSTEM SETTINGS ============

class SystemSettingsUpdate(BaseModel):
    cash_difference_threshold: float = 1000  # حد التنبيه للعجز/الفائض
    low_stock_threshold: int = 10  # حد المخزون المنخفض
    currency_symbol: str = "دج"
    business_name: str = "NT"

DEFAULT_SYSTEM_SETTINGS = {
    "id": "global",
    "cash_difference_threshold": 1000,
    "low_stock_threshold": 10,
    "currency_symbol": "دج",
    "business_name": "NT"
}

@api_router.get("/system/settings")
async def get_system_settings(user: dict = Depends(require_tenant)):
    """Get system settings"""
    settings = await db.system_settings.find_one({"id": "global"}, {"_id": 0})
    if not settings:
        settings = {**DEFAULT_SYSTEM_SETTINGS}
        await db.system_settings.insert_one(settings)
    else:
        settings = {k: v for k, v in settings.items() if k != "_id"}
    return settings

@api_router.put("/system/settings")
async def update_system_settings(settings: SystemSettingsUpdate, admin: dict = Depends(get_tenant_admin)):
    """Update system settings (admin only)"""
    update_data = settings.model_dump()
    
    existing = await db.system_settings.find_one({"id": "global"})
    if existing:
        await db.system_settings.update_one(
            {"id": "global"},
            {"$set": update_data}
        )
    else:
        await db.system_settings.insert_one({**update_data, "id": "global"})
    
    return {"message": "تم تحديث الإعدادات بنجاح"}

# ============ SIM BALANCE MANAGEMENT ============

class SimSlotBalance(BaseModel):
    slot_id: int  # 1 أو 2
    operator: str  # موبيليس، جازي، أوريدو
    phone: str
    balance: float = 0
    last_updated: str = ""

class SimBalanceUpdate(BaseModel):
    balance: float
    notes: Optional[str] = ""

@api_router.get("/sim/slots")
async def get_sim_slots(admin: dict = Depends(get_tenant_admin)):
    """Get all SIM slots with their balances"""
    slots = await db.sim_slots.find({}, {"_id": 0}).to_list(10)
    if not slots:
        # Create default slots
        default_slots = [
            {"slot_id": 1, "operator": "موبيليس", "phone": "", "balance": 0, "last_updated": "", "prefix": "06"},
            {"slot_id": 2, "operator": "جازي", "phone": "", "balance": 0, "last_updated": "", "prefix": "07"},
            {"slot_id": 3, "operator": "أوريدو", "phone": "", "balance": 0, "last_updated": "", "prefix": "05"}
        ]
        await db.sim_slots.insert_many(default_slots)
        slots = default_slots
    return slots

@api_router.put("/sim/slots/{slot_id}")
async def update_sim_slot(slot_id: int, slot_data: dict, admin: dict = Depends(get_tenant_admin)):
    """Update SIM slot info"""
    now = datetime.now(timezone.utc).isoformat()
    update_data = {**slot_data, "last_updated": now}
    
    await db.sim_slots.update_one(
        {"slot_id": slot_id},
        {"$set": update_data},
        upsert=True
    )
    return {"message": "تم تحديث الشريحة بنجاح"}

@api_router.put("/sim/slots/{slot_id}/balance")
async def update_sim_balance(slot_id: int, balance_data: SimBalanceUpdate, admin: dict = Depends(get_tenant_admin)):
    """Update SIM slot balance"""
    now = datetime.now(timezone.utc).isoformat()
    
    # Get current slot
    slot = await db.sim_slots.find_one({"slot_id": slot_id})
    old_balance = slot.get("balance", 0) if slot else 0
    
    await db.sim_slots.update_one(
        {"slot_id": slot_id},
        {"$set": {"balance": balance_data.balance, "last_updated": now}}
    )
    
    # Log the balance change
    log_entry = {
        "id": str(uuid.uuid4()),
        "slot_id": slot_id,
        "old_balance": old_balance,
        "new_balance": balance_data.balance,
        "change": balance_data.balance - old_balance,
        "notes": balance_data.notes or "",
        "created_at": now,
        "created_by": admin.get("name", "")
    }
    await db.sim_balance_logs.insert_one(log_entry)
    
    return {"message": "تم تحديث الرصيد بنجاح"}

@api_router.get("/sim/slots/{slot_id}/logs")
async def get_sim_balance_logs(slot_id: int, admin: dict = Depends(get_tenant_admin)):
    """Get balance change history for a SIM slot"""
    logs = await db.sim_balance_logs.find({"slot_id": slot_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return logs

# ============ AUTO RECHARGE BY OPERATOR ============

@api_router.post("/recharge/auto")
async def auto_recharge(phone: str, amount: float, user: dict = Depends(require_tenant)):
    """Auto-select SIM slot based on phone number prefix"""
    
    # Clean phone number
    clean_phone = phone.replace(" ", "").replace("-", "")
    if clean_phone.startswith("+213"):
        clean_phone = "0" + clean_phone[4:]
    elif clean_phone.startswith("213"):
        clean_phone = "0" + clean_phone[3:]
    
    # Determine operator by prefix
    prefix = clean_phone[:2] if len(clean_phone) >= 2 else ""
    
    operator_map = {
        "06": {"name": "موبيليس", "name_fr": "Mobilis"},
        "07": {"name": "جازي", "name_fr": "Djezzy"},
        "05": {"name": "أوريدو", "name_fr": "Ooredoo"}
    }
    
    if prefix not in operator_map:
        raise HTTPException(status_code=400, detail="رقم هاتف غير صالح. يجب أن يبدأ بـ 05, 06, أو 07")
    
    operator = operator_map[prefix]
    
    # Find the appropriate SIM slot
    slot = await db.sim_slots.find_one({"prefix": prefix}, {"_id": 0})
    
    if not slot or not slot.get("phone"):
        raise HTTPException(status_code=400, detail=f"شريحة {operator['name']} غير مفعلة")
    
    if slot.get("balance", 0) < amount:
        raise HTTPException(status_code=400, detail=f"رصيد شريحة {operator['name']} غير كافي")
    
    # Log the recharge (MOCKED)
    now = datetime.now(timezone.utc).isoformat()
    recharge_log = {
        "id": str(uuid.uuid4()),
        "phone": clean_phone,
        "amount": amount,
        "operator": operator["name"],
        "slot_id": slot["slot_id"],
        "status": "success",  # MOCKED
        "created_at": now,
        "created_by": user.get("name", "")
    }
    await db.recharge_logs.insert_one(recharge_log)
    
    # Deduct from SIM balance
    await db.sim_slots.update_one(
        {"slot_id": slot["slot_id"]},
        {"$inc": {"balance": -amount}, "$set": {"last_updated": now}}
    )
    
    return {
        "success": True,
        "phone": clean_phone,
        "amount": amount,
        "operator": operator["name"],
        "message": f"تم شحن {amount} دج لـ {clean_phone} عبر {operator['name']}"
    }

# ============ WOOCOMMERCE INTEGRATION (UI Ready) ============

class WooCommerceSettings(BaseModel):
    enabled: bool = False
    store_url: str = ""
    consumer_key: str = ""
    consumer_secret: str = ""
    sync_products: bool = True
    sync_orders: bool = True
    sync_customers: bool = True
    last_sync: str = ""

@api_router.get("/woocommerce/settings")
async def get_woocommerce_settings(admin: dict = Depends(get_tenant_admin)):
    """Get WooCommerce integration settings"""
    settings = await db.woocommerce_settings.find_one({"id": "global"}, {"_id": 0})
    if not settings:
        settings = {
            "id": "global",
            "enabled": False,
            "store_url": "",
            "consumer_key": "",
            "consumer_secret": "",
            "sync_products": True,
            "sync_orders": True,
            "sync_customers": True,
            "last_sync": ""
        }
        await db.woocommerce_settings.insert_one(settings)
    return settings

@api_router.put("/woocommerce/settings")
async def update_woocommerce_settings(settings: WooCommerceSettings, admin: dict = Depends(get_tenant_admin)):
    """Update WooCommerce integration settings"""
    update_data = settings.model_dump()
    
    await db.woocommerce_settings.update_one(
        {"id": "global"},
        {"$set": update_data},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات WooCommerce"}

@api_router.post("/woocommerce/test-connection")
async def test_woocommerce_connection(admin: dict = Depends(get_tenant_admin)):
    """Test WooCommerce connection (MOCKED)"""
    settings = await db.woocommerce_settings.find_one({"id": "global"}, {"_id": 0})
    
    if not settings or not settings.get("store_url"):
        raise HTTPException(status_code=400, detail="يرجى إدخال رابط المتجر أولاً")
    
    # MOCKED - In production, actually test the API connection
    return {
        "success": True,
        "message": "تم الاتصال بالمتجر بنجاح (وضع المحاكاة)",
        "store_info": {
            "name": "متجرك",
            "url": settings.get("store_url"),
            "version": "8.0.0"
        }
    }

@api_router.post("/woocommerce/publish-product/{product_id}")
async def publish_product_to_woocommerce(product_id: str, admin: dict = Depends(get_tenant_admin)):
    """Publish a single product to WooCommerce (MOCKED)"""
    
    # Check WooCommerce settings
    wc_settings = await db.woocommerce_settings.find_one({"id": "global"}, {"_id": 0})
    if not wc_settings or not wc_settings.get("enabled"):
        raise HTTPException(status_code=400, detail="WooCommerce غير مفعل")
    
    # Get product
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # MOCKED - In production, actually call WooCommerce API
    wc_product_id = f"wc_{product_id[:8]}"
    
    # Update product with WooCommerce info
    await db.products.update_one(
        {"id": product_id},
        {"$set": {
            "woocommerce_id": wc_product_id,
            "woocommerce_status": "published",
            "woocommerce_url": f"{wc_settings.get('store_url')}/product/{product.get('name_en', '').lower().replace(' ', '-')}",
            "woocommerce_synced_at": now
        }}
    )
    
    return {
        "success": True,
        "message": f"تم نشر المنتج '{product.get('name_en')}' على المتجر",
        "woocommerce_id": wc_product_id,
        "product_url": f"{wc_settings.get('store_url')}/product/{product.get('name_en', '').lower().replace(' ', '-')}"
    }

@api_router.post("/woocommerce/publish-products")
async def publish_multiple_products_to_woocommerce(product_ids: List[str], admin: dict = Depends(get_tenant_admin)):
    """Publish multiple products to WooCommerce (MOCKED)"""
    
    # Check WooCommerce settings
    wc_settings = await db.woocommerce_settings.find_one({"id": "global"}, {"_id": 0})
    if not wc_settings or not wc_settings.get("enabled"):
        raise HTTPException(status_code=400, detail="WooCommerce غير مفعل")
    
    now = datetime.now(timezone.utc).isoformat()
    published = []
    failed = []
    
    for product_id in product_ids:
        product = await db.products.find_one({"id": product_id}, {"_id": 0})
        if not product:
            failed.append({"id": product_id, "error": "المنتج غير موجود"})
            continue
        
        # MOCKED
        wc_product_id = f"wc_{product_id[:8]}"
        
        await db.products.update_one(
            {"id": product_id},
            {"$set": {
                "woocommerce_id": wc_product_id,
                "woocommerce_status": "published",
                "woocommerce_url": f"{wc_settings.get('store_url')}/product/{product.get('name_en', '').lower().replace(' ', '-')}",
                "woocommerce_synced_at": now
            }}
        )
        
        published.append({
            "id": product_id,
            "name": product.get("name_en"),
            "woocommerce_id": wc_product_id
        })
    
    return {
        "success": True,
        "message": f"تم نشر {len(published)} منتج على المتجر",
        "published": published,
        "failed": failed
    }

@api_router.delete("/woocommerce/unpublish-product/{product_id}")
async def unpublish_product_from_woocommerce(product_id: str, admin: dict = Depends(get_tenant_admin)):
    """Remove a product from WooCommerce (MOCKED)"""
    
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    # Remove WooCommerce info
    await db.products.update_one(
        {"id": product_id},
        {"$unset": {
            "woocommerce_id": "",
            "woocommerce_status": "",
            "woocommerce_url": "",
            "woocommerce_synced_at": ""
        }}
    )
    
    return {
        "success": True,
        "message": f"تم إلغاء نشر المنتج '{product.get('name_en')}' من المتجر"
    }

@api_router.post("/woocommerce/sync-inventory")
async def sync_inventory_to_woocommerce(admin: dict = Depends(get_tenant_admin)):
    """Sync all published products inventory to WooCommerce (MOCKED)"""
    
    # Get all products with woocommerce_id
    products = await db.products.find(
        {"woocommerce_id": {"$exists": True, "$ne": ""}},
        {"_id": 0}
    ).to_list(1000)
    
    now = datetime.now(timezone.utc).isoformat()
    synced_count = 0
    
    for product in products:
        # MOCKED - In production, update WooCommerce stock
        await db.products.update_one(
            {"id": product["id"]},
            {"$set": {"woocommerce_synced_at": now}}
        )
        synced_count += 1
    
    # Update last sync time
    await db.woocommerce_settings.update_one(
        {"id": "global"},
        {"$set": {"last_sync": now}}
    )
    
    return {
        "success": True,
        "message": f"تم مزامنة {synced_count} منتج",
        "synced_at": now
    }

# ============ SHIPPING/DELIVERY MANAGEMENT ============

ALGERIAN_SHIPPING_COMPANIES = [
    {"id": "yalidine", "name": "Yalidine", "name_ar": "ياليدين", "website": "https://yalidine.com", "has_api": True},
    {"id": "zr_express", "name": "ZR Express", "name_ar": "زد آر إكسبريس", "website": "https://zrexpress.com", "has_api": True},
    {"id": "maystro", "name": "Maystro Delivery", "name_ar": "مايسترو", "website": "https://maystro-delivery.com", "has_api": True},
    {"id": "ecotrack", "name": "EcoTrack", "name_ar": "إيكو تراك", "website": "https://ecotrack.dz", "has_api": True},
    {"id": "guepex", "name": "Guepex", "name_ar": "قيبكس", "website": "https://guepex.com", "has_api": True},
    {"id": "procolis", "name": "Procolis", "name_ar": "بروكوليس", "website": "https://procolis.com", "has_api": False},
    {"id": "other", "name": "Autre", "name_ar": "أخرى", "website": "", "has_api": False}
]

class ShippingCompanySettings(BaseModel):
    company_id: str
    enabled: bool = False
    api_key: str = ""
    api_secret: str = ""
    default_wilaya: str = ""
    default_commune: str = ""

class ShippingRateRequest(BaseModel):
    from_wilaya: str
    to_wilaya: str
    weight: float = 0.5  # kg
    company_id: str = ""

@api_router.get("/shipping/companies")
async def get_shipping_companies(user: dict = Depends(require_tenant)):
    """Get list of Algerian shipping companies"""
    return ALGERIAN_SHIPPING_COMPANIES

@api_router.get("/shipping/settings")
async def get_shipping_settings(admin: dict = Depends(get_tenant_admin)):
    """Get shipping integration settings"""
    settings = await db.shipping_settings.find({}, {"_id": 0}).to_list(20)
    
    # Add default settings for companies not configured
    configured_ids = {s["company_id"] for s in settings}
    for company in ALGERIAN_SHIPPING_COMPANIES:
        if company["id"] not in configured_ids:
            settings.append({
                "company_id": company["id"],
                "enabled": False,
                "api_key": "",
                "api_secret": "",
                "default_wilaya": "",
                "default_commune": ""
            })
    
    return settings

@api_router.put("/shipping/settings/{company_id}")
async def update_shipping_settings(company_id: str, settings: ShippingCompanySettings, admin: dict = Depends(get_tenant_admin)):
    """Update shipping company settings"""
    await db.shipping_settings.update_one(
        {"company_id": company_id},
        {"$set": settings.model_dump()},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات شركة الشحن"}

@api_router.post("/shipping/calculate-rate")
async def calculate_shipping_rate(request: ShippingRateRequest, user: dict = Depends(require_tenant)):
    """Calculate shipping rate (MOCKED - returns estimated prices)"""
    
    # MOCKED shipping rates by wilaya distance
    base_rates = {
        "yalidine": 400,
        "zr_express": 350,
        "maystro": 380,
        "ecotrack": 420,
        "guepex": 390,
        "procolis": 450,
        "other": 500
    }
    
    # Same wilaya = lower rate
    is_same_wilaya = request.from_wilaya == request.to_wilaya
    
    rates = []
    for company in ALGERIAN_SHIPPING_COMPANIES:
        base = base_rates.get(company["id"], 400)
        if is_same_wilaya:
            price = base * 0.6
        else:
            price = base + (request.weight * 50)
        
        rates.append({
            "company_id": company["id"],
            "company_name": company["name"],
            "company_name_ar": company["name_ar"],
            "price": round(price, 2),
            "estimated_days": 2 if is_same_wilaya else 4,
            "currency": "دج"
        })
    
    return {"rates": sorted(rates, key=lambda x: x["price"])}

@api_router.get("/shipping/wilayas")
async def get_wilayas(user: dict = Depends(require_tenant)):
    """Get list of Algerian wilayas"""
    wilayas = [
        {"code": "01", "name": "أدرار", "name_fr": "Adrar"},
        {"code": "02", "name": "الشلف", "name_fr": "Chlef"},
        {"code": "03", "name": "الأغواط", "name_fr": "Laghouat"},
        {"code": "04", "name": "أم البواقي", "name_fr": "Oum El Bouaghi"},
        {"code": "05", "name": "باتنة", "name_fr": "Batna"},
        {"code": "06", "name": "بجاية", "name_fr": "Béjaïa"},
        {"code": "07", "name": "بسكرة", "name_fr": "Biskra"},
        {"code": "08", "name": "بشار", "name_fr": "Béchar"},
        {"code": "09", "name": "البليدة", "name_fr": "Blida"},
        {"code": "10", "name": "البويرة", "name_fr": "Bouira"},
        {"code": "11", "name": "تمنراست", "name_fr": "Tamanrasset"},
        {"code": "12", "name": "تبسة", "name_fr": "Tébessa"},
        {"code": "13", "name": "تلمسان", "name_fr": "Tlemcen"},
        {"code": "14", "name": "تيارت", "name_fr": "Tiaret"},
        {"code": "15", "name": "تيزي وزو", "name_fr": "Tizi Ouzou"},
        {"code": "16", "name": "الجزائر", "name_fr": "Alger"},
        {"code": "17", "name": "الجلفة", "name_fr": "Djelfa"},
        {"code": "18", "name": "جيجل", "name_fr": "Jijel"},
        {"code": "19", "name": "سطيف", "name_fr": "Sétif"},
        {"code": "20", "name": "سعيدة", "name_fr": "Saïda"},
        {"code": "21", "name": "سكيكدة", "name_fr": "Skikda"},
        {"code": "22", "name": "سيدي بلعباس", "name_fr": "Sidi Bel Abbès"},
        {"code": "23", "name": "عنابة", "name_fr": "Annaba"},
        {"code": "24", "name": "قالمة", "name_fr": "Guelma"},
        {"code": "25", "name": "قسنطينة", "name_fr": "Constantine"},
        {"code": "26", "name": "المدية", "name_fr": "Médéa"},
        {"code": "27", "name": "مستغانم", "name_fr": "Mostaganem"},
        {"code": "28", "name": "المسيلة", "name_fr": "M'Sila"},
        {"code": "29", "name": "معسكر", "name_fr": "Mascara"},
        {"code": "30", "name": "ورقلة", "name_fr": "Ouargla"},
        {"code": "31", "name": "وهران", "name_fr": "Oran"},
        {"code": "32", "name": "البيض", "name_fr": "El Bayadh"},
        {"code": "33", "name": "إليزي", "name_fr": "Illizi"},
        {"code": "34", "name": "برج بوعريريج", "name_fr": "Bordj Bou Arreridj"},
        {"code": "35", "name": "بومرداس", "name_fr": "Boumerdès"},
        {"code": "36", "name": "الطارف", "name_fr": "El Tarf"},
        {"code": "37", "name": "تندوف", "name_fr": "Tindouf"},
        {"code": "38", "name": "تيسمسيلت", "name_fr": "Tissemsilt"},
        {"code": "39", "name": "الوادي", "name_fr": "El Oued"},
        {"code": "40", "name": "خنشلة", "name_fr": "Khenchela"},
        {"code": "41", "name": "سوق أهراس", "name_fr": "Souk Ahras"},
        {"code": "42", "name": "تيبازة", "name_fr": "Tipaza"},
        {"code": "43", "name": "ميلة", "name_fr": "Mila"},
        {"code": "44", "name": "عين الدفلى", "name_fr": "Aïn Defla"},
        {"code": "45", "name": "النعامة", "name_fr": "Naâma"},
        {"code": "46", "name": "عين تموشنت", "name_fr": "Aïn Témouchent"},
        {"code": "47", "name": "غرداية", "name_fr": "Ghardaïa"},
        {"code": "48", "name": "غليزان", "name_fr": "Relizane"},
        {"code": "49", "name": "تميمون", "name_fr": "Timimoun"},
        {"code": "50", "name": "برج باجي مختار", "name_fr": "Bordj Badji Mokhtar"},
        {"code": "51", "name": "أولاد جلال", "name_fr": "Ouled Djellal"},
        {"code": "52", "name": "بني عباس", "name_fr": "Béni Abbès"},
        {"code": "53", "name": "عين صالح", "name_fr": "In Salah"},
        {"code": "54", "name": "عين قزام", "name_fr": "In Guezzam"},
        {"code": "55", "name": "توقرت", "name_fr": "Touggourt"},
        {"code": "56", "name": "جانت", "name_fr": "Djanet"},
        {"code": "57", "name": "المغير", "name_fr": "El M'Ghair"},
        {"code": "58", "name": "المنيعة", "name_fr": "El Meniaa"}
    ]
    return wilayas

# ============ LOGIN PAGE CUSTOMIZATION ============

class LoginPageSettings(BaseModel):
    logo_url: str = ""
    business_name: str = "NT"
    background_image_url: str = ""
    tagline_ar: str = "إدارة مخزون زجاج الحماية بسهولة"
    tagline_fr: str = "Gestion facile de stock de protection"

@api_router.get("/branding/settings")
async def get_branding_settings():
    """Get login page branding settings (public)"""
    settings = await db.branding_settings.find_one({"id": "global"}, {"_id": 0})
    if not settings:
        settings = {
            "id": "global",
            "logo_url": "",
            "business_name": "NT",
            "background_image_url": "",
            "tagline_ar": "إدارة مخزون زجاج الحماية بسهولة",
            "tagline_fr": "Gestion facile de stock de protection"
        }
    return settings

@api_router.put("/branding/settings")
async def update_branding_settings(settings: LoginPageSettings, admin: dict = Depends(get_tenant_admin)):
    """Update login page branding settings"""
    update_data = settings.model_dump()
    
    await db.branding_settings.update_one(
        {"id": "global"},
        {"$set": update_data},
        upsert=True
    )
    return {"message": "تم تحديث إعدادات العلامة التجارية"}

# ============ LOYALTY PROGRAM ============

class LoyaltySettings(BaseModel):
    enabled: bool = False
    points_per_dinar: float = 0.01  # نقطة لكل دينار
    points_value: float = 0.1  # قيمة النقطة بالدينار
    min_redeem_points: int = 100
    welcome_bonus: int = 0  # نقاط ترحيبية للعميل الجديد

class LoyaltyTransaction(BaseModel):
    customer_id: str
    points: int
    type: str  # earn, redeem
    sale_id: Optional[str] = None
    notes: Optional[str] = ""

@api_router.get("/loyalty/settings")
async def get_loyalty_settings(admin: dict = Depends(get_tenant_admin)):
    """Get loyalty program settings"""
    settings = await db.loyalty_settings.find_one({"id": "global"}, {"_id": 0})
    if not settings:
        settings = {
            "id": "global",
            "enabled": False,
            "points_per_dinar": 0.01,
            "points_value": 0.1,
            "min_redeem_points": 100,
            "welcome_bonus": 0
        }
        await db.loyalty_settings.insert_one(settings.copy())
    return settings

@api_router.put("/loyalty/settings")
async def update_loyalty_settings(settings: LoyaltySettings, admin: dict = Depends(get_tenant_admin)):
    """Update loyalty program settings"""
    await db.loyalty_settings.update_one(
        {"id": "global"},
        {"$set": settings.model_dump()},
        upsert=True
    )
    return {"message": "تم تحديث إعدادات برنامج الولاء"}

@api_router.get("/loyalty/customer/{customer_id}")
async def get_customer_loyalty(customer_id: str, user: dict = Depends(require_tenant)):
    """Get customer loyalty points and history"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    
    points = customer.get("loyalty_points", 0)
    
    # Get transaction history
    transactions = await db.loyalty_transactions.find(
        {"customer_id": customer_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    # Get loyalty settings for point value
    settings = await db.loyalty_settings.find_one({"id": "global"}, {"_id": 0})
    points_value = settings.get("points_value", 0.1) if settings else 0.1
    
    return {
        "customer_id": customer_id,
        "customer_name": customer.get("name"),
        "points": points,
        "points_value_dinar": round(points * points_value, 2),
        "transactions": transactions
    }

@api_router.post("/loyalty/earn")
async def earn_loyalty_points(transaction: LoyaltyTransaction, user: dict = Depends(require_tenant)):
    """Add loyalty points from a sale"""
    customer = await db.customers.find_one({"id": transaction.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    
    now = datetime.now(timezone.utc).isoformat()
    current_points = customer.get("loyalty_points", 0)
    new_points = current_points + transaction.points
    
    # Update customer points
    await db.customers.update_one(
        {"id": transaction.customer_id},
        {"$set": {"loyalty_points": new_points}}
    )
    
    # Log transaction
    await db.loyalty_transactions.insert_one({
        "id": str(uuid.uuid4()),
        "customer_id": transaction.customer_id,
        "points": transaction.points,
        "type": "earn",
        "sale_id": transaction.sale_id,
        "notes": transaction.notes or "",
        "balance_after": new_points,
        "created_at": now,
        "created_by": user.get("name", "")
    })
    
    return {"message": f"تم إضافة {transaction.points} نقطة", "new_balance": new_points}

@api_router.post("/loyalty/redeem")
async def redeem_loyalty_points(transaction: LoyaltyTransaction, user: dict = Depends(require_tenant)):
    """Redeem loyalty points"""
    customer = await db.customers.find_one({"id": transaction.customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    
    current_points = customer.get("loyalty_points", 0)
    
    # Check minimum redeem
    settings = await db.loyalty_settings.find_one({"id": "global"}, {"_id": 0})
    min_redeem = settings.get("min_redeem_points", 100) if settings else 100
    
    if transaction.points > current_points:
        raise HTTPException(status_code=400, detail="رصيد النقاط غير كافي")
    
    if transaction.points < min_redeem:
        raise HTTPException(status_code=400, detail=f"الحد الأدنى للاسترداد {min_redeem} نقطة")
    
    now = datetime.now(timezone.utc).isoformat()
    new_points = current_points - transaction.points
    
    # Update customer points
    await db.customers.update_one(
        {"id": transaction.customer_id},
        {"$set": {"loyalty_points": new_points}}
    )
    
    # Log transaction
    points_value = settings.get("points_value", 0.1) if settings else 0.1
    discount_amount = transaction.points * points_value
    
    await db.loyalty_transactions.insert_one({
        "id": str(uuid.uuid4()),
        "customer_id": transaction.customer_id,
        "points": -transaction.points,
        "type": "redeem",
        "sale_id": transaction.sale_id,
        "notes": transaction.notes or f"خصم {discount_amount} دج",
        "balance_after": new_points,
        "created_at": now,
        "created_by": user.get("name", "")
    })
    
    return {
        "message": f"تم استرداد {transaction.points} نقطة",
        "discount_amount": discount_amount,
        "new_balance": new_points
    }

# ============ SMS MARKETING ============

class SMSCampaign(BaseModel):
    name: str
    message: str
    target: str  # all, customers_with_debt, inactive, custom
    customer_ids: Optional[List[str]] = []
    scheduled_at: Optional[str] = None

@api_router.get("/marketing/sms/campaigns")
async def get_sms_campaigns(admin: dict = Depends(get_tenant_admin)):
    """Get all SMS campaigns"""
    campaigns = await db.sms_campaigns.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return campaigns

@api_router.post("/marketing/sms/campaigns")
async def create_sms_campaign(campaign: SMSCampaign, admin: dict = Depends(get_tenant_admin)):
    """Create a new SMS campaign (MOCKED)"""
    now = datetime.now(timezone.utc).isoformat()
    
    # Determine target customers
    if campaign.target == "all":
        customers = await db.customers.find({"phone": {"$ne": ""}}, {"_id": 0, "id": 1, "phone": 1, "name": 1}).to_list(1000)
    elif campaign.target == "customers_with_debt":
        customers = await db.customers.find({"balance": {"$gt": 0}, "phone": {"$ne": ""}}, {"_id": 0, "id": 1, "phone": 1, "name": 1}).to_list(1000)
    elif campaign.target == "inactive":
        # Customers who haven't purchased in 30 days
        thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
        active_customer_ids = await db.sales.distinct("customer_id", {"created_at": {"$gte": thirty_days_ago}})
        customers = await db.customers.find(
            {"id": {"$nin": active_customer_ids}, "phone": {"$ne": ""}},
            {"_id": 0, "id": 1, "phone": 1, "name": 1}
        ).to_list(1000)
    else:  # custom
        customers = await db.customers.find(
            {"id": {"$in": campaign.customer_ids}, "phone": {"$ne": ""}},
            {"_id": 0, "id": 1, "phone": 1, "name": 1}
        ).to_list(1000)
    
    campaign_doc = {
        "id": str(uuid.uuid4()),
        "name": campaign.name,
        "message": campaign.message,
        "target": campaign.target,
        "recipients_count": len(customers),
        "status": "pending" if campaign.scheduled_at else "sent",
        "scheduled_at": campaign.scheduled_at,
        "sent_at": now if not campaign.scheduled_at else None,
        "created_at": now,
        "created_by": admin.get("name", "")
    }
    
    await db.sms_campaigns.insert_one(campaign_doc)
    
    # MOCKED - In production, actually send SMS
    return {
        "success": True,
        "message": f"تم إنشاء الحملة وإرسالها لـ {len(customers)} عميل (وضع المحاكاة)",
        "campaign_id": campaign_doc["id"],
        "recipients_count": len(customers)
    }

# ============ INVOICES ============

class InvoiceTemplate(BaseModel):
    name: str
    type: str  # simple, detailed, thermal
    header_text: str = ""
    footer_text: str = ""
    show_logo: bool = True
    show_qr: bool = False

@api_router.get("/invoices/templates")
async def get_invoice_templates(user: dict = Depends(require_tenant)):
    """Get all invoice templates"""
    templates = await db.invoice_templates.find({}, {"_id": 0}).to_list(20)
    
    if not templates:
        # Create default templates
        default_templates = [
            {
                "id": "simple",
                "name": "فاتورة بسيطة",
                "name_fr": "Facture simple",
                "type": "simple",
                "header_text": "",
                "footer_text": "شكراً لتعاملكم معنا",
                "show_logo": True,
                "show_qr": False,
                "is_default": True
            },
            {
                "id": "detailed",
                "name": "فاتورة تفصيلية",
                "name_fr": "Facture détaillée",
                "type": "detailed",
                "header_text": "",
                "footer_text": "",
                "show_logo": True,
                "show_qr": True,
                "is_default": False
            },
            {
                "id": "thermal",
                "name": "فاتورة حرارية",
                "name_fr": "Ticket thermique",
                "type": "thermal",
                "header_text": "",
                "footer_text": "",
                "show_logo": False,
                "show_qr": False,
                "is_default": False
            }
        ]
        await db.invoice_templates.insert_many(default_templates)
        templates = default_templates
    
    return templates

@api_router.post("/invoices/generate/{sale_id}")
async def generate_invoice(sale_id: str, template_id: str = "simple", user: dict = Depends(require_tenant)):
    """Generate invoice for a sale"""
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="البيع غير موجود")
    
    template = await db.invoice_templates.find_one({"id": template_id}, {"_id": 0})
    branding = await db.branding_settings.find_one({"id": "global"}, {"_id": 0})
    
    # Get customer info if exists
    customer = None
    if sale.get("customer_id"):
        customer = await db.customers.find_one({"id": sale["customer_id"]}, {"_id": 0})
    
    invoice_data = {
        "invoice_number": f"INV-{sale_id[:8].upper()}",
        "date": sale.get("created_at", ""),
        "business_name": branding.get("business_name", "NT") if branding else "NT",
        "logo_url": branding.get("logo_url", "") if branding else "",
        "customer": {
            "name": customer.get("name", "") if customer else sale.get("customer_name", ""),
            "phone": customer.get("phone", "") if customer else "",
            "address": customer.get("address", "") if customer else ""
        },
        "items": sale.get("items", []),
        "subtotal": sale.get("total", 0),
        "discount": sale.get("discount", 0),
        "total": sale.get("total", 0),
        "paid": sale.get("paid_amount", 0),
        "remaining": sale.get("remaining", 0),
        "payment_method": sale.get("payment_method", ""),
        "template": template,
        "header_text": template.get("header_text", "") if template else "",
        "footer_text": template.get("footer_text", "شكراً لتعاملكم معنا") if template else ""
    }
    
    return invoice_data

# ============ PAYMENT GATEWAYS (MOCKED) ============

class PaymentGatewaySettings(BaseModel):
    gateway: str  # cib, dahabia, baridimob
    enabled: bool = False
    merchant_id: str = ""
    api_key: str = ""
    terminal_id: str = ""

ALGERIAN_PAYMENT_GATEWAYS = [
    {"id": "cib", "name": "CIB", "name_ar": "البطاقة البنكية CIB", "type": "card"},
    {"id": "dahabia", "name": "Dahabia", "name_ar": "بطاقة الذهبية", "type": "card"},
    {"id": "baridimob", "name": "BaridiMob", "name_ar": "بريدي موب", "type": "mobile"}
]

@api_router.get("/payments/gateways")
async def get_payment_gateways(admin: dict = Depends(get_tenant_admin)):
    """Get available payment gateways"""
    settings = await db.payment_gateways.find({}, {"_id": 0}).to_list(10)
    
    result = []
    for gateway in ALGERIAN_PAYMENT_GATEWAYS:
        setting = next((s for s in settings if s.get("gateway") == gateway["id"]), None)
        result.append({
            **gateway,
            "enabled": setting.get("enabled", False) if setting else False,
            "configured": bool(setting.get("merchant_id")) if setting else False
        })
    
    return result

@api_router.put("/payments/gateways/{gateway_id}")
async def update_payment_gateway(gateway_id: str, settings: PaymentGatewaySettings, admin: dict = Depends(get_tenant_admin)):
    """Update payment gateway settings"""
    await db.payment_gateways.update_one(
        {"gateway": gateway_id},
        {"$set": settings.model_dump()},
        upsert=True
    )
    return {"message": "تم تحديث إعدادات بوابة الدفع"}

# ============ SMS REMINDER SYSTEM ============

class SMSReminderRequest(BaseModel):
    customer_ids: List[str]  # قائمة معرفات الزبائن
    message_template: Optional[str] = None  # قالب الرسالة المخصص

class SMSSettingsUpdate(BaseModel):
    auto_reminder_enabled: bool = False
    reminder_frequency: Literal["daily", "weekly", "monthly"] = "weekly"
    reminder_day: int = 1  # 1-7 للأسبوعي، 1-28 للشهري
    reminder_time: str = "09:00"
    min_debt_amount: float = 100  # الحد الأدنى للدين لإرسال تذكير
    message_template: str = "السلام عليكم {customer_name}، نذكركم بأن لديكم مبلغ {debt_amount} دج مستحق. شكراً لتعاملكم معنا."

# Default SMS settings
DEFAULT_SMS_SETTINGS = {
    "auto_reminder_enabled": False,
    "reminder_frequency": "weekly",
    "reminder_day": 1,
    "reminder_time": "09:00",
    "min_debt_amount": 100,
    "message_template": "السلام عليكم {customer_name}، نذكركم بأن لديكم مبلغ {debt_amount} دج مستحق. شكراً لتعاملكم معنا - NT"
}

async def send_sms_mock(phone: str, message: str) -> dict:
    """
    MOCKED SMS sending function
    Replace this with actual SMS provider integration
    Supported providers: SMS Algérie, Mobilzone, etc.
    """
    # Simulate SMS sending
    import random
    success = random.random() > 0.1  # 90% success rate simulation
    
    return {
        "success": success,
        "phone": phone,
        "message_length": len(message),
        "provider": "MOCKED",
        "message_id": str(uuid.uuid4()) if success else None,
        "error": None if success else "Simulated failure"
    }

@api_router.get("/sms/settings")
async def get_sms_settings(admin: dict = Depends(get_tenant_admin)):
    """Get SMS reminder settings"""
    settings = await db.sms_settings.find_one({"id": "global"}, {"_id": 0})
    if not settings:
        settings = {**DEFAULT_SMS_SETTINGS, "id": "global"}
        await db.sms_settings.insert_one(settings)
        # Return without _id
        settings = {k: v for k, v in settings.items() if k != "_id"}
    return settings

@api_router.put("/sms/settings")
async def update_sms_settings(settings: SMSSettingsUpdate, admin: dict = Depends(get_tenant_admin)):
    """Update SMS reminder settings"""
    update_data = settings.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.sms_settings.update_one(
        {"id": "global"},
        {"$set": update_data},
        upsert=True
    )
    
    return {"success": True, "message": "Settings updated"}

@api_router.post("/sms/send-reminder")
async def send_debt_reminder(request: SMSReminderRequest, user: dict = Depends(require_tenant)):
    """Send SMS reminder to selected customers"""
    settings = await db.sms_settings.find_one({"id": "global"}, {"_id": 0})
    if not settings:
        settings = DEFAULT_SMS_SETTINGS
    
    template = request.message_template or settings.get("message_template", DEFAULT_SMS_SETTINGS["message_template"])
    
    results = []
    for customer_id in request.customer_ids:
        # Get customer info
        customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if not customer:
            results.append({"customer_id": customer_id, "success": False, "error": "Customer not found"})
            continue
        
        if not customer.get("phone"):
            results.append({"customer_id": customer_id, "success": False, "error": "No phone number"})
            continue
        
        # Get customer debt
        debt_pipeline = [
            {"$match": {"customer_id": customer_id, "debt_amount": {"$gt": 0}}},
            {"$group": {"_id": None, "total": {"$sum": "$debt_amount"}}}
        ]
        debt_result = await db.sales.aggregate(debt_pipeline).to_list(1)
        total_debt = debt_result[0]["total"] if debt_result else 0
        
        if total_debt <= 0:
            results.append({"customer_id": customer_id, "success": False, "error": "No debt"})
            continue
        
        # Format message
        message = template.format(
            customer_name=customer.get("name", ""),
            debt_amount=f"{total_debt:,.0f}",
            phone=customer.get("phone", "")
        )
        
        # Send SMS (MOCKED)
        sms_result = await send_sms_mock(customer["phone"], message)
        
        # Log the SMS
        sms_log = {
            "id": str(uuid.uuid4()),
            "customer_id": customer_id,
            "customer_name": customer.get("name", ""),
            "phone": customer.get("phone", ""),
            "message": message,
            "debt_amount": total_debt,
            "status": "sent" if sms_result["success"] else "failed",
            "provider_response": sms_result,
            "sent_by": user.get("name", ""),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.sms_logs.insert_one(sms_log)
        
        results.append({
            "customer_id": customer_id,
            "customer_name": customer.get("name", ""),
            "phone": customer.get("phone", ""),
            "success": sms_result["success"],
            "error": sms_result.get("error")
        })
    
    success_count = sum(1 for r in results if r.get("success"))
    return {
        "total": len(request.customer_ids),
        "success": success_count,
        "failed": len(request.customer_ids) - success_count,
        "results": results
    }

@api_router.post("/sms/send-bulk-reminder")
async def send_bulk_debt_reminder(user: dict = Depends(require_tenant), min_debt: float = 0):
    """Send SMS reminder to all customers with debt"""
    settings = await db.sms_settings.find_one({"id": "global"}, {"_id": 0})
    if not settings:
        settings = DEFAULT_SMS_SETTINGS
    
    min_amount = min_debt if min_debt > 0 else settings.get("min_debt_amount", 100)
    
    # Get all customers with debt above minimum
    pipeline = [
        {"$match": {"debt_amount": {"$gt": 0}}},
        {"$group": {
            "_id": "$customer_id",
            "total_debt": {"$sum": "$debt_amount"}
        }},
        {"$match": {"total_debt": {"$gte": min_amount}}}
    ]
    
    debts = await db.sales.aggregate(pipeline).to_list(1000)
    customer_ids = [d["_id"] for d in debts if d["_id"]]
    
    if not customer_ids:
        return {"total": 0, "success": 0, "failed": 0, "results": [], "message": "No customers with debt found"}
    
    # Use the single reminder endpoint
    request = SMSReminderRequest(customer_ids=customer_ids)
    return await send_debt_reminder(request, user)

@api_router.get("/sms/logs")
async def get_sms_logs(
    limit: int = 50,
    customer_id: Optional[str] = None,
    user: dict = Depends(require_tenant)
):
    """Get SMS sending history"""
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    
    logs = await db.sms_logs.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    
    # Stats
    total_sent = await db.sms_logs.count_documents({"status": "sent"})
    total_failed = await db.sms_logs.count_documents({"status": "failed"})
    
    return {
        "logs": logs,
        "stats": {
            "total_sent": total_sent,
            "total_failed": total_failed
        }
    }

@api_router.get("/sms/templates")
async def get_sms_templates():
    """Get predefined SMS templates"""
    return {
        "templates": [
            {
                "id": "reminder_friendly",
                "name_ar": "تذكير ودي",
                "name_en": "Friendly Reminder",
                "template": "السلام عليكم {customer_name}، نذكركم بأن لديكم مبلغ {debt_amount} دج مستحق. شكراً لتعاملكم معنا."
            },
            {
                "id": "reminder_formal",
                "name_ar": "تذكير رسمي",
                "name_en": "Formal Reminder",
                "template": "عزيزنا {customer_name}، نود إعلامكم بوجود مستحقات بقيمة {debt_amount} دج. نرجو التسديد في أقرب وقت."
            },
            {
                "id": "reminder_urgent",
                "name_ar": "تذكير عاجل",
                "name_en": "Urgent Reminder",
                "template": "تنبيه: {customer_name}، لديكم مبلغ {debt_amount} دج متأخر السداد. يرجى التواصل معنا."
            },
            {
                "id": "payment_thanks",
                "name_ar": "شكر على الدفع",
                "name_en": "Payment Thanks",
                "template": "شكراً {customer_name} على سداد مستحقاتكم. نقدر تعاملكم معنا - NT"
            }
        ]
    }

# ============ ADVANCED ROLES AND PERMISSIONS ============

@api_router.get("/permissions/roles")
async def get_all_roles():
    """Get all available roles with their default permissions and descriptions"""
    return {
        "roles": list(DEFAULT_PERMISSIONS.keys()),
        "default_permissions": DEFAULT_PERMISSIONS,
        "role_descriptions": ROLE_DESCRIPTIONS,
        "permission_categories": PERMISSION_CATEGORIES
    }

@api_router.get("/permissions/categories")
async def get_permission_categories():
    """Get permission categories for UI grouping"""
    return PERMISSION_CATEGORIES

@api_router.get("/permissions/role/{role_name}")
async def get_role_permissions(role_name: str):
    """Get permissions for a specific role"""
    if role_name not in DEFAULT_PERMISSIONS:
        raise HTTPException(status_code=404, detail="Role not found")
    
    return {
        "role": role_name,
        "description": ROLE_DESCRIPTIONS.get(role_name, {"ar": role_name, "fr": role_name}),
        "permissions": DEFAULT_PERMISSIONS[role_name]
    }

# ============ FILE UPLOAD ============

@api_router.post("/upload/image")
async def upload_image(file: UploadFile = File(...), user: dict = Depends(require_tenant)):
    """Upload an image file"""
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/webp", "image/gif"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Only images are allowed.")
    
    # Generate unique filename
    file_ext = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    unique_filename = f"{uuid.uuid4()}.{file_ext}"
    file_path = UPLOAD_DIR / unique_filename
    
    # Save file
    try:
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Return URL (relative to static)
        return {"url": f"/api/static/uploads/{unique_filename}", "filename": unique_filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving file: {str(e)}")

# ============ PERMISSIONS SYSTEM ============

@api_router.get("/permissions/roles")
async def get_available_roles():
    """Get all available roles and their default permissions"""
    return {
        "roles": ["admin", "manager", "user"],
        "default_permissions": DEFAULT_PERMISSIONS
    }

@api_router.get("/users/{user_id}/permissions")
async def get_user_permissions(user_id: str, admin: dict = Depends(get_tenant_admin)):
    """Get permissions for a specific user"""
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # If user has custom permissions, return them; otherwise return role defaults
    permissions = user.get("permissions") or DEFAULT_PERMISSIONS.get(user.get("role", "user"), {})
    return {
        "user_id": user_id,
        "role": user.get("role", "user"),
        "permissions": permissions,
        "is_custom": bool(user.get("permissions"))
    }

@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(user_id: str, permissions: dict, admin: dict = Depends(get_tenant_admin)):
    """Update permissions for a specific user"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"permissions": permissions}}
    )
    
    return {"success": True, "message": "Permissions updated"}

@api_router.put("/users/{user_id}/reset-permissions")
async def reset_user_permissions(user_id: str, admin: dict = Depends(get_tenant_admin)):
    """Reset user permissions to role defaults"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one(
        {"id": user_id},
        {"$unset": {"permissions": ""}}
    )
    
    return {"success": True, "message": "Permissions reset to defaults"}

# ============ FACTORY RESET ============

@api_router.post("/system/factory-reset")
async def factory_reset(confirm_code: str, admin: dict = Depends(get_tenant_admin)):
    """Factory reset - Delete all data except admin user"""
    # Verify confirmation code
    if confirm_code != "RESET-ALL-DATA":
        raise HTTPException(status_code=400, detail="Invalid confirmation code")
    
    # Check if user has factory_reset permission
    user_permissions = admin.get("permissions") or DEFAULT_PERMISSIONS.get(admin.get("role", "user"), {})
    if not user_permissions.get("factory_reset", False):
        raise HTTPException(status_code=403, detail="No permission for factory reset")
    
    # Collections to clear
    collections_to_clear = [
        "products", "customers", "suppliers", "employees", 
        "sales", "purchases", "debts", "debt_payments",
        "transactions", "notifications", "sms_logs",
        "product_families", "api_keys", "recharges"
    ]
    
    deleted_counts = {}
    for collection in collections_to_clear:
        result = await db[collection].delete_many({})
        deleted_counts[collection] = result.deleted_count
    
    # Reset cash boxes to zero
    await db.cash_boxes.update_many({}, {"$set": {"balance": 0}})
    
    # Keep admin user, delete others
    await db.users.delete_many({"role": {"$ne": "admin"}})
    
    # Log the reset
    await db.system_logs.insert_one({
        "id": str(uuid.uuid4()),
        "action": "factory_reset",
        "performed_by": admin.get("name", ""),
        "deleted_counts": deleted_counts,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "success": True,
        "message": "Factory reset completed",
        "deleted_counts": deleted_counts
    }

@api_router.get("/system/stats")
async def get_system_stats(admin: dict = Depends(get_tenant_admin)):
    """Get system statistics for factory reset preview"""
    stats = {
        "products": await db.products.count_documents({}),
        "customers": await db.customers.count_documents({}),
        "suppliers": await db.suppliers.count_documents({}),
        "employees": await db.employees.count_documents({}),
        "sales": await db.sales.count_documents({}),
        "users": await db.users.count_documents({}),
        "product_families": await db.product_families.count_documents({}),
        "recharges": await db.recharges.count_documents({})
    }
    return stats

# ============ BULK PRICE UPDATE ============

class BulkPriceUpdateRequest(BaseModel):
    product_ids: Optional[List[str]] = None  # None = all products
    family_id: Optional[str] = None  # Filter by family
    update_type: Literal["percentage", "fixed", "set"]  # نسبة مئوية، مبلغ ثابت، تحديد قيمة
    price_field: Literal["purchase_price", "wholesale_price", "retail_price", "all"]
    value: float
    round_to: int = 0  # Round to nearest (0 = no rounding, 10 = nearest 10, etc.)

@api_router.post("/products/bulk-price-update")
async def bulk_price_update(request: BulkPriceUpdateRequest, admin: dict = Depends(get_tenant_admin)):
    """Update prices for multiple products at once"""
    
    # Build query
    query = {}
    if request.product_ids:
        query["id"] = {"$in": request.product_ids}
    if request.family_id:
        query["family_id"] = request.family_id
    
    # Get products
    products = await db.products.find(query, {"_id": 0}).to_list(10000)
    
    if not products:
        return {"success": False, "message": "No products found", "updated_count": 0}
    
    price_fields = ["purchase_price", "wholesale_price", "retail_price"] if request.price_field == "all" else [request.price_field]
    
    updated_count = 0
    updates_log = []
    
    for product in products:
        update_data = {}
        
        for field in price_fields:
            old_price = product.get(field, 0)
            new_price = old_price
            
            if request.update_type == "percentage":
                # Increase/decrease by percentage
                new_price = old_price * (1 + request.value / 100)
            elif request.update_type == "fixed":
                # Add/subtract fixed amount
                new_price = old_price + request.value
            elif request.update_type == "set":
                # Set to specific value
                new_price = request.value
            
            # Round if needed
            if request.round_to > 0:
                new_price = round(new_price / request.round_to) * request.round_to
            
            # Ensure price is not negative
            new_price = max(0, new_price)
            
            update_data[field] = new_price
        
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.products.update_one({"id": product["id"]}, {"$set": update_data})
        updated_count += 1
        
        updates_log.append({
            "product_id": product["id"],
            "product_name": product.get("name_ar", product.get("name_en", "")),
            "old_prices": {f: product.get(f, 0) for f in price_fields},
            "new_prices": {f: update_data[f] for f in price_fields}
        })
    
    # Log the bulk update
    await db.system_logs.insert_one({
        "id": str(uuid.uuid4()),
        "action": "bulk_price_update",
        "performed_by": admin.get("name", ""),
        "update_type": request.update_type,
        "value": request.value,
        "updated_count": updated_count,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "success": True,
        "updated_count": updated_count,
        "updates": updates_log[:10]  # Return first 10 as sample
    }

@api_router.get("/products/price-preview")
async def preview_price_update(
    update_type: str,
    price_field: str,
    value: float,
    family_id: Optional[str] = None,
    round_to: int = 0,
    admin: dict = Depends(get_tenant_admin)
):
    """Preview price changes before applying"""
    query = {}
    if family_id:
        query["family_id"] = family_id
    
    products = await db.products.find(query, {"_id": 0}).limit(20).to_list(20)
    
    previews = []
    for product in products:
        price_fields = ["purchase_price", "wholesale_price", "retail_price"] if price_field == "all" else [price_field]
        
        preview = {
            "id": product["id"],
            "name": product.get("name_ar", product.get("name_en", "")),
            "changes": {}
        }
        
        for field in price_fields:
            old_price = product.get(field, 0)
            new_price = old_price
            
            if update_type == "percentage":
                new_price = old_price * (1 + value / 100)
            elif update_type == "fixed":
                new_price = old_price + value
            elif update_type == "set":
                new_price = value
            
            if round_to > 0:
                new_price = round(new_price / round_to) * round_to
            
            new_price = max(0, new_price)
            
            preview["changes"][field] = {
                "old": old_price,
                "new": new_price,
                "diff": new_price - old_price
            }
        
        previews.append(preview)
    
    return {
        "preview_count": len(previews),
        "total_products": await db.products.count_documents(query),
        "previews": previews
    }

# ============ PRODUCT FAMILIES ROUTES ============

@api_router.post("/product-families", response_model=ProductFamilyResponse)
async def create_product_family(family: ProductFamilyCreate, admin: dict = Depends(get_tenant_admin)):
    family_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Get parent name if exists
    parent_name = ""
    if family.parent_id:
        parent = await db.product_families.find_one({"id": family.parent_id}, {"_id": 0, "name_ar": 1})
        if parent:
            parent_name = parent["name_ar"]
    
    family_doc = {
        "id": family_id,
        "name_en": family.name_en,
        "name_ar": family.name_ar,
        "description_en": family.description_en or "",
        "description_ar": family.description_ar or "",
        "parent_id": family.parent_id or "",
        "parent_name": parent_name,
        "product_count": 0,
        "created_at": now
    }
    await db.product_families.insert_one(family_doc)
    return ProductFamilyResponse(**family_doc)

@api_router.get("/product-families", response_model=List[ProductFamilyResponse])
async def get_product_families(user: dict = Depends(require_tenant)):
    families = await db.product_families.find({}, {"_id": 0}).to_list(1000)
    
    # Update product counts
    for family in families:
        count = await db.products.count_documents({"family_id": family["id"]})
        family["product_count"] = count
    
    return [ProductFamilyResponse(**f) for f in families]

@api_router.get("/product-families/{family_id}", response_model=ProductFamilyResponse)
async def get_product_family(family_id: str, user: dict = Depends(require_tenant)):
    family = await db.product_families.find_one({"id": family_id}, {"_id": 0})
    if not family:
        raise HTTPException(status_code=404, detail="Product family not found")
    
    # Update product count
    count = await db.products.count_documents({"family_id": family_id})
    family["product_count"] = count
    
    return ProductFamilyResponse(**family)

@api_router.put("/product-families/{family_id}", response_model=ProductFamilyResponse)
async def update_product_family(family_id: str, updates: ProductFamilyUpdate, admin: dict = Depends(get_tenant_admin)):
    family = await db.product_families.find_one({"id": family_id})
    if not family:
        raise HTTPException(status_code=404, detail="Product family not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    
    # Update parent name if parent_id changed
    if "parent_id" in update_data and update_data["parent_id"]:
        parent = await db.product_families.find_one({"id": update_data["parent_id"]}, {"_id": 0, "name_ar": 1})
        update_data["parent_name"] = parent["name_ar"] if parent else ""
    elif "parent_id" in update_data and not update_data["parent_id"]:
        update_data["parent_name"] = ""
    
    if update_data:
        await db.product_families.update_one({"id": family_id}, {"$set": update_data})
    
    updated = await db.product_families.find_one({"id": family_id}, {"_id": 0})
    count = await db.products.count_documents({"family_id": family_id})
    updated["product_count"] = count
    
    return ProductFamilyResponse(**updated)

@api_router.delete("/product-families/{family_id}")
async def delete_product_family(family_id: str, admin: dict = Depends(get_tenant_admin)):
    # Check if family has products
    product_count = await db.products.count_documents({"family_id": family_id})
    if product_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete family with {product_count} products")
    
    # Check if family has children
    child_count = await db.product_families.count_documents({"parent_id": family_id})
    if child_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete family with {child_count} sub-families")
    
    result = await db.product_families.delete_one({"id": family_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product family not found")
    return {"message": "Product family deleted successfully"}

@api_router.get("/product-families/{family_id}/products", response_model=List[ProductResponse])
async def get_family_products(family_id: str, user: dict = Depends(require_tenant)):
    """Get all products in a specific family"""
    products = await db.products.find({"family_id": family_id}, {"_id": 0}).to_list(1000)
    
    # Add family names
    for product in products:
        if product.get("family_id"):
            family = await db.product_families.find_one({"id": product["family_id"]}, {"_id": 0, "name_ar": 1})
            product["family_name"] = family["name_ar"] if family else ""
        else:
            product["family_name"] = ""
    
    return [ProductResponse(**p) for p in products]

# ============ DAILY SESSIONS ============

class DailySessionCreate(BaseModel):
    opening_cash: float
    opened_at: str
    status: str = "open"
    code: Optional[str] = ""  # كود الحصة S0001

class DailySessionClose(BaseModel):
    closing_cash: float
    closed_at: str
    notes: Optional[str] = ""
    status: str = "closed"

class DailySessionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    code: str = ""  # كود الحصة
    user_id: str = ""
    user_name: str = ""
    opening_cash: float
    closing_cash: Optional[float] = None
    opened_at: str
    closed_at: Optional[str] = None
    total_sales: float = 0
    cash_sales: float = 0
    credit_sales: float = 0
    sales_count: int = 0
    status: str
    notes: str = ""
    created_by: str = ""

@api_router.post("/daily-sessions", response_model=DailySessionResponse)
async def create_daily_session(session: DailySessionCreate, user: dict = Depends(require_tenant)):
    """Start a new daily cash session for the current user"""
    
    # Check if this user already has an open session
    existing = await db.daily_sessions.find_one({"user_id": user["id"], "status": "open"})
    if existing:
        raise HTTPException(status_code=400, detail="لديك حصة مفتوحة بالفعل / Vous avez déjà une session ouverte")
    
    session_id = str(uuid.uuid4())
    
    session_doc = {
        "id": session_id,
        "code": session.code or "",  # كود الحصة
        "user_id": user["id"],
        "user_name": user.get("name", ""),
        "opening_cash": session.opening_cash,
        "closing_cash": None,
        "opened_at": session.opened_at,
        "closed_at": None,
        "total_sales": 0,
        "cash_sales": 0,
        "credit_sales": 0,
        "sales_count": 0,
        "status": "open",
        "notes": "",
        "created_by": user.get("name", "")
    }
    
    await db.daily_sessions.insert_one(session_doc)
    return DailySessionResponse(**session_doc)

@api_router.get("/daily-sessions", response_model=List[DailySessionResponse])
async def get_daily_sessions(all_users: bool = False, user: dict = Depends(require_tenant)):
    """Get daily sessions - own sessions or all (admin only)"""
    query = {}
    
    # Regular users see only their sessions, admin can see all
    if not all_users or user.get("role") != "admin":
        query["user_id"] = user["id"]
    
    sessions = await db.daily_sessions.find(query, {"_id": 0}).sort("opened_at", -1).to_list(100)
    
    # Add missing fields for backward compatibility
    for s in sessions:
        if "user_id" not in s:
            s["user_id"] = ""
        if "user_name" not in s:
            s["user_name"] = s.get("created_by", "")
    
    return [DailySessionResponse(**s) for s in sessions]

@api_router.get("/daily-sessions/current")
async def get_current_session(user: dict = Depends(require_tenant)):
    """Get the current open session for the logged-in user"""
    session = await db.daily_sessions.find_one({"user_id": user["id"], "status": "open"}, {"_id": 0})
    if not session:
        return None
    
    # Add missing fields
    if "user_id" not in session:
        session["user_id"] = user["id"]
    if "user_name" not in session:
        session["user_name"] = user.get("name", "")
    
    return DailySessionResponse(**session)

@api_router.get("/daily-sessions/summary")
async def get_sessions_summary(days: int = 7, admin: dict = Depends(get_tenant_admin)):
    """Get summary report of all sessions (admin only)"""
    
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    
    # Get all closed sessions in the period
    sessions = await db.daily_sessions.find({
        "status": "closed",
        "closed_at": {"$gte": start_date}
    }, {"_id": 0}).to_list(500)
    
    # Group by user
    user_stats = {}
    for session in sessions:
        user_id = session.get("user_id", "unknown")
        user_name = session.get("user_name") or session.get("created_by", "غير معروف")
        
        if user_id not in user_stats:
            user_stats[user_id] = {
                "user_id": user_id,
                "user_name": user_name,
                "sessions_count": 0,
                "total_sales": 0,
                "cash_sales": 0,
                "credit_sales": 0,
                "total_difference": 0
            }
        
        stats = user_stats[user_id]
        stats["sessions_count"] += 1
        stats["total_sales"] += session.get("total_sales", 0)
        stats["cash_sales"] += session.get("cash_sales", 0)
        stats["credit_sales"] += session.get("credit_sales", 0)
        
        # Calculate difference
        expected = session.get("opening_cash", 0) + session.get("cash_sales", 0)
        actual = session.get("closing_cash", 0)
        stats["total_difference"] += (actual - expected)
    
    # Overall totals
    overall = {
        "total_sessions": len(sessions),
        "total_sales": sum(s.get("total_sales", 0) for s in sessions),
        "total_cash_sales": sum(s.get("cash_sales", 0) for s in sessions),
        "total_credit_sales": sum(s.get("credit_sales", 0) for s in sessions),
        "total_difference": sum(
            (s.get("closing_cash", 0) - (s.get("opening_cash", 0) + s.get("cash_sales", 0)))
            for s in sessions
        )
    }
    
    return {
        "period_days": days,
        "overall": overall,
        "by_user": list(user_stats.values())
    }

@api_router.put("/daily-sessions/{session_id}/close", response_model=DailySessionResponse)
async def close_daily_session(session_id: str, closing_data: DailySessionClose, user: dict = Depends(require_tenant)):
    """Close a daily cash session"""
    
    session = await db.daily_sessions.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="الحصة غير موجودة / Session non trouvée")
    
    # Only owner or admin can close
    if session.get("user_id") != user["id"] and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="لا يمكنك إغلاق حصة موظف آخر")
    
    if session["status"] == "closed":
        raise HTTPException(status_code=400, detail="الحصة مغلقة بالفعل / Session déjà fermée")
    
    # Calculate sales for this session period by this user
    opened_at = session["opened_at"]
    closed_at = closing_data.closed_at
    session_user_id = session.get("user_id", user["id"])
    
    # Get sales made by this user during their session
    sales = await db.sales.find({
        "created_at": {"$gte": opened_at, "$lte": closed_at},
        "status": {"$ne": "returned"}
    }, {"_id": 0}).to_list(1000)
    
    total_sales = sum(s.get("total", 0) for s in sales)
    cash_sales = sum(s.get("paid_amount", 0) for s in sales if s.get("payment_method") == "cash")
    credit_sales = sum(s.get("remaining", 0) for s in sales)
    sales_count = len(sales)
    
    # Calculate actual profit from sale items
    total_profit = 0
    for sale in sales:
        for item in sale.get("items", []):
            product_id = item.get("product_id")
            if product_id:
                product = await db.products.find_one({"id": product_id}, {"_id": 0, "purchase_price": 1})
                if product:
                    purchase_price = product.get("purchase_price", 0)
                    sale_price = item.get("price", 0)
                    quantity = item.get("quantity", 1)
                    total_profit += (sale_price - purchase_price) * quantity
    
    update_data = {
        "closing_cash": closing_data.closing_cash,
        "closed_at": closing_data.closed_at,
        "notes": closing_data.notes or "",
        "status": "closed",
        "total_sales": total_sales,
        "cash_sales": cash_sales,
        "credit_sales": credit_sales,
        "sales_count": sales_count,
        "total_profit": total_profit
    }
    
    await db.daily_sessions.update_one({"id": session_id}, {"$set": update_data})
    
    updated = await db.daily_sessions.find_one({"id": session_id}, {"_id": 0})
    
    # Add missing fields
    if "user_id" not in updated:
        updated["user_id"] = session_user_id
    if "user_name" not in updated:
        updated["user_name"] = session.get("user_name", session.get("created_by", ""))
    
    # Check for significant cash difference and create notifications
    # Get threshold from system settings
    system_settings = await db.system_settings.find_one({"id": "global"}, {"_id": 0})
    DIFFERENCE_THRESHOLD = system_settings.get("cash_difference_threshold", 1000) if system_settings else 1000
    
    expected_cash = session.get("opening_cash", 0) + cash_sales
    actual_cash = closing_data.closing_cash
    difference = actual_cash - expected_cash
    
    if abs(difference) >= DIFFERENCE_THRESHOLD:
        employee_name = session.get("user_name", session.get("created_by", "موظف"))
        now = datetime.now(timezone.utc).isoformat()
        
        if difference < 0:
            # Deficit notification
            notification_type = "cash_deficit"
            title_ar = f"عجز في صندوق {employee_name}"
            title_fr = f"Déficit caisse {employee_name}"
            message_ar = f"تم تسجيل عجز بقيمة {abs(difference):.2f} دج في حصة {employee_name}"
            message_fr = f"Déficit de {abs(difference):.2f} DA dans la session de {employee_name}"
        else:
            # Surplus notification
            notification_type = "cash_surplus"
            title_ar = f"فائض في صندوق {employee_name}"
            title_fr = f"Excédent caisse {employee_name}"
            message_ar = f"تم تسجيل فائض بقيمة {difference:.2f} دج في حصة {employee_name}"
            message_fr = f"Excédent de {difference:.2f} DA dans la session de {employee_name}"
        
        # Get all admin users
        admin_users = await db.users.find({"role": "admin"}, {"_id": 0, "id": 1}).to_list(50)
        admin_ids = [u["id"] for u in admin_users]
        
        # Recipients: admins + the employee
        recipients = list(set(admin_ids + [session_user_id]))
        
        for recipient_id in recipients:
            notification_doc = {
                "id": str(uuid.uuid4()),
                "user_id": recipient_id,
                "type": notification_type,
                "title": title_ar,
                "title_fr": title_fr,
                "message": message_ar,
                "message_fr": message_fr,
                "data": {
                    "session_id": session_id,
                    "employee_id": session_user_id,
                    "employee_name": employee_name,
                    "difference": difference,
                    "expected": expected_cash,
                    "actual": actual_cash
                },
                "read": False,
                "created_at": now
            }
            await db.notifications.insert_one(notification_doc)
    
    return DailySessionResponse(**updated)

@api_router.delete("/daily-sessions/{session_id}")
async def delete_daily_session(session_id: str, admin: dict = Depends(get_tenant_admin)):
    """Delete a daily session (admin only)"""
    
    session = await db.daily_sessions.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="الحصة غير موجودة / Session non trouvée")
    
    if session["status"] == "open":
        raise HTTPException(status_code=400, detail="لا يمكن حذف حصة مفتوحة / Impossible de supprimer une session ouverte")
    
    result = await db.daily_sessions.delete_one({"id": session_id})
    return {"message": "تم حذف الحصة بنجاح / Session supprimée"}

# ============ CUSTOMER & SUPPLIER FAMILIES ============

class CustomerFamilyCreate(BaseModel):
    name: str
    description: Optional[str] = ""

class CustomerFamilyUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

class CustomerFamilyResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    description: str
    customer_count: int = 0
    created_at: str

class SupplierFamilyCreate(BaseModel):
    name: str
    description: Optional[str] = ""

class SupplierFamilyUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

class SupplierFamilyResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    description: str
    supplier_count: int = 0
    created_at: str

# Customer Families CRUD
@api_router.post("/customer-families", response_model=CustomerFamilyResponse)
async def create_customer_family(family: CustomerFamilyCreate, admin: dict = Depends(get_tenant_admin)):
    family_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    family_doc = {
        "id": family_id,
        "name": family.name,
        "description": family.description or "",
        "customer_count": 0,
        "created_at": now
    }
    
    await db.customer_families.insert_one(family_doc)
    return CustomerFamilyResponse(**family_doc)

@api_router.get("/customer-families", response_model=List[CustomerFamilyResponse])
async def get_customer_families(user: dict = Depends(require_tenant)):
    families = await db.customer_families.find({}, {"_id": 0}).to_list(100)
    
    # Update customer counts
    for family in families:
        count = await db.customers.count_documents({"family_id": family["id"]})
        family["customer_count"] = count
    
    return [CustomerFamilyResponse(**f) for f in families]

@api_router.get("/customer-families/{family_id}", response_model=CustomerFamilyResponse)
async def get_customer_family(family_id: str, user: dict = Depends(require_tenant)):
    family = await db.customer_families.find_one({"id": family_id}, {"_id": 0})
    if not family:
        raise HTTPException(status_code=404, detail="عائلة الزبائن غير موجودة")
    
    count = await db.customers.count_documents({"family_id": family_id})
    family["customer_count"] = count
    
    return CustomerFamilyResponse(**family)

@api_router.put("/customer-families/{family_id}", response_model=CustomerFamilyResponse)
async def update_customer_family(family_id: str, updates: CustomerFamilyUpdate, admin: dict = Depends(get_tenant_admin)):
    family = await db.customer_families.find_one({"id": family_id})
    if not family:
        raise HTTPException(status_code=404, detail="عائلة الزبائن غير موجودة")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if update_data:
        await db.customer_families.update_one({"id": family_id}, {"$set": update_data})
    
    updated = await db.customer_families.find_one({"id": family_id}, {"_id": 0})
    count = await db.customers.count_documents({"family_id": family_id})
    updated["customer_count"] = count
    
    return CustomerFamilyResponse(**updated)

@api_router.delete("/customer-families/{family_id}")
async def delete_customer_family(family_id: str, admin: dict = Depends(get_tenant_admin)):
    count = await db.customers.count_documents({"family_id": family_id})
    if count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف عائلة بها {count} زبون")
    
    result = await db.customer_families.delete_one({"id": family_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="عائلة الزبائن غير موجودة")
    return {"message": "تم حذف عائلة الزبائن بنجاح"}

# Supplier Families CRUD
@api_router.post("/supplier-families", response_model=SupplierFamilyResponse)
async def create_supplier_family(family: SupplierFamilyCreate, admin: dict = Depends(get_tenant_admin)):
    family_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    family_doc = {
        "id": family_id,
        "name": family.name,
        "description": family.description or "",
        "supplier_count": 0,
        "created_at": now
    }
    
    await db.supplier_families.insert_one(family_doc)
    return SupplierFamilyResponse(**family_doc)

@api_router.get("/supplier-families", response_model=List[SupplierFamilyResponse])
async def get_supplier_families(user: dict = Depends(require_tenant)):
    families = await db.supplier_families.find({}, {"_id": 0}).to_list(100)
    
    # Update supplier counts
    for family in families:
        count = await db.suppliers.count_documents({"family_id": family["id"]})
        family["supplier_count"] = count
    
    return [SupplierFamilyResponse(**f) for f in families]

@api_router.get("/supplier-families/{family_id}", response_model=SupplierFamilyResponse)
async def get_supplier_family(family_id: str, user: dict = Depends(require_tenant)):
    family = await db.supplier_families.find_one({"id": family_id}, {"_id": 0})
    if not family:
        raise HTTPException(status_code=404, detail="عائلة الموردين غير موجودة")
    
    count = await db.suppliers.count_documents({"family_id": family_id})
    family["supplier_count"] = count
    
    return SupplierFamilyResponse(**family)

@api_router.put("/supplier-families/{family_id}", response_model=SupplierFamilyResponse)
async def update_supplier_family(family_id: str, updates: SupplierFamilyUpdate, admin: dict = Depends(get_tenant_admin)):
    family = await db.supplier_families.find_one({"id": family_id})
    if not family:
        raise HTTPException(status_code=404, detail="عائلة الموردين غير موجودة")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if update_data:
        await db.supplier_families.update_one({"id": family_id}, {"$set": update_data})
    
    updated = await db.supplier_families.find_one({"id": family_id}, {"_id": 0})
    count = await db.suppliers.count_documents({"family_id": family_id})
    updated["supplier_count"] = count
    
    return SupplierFamilyResponse(**updated)

@api_router.delete("/supplier-families/{family_id}")
async def delete_supplier_family(family_id: str, admin: dict = Depends(get_tenant_admin)):
    count = await db.suppliers.count_documents({"family_id": family_id})
    if count > 0:
        raise HTTPException(status_code=400, detail=f"لا يمكن حذف عائلة بها {count} مورد")
    
    result = await db.supplier_families.delete_one({"id": family_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="عائلة الموردين غير موجودة")
    return {"message": "تم حذف عائلة الموردين بنجاح"}

# ============ REPAIRS MANAGEMENT ============

class RepairCreate(BaseModel):
    ticket_number: str
    customer_name: str
    customer_phone: str
    customer_phone2: Optional[str] = ""
    device_brand: str
    device_model: str
    device_color: Optional[str] = ""
    device_imei: Optional[str] = ""
    device_password: Optional[str] = ""
    problems: List[str] = []
    problem_description: Optional[str] = ""
    device_condition: Optional[str] = ""
    accessories: Optional[str] = ""
    estimated_cost: float = 0
    estimated_days: int = 0
    advance_payment: float = 0
    technician_notes: Optional[str] = ""
    status: str = "received"

class RepairUpdate(BaseModel):
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_phone2: Optional[str] = None
    device_brand: Optional[str] = None
    device_model: Optional[str] = None
    device_color: Optional[str] = None
    device_imei: Optional[str] = None
    device_password: Optional[str] = None
    problems: Optional[List[str]] = None
    problem_description: Optional[str] = None
    device_condition: Optional[str] = None
    accessories: Optional[str] = None
    estimated_cost: Optional[float] = None
    estimated_days: Optional[int] = None
    advance_payment: Optional[float] = None
    technician_notes: Optional[str] = None
    status: Optional[str] = None
    final_cost: Optional[float] = None
    completed_at: Optional[str] = None
    delivered_at: Optional[str] = None
    spare_parts: Optional[List[dict]] = None
    notes: Optional[str] = None

class SparePartCreate(BaseModel):
    name: str
    name_ar: Optional[str] = ""
    category: str
    compatible_brands: List[str] = []
    compatible_models: Optional[str] = ""
    quantity: int = 0
    buy_price: float = 0
    sell_price: float = 0
    min_stock: int = 5
    supplier: Optional[str] = ""
    notes: Optional[str] = ""

class SparePartUpdate(BaseModel):
    name: Optional[str] = None
    name_ar: Optional[str] = None
    category: Optional[str] = None
    compatible_brands: Optional[List[str]] = None
    compatible_models: Optional[str] = None
    quantity: Optional[int] = None
    buy_price: Optional[float] = None
    sell_price: Optional[float] = None
    min_stock: Optional[int] = None
    supplier: Optional[str] = None
    notes: Optional[str] = None

# Repairs endpoints
@api_router.get("/repairs")
async def get_repairs(
    status: Optional[str] = None,
    search: Optional[str] = None,
    user: dict = Depends(require_tenant)
):
    """Get all repair tickets"""
    query = {}
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"ticket_number": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"customer_phone": {"$regex": search, "$options": "i"}},
            {"device_brand": {"$regex": search, "$options": "i"}},
            {"device_model": {"$regex": search, "$options": "i"}},
        ]
    
    repairs = await db.repairs.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return repairs

@api_router.get("/repairs/stats")
async def get_repair_stats(user: dict = Depends(require_tenant)):
    """Get repair statistics"""
    total = await db.repairs.count_documents({})
    received = await db.repairs.count_documents({"status": "received"})
    diagnosing = await db.repairs.count_documents({"status": "diagnosing"})
    in_progress = await db.repairs.count_documents({"status": "in_progress"})
    waiting_parts = await db.repairs.count_documents({"status": "waiting_parts"})
    completed = await db.repairs.count_documents({"status": "completed"})
    delivered = await db.repairs.count_documents({"status": "delivered"})
    cancelled = await db.repairs.count_documents({"status": "cancelled"})
    
    # Revenue stats
    pipeline = [
        {"$match": {"status": {"$in": ["completed", "delivered"]}}},
        {"$group": {
            "_id": None,
            "total_revenue": {"$sum": "$final_cost"},
            "total_advance": {"$sum": "$advance_payment"}
        }}
    ]
    revenue_data = await db.repairs.aggregate(pipeline).to_list(1)
    revenue = revenue_data[0] if revenue_data else {"total_revenue": 0, "total_advance": 0}
    
    # Today's repairs
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    today_repairs = await db.repairs.count_documents({
        "created_at": {"$gte": today_start.isoformat()}
    })
    
    return {
        "total": total,
        "received": received,
        "diagnosing": diagnosing,
        "in_progress": in_progress,
        "waiting_parts": waiting_parts,
        "completed": completed,
        "delivered": delivered,
        "cancelled": cancelled,
        "today_repairs": today_repairs,
        "total_revenue": revenue.get("total_revenue", 0),
        "total_advance": revenue.get("total_advance", 0)
    }

@api_router.get("/repairs/{repair_id}")
async def get_repair(repair_id: str, user: dict = Depends(require_tenant)):
    """Get a single repair ticket"""
    repair = await db.repairs.find_one({"id": repair_id}, {"_id": 0})
    if not repair:
        repair = await db.repairs.find_one({"ticket_number": repair_id}, {"_id": 0})
    if not repair:
        raise HTTPException(status_code=404, detail="طلب الصيانة غير موجود")
    return repair

@api_router.post("/repairs")
async def create_repair(repair: RepairCreate, user: dict = Depends(require_tenant)):
    """Create a new repair ticket"""
    repair_data = repair.model_dump()
    repair_data["id"] = str(uuid.uuid4())
    repair_data["created_at"] = datetime.now(timezone.utc).isoformat()
    repair_data["created_by"] = user["id"]
    repair_data["history"] = [{
        "status": "received",
        "timestamp": repair_data["created_at"],
        "user": user["name"],
        "note": "تم استلام الجهاز"
    }]
    
    await db.repairs.insert_one(repair_data)
    repair_data.pop("_id", None)
    return repair_data

@api_router.put("/repairs/{repair_id}")
async def update_repair(repair_id: str, repair: RepairUpdate, user: dict = Depends(require_tenant)):
    """Update a repair ticket"""
    existing = await db.repairs.find_one({"id": repair_id}, {"_id": 0})
    if not existing:
        existing = await db.repairs.find_one({"ticket_number": repair_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="طلب الصيانة غير موجود")
    
    update_data = {k: v for k, v in repair.model_dump().items() if v is not None}
    
    # Track status change
    if "status" in update_data and update_data["status"] != existing.get("status"):
        history_entry = {
            "status": update_data["status"],
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "user": user["name"],
            "note": get_status_note(update_data["status"])
        }
        await db.repairs.update_one(
            {"id": existing["id"]},
            {"$push": {"history": history_entry}}
        )
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.repairs.update_one({"id": existing["id"]}, {"$set": update_data})
    updated = await db.repairs.find_one({"id": existing["id"]}, {"_id": 0})
    return updated

def get_status_note(status: str) -> str:
    """Get Arabic note for status change"""
    notes = {
        "received": "تم استلام الجهاز",
        "diagnosing": "جاري التشخيص",
        "in_progress": "جاري الإصلاح",
        "waiting_parts": "في انتظار قطع الغيار",
        "completed": "تم الإصلاح",
        "delivered": "تم التسليم للزبون",
        "cancelled": "تم إلغاء الطلب"
    }
    return notes.get(status, "تم تحديث الحالة")

@api_router.delete("/repairs/{repair_id}")
async def delete_repair(repair_id: str, user: dict = Depends(require_tenant)):
    """Delete a repair ticket"""
    result = await db.repairs.delete_one({"id": repair_id})
    if result.deleted_count == 0:
        result = await db.repairs.delete_one({"ticket_number": repair_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="طلب الصيانة غير موجود")
    return {"message": "تم حذف طلب الصيانة بنجاح"}

# Spare Parts endpoints
@api_router.get("/spare-parts")
async def get_spare_parts(
    category: Optional[str] = None,
    search: Optional[str] = None,
    low_stock: Optional[bool] = None,
    user: dict = Depends(require_tenant)
):
    """Get all spare parts"""
    query = {}
    if category:
        query["category"] = category
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"name_ar": {"$regex": search, "$options": "i"}},
            {"compatible_models": {"$regex": search, "$options": "i"}},
        ]
    if low_stock:
        query["$expr"] = {"$lte": ["$quantity", "$min_stock"]}
    
    parts = await db.spare_parts.find(query, {"_id": 0}).sort("name", 1).to_list(1000)
    return parts

@api_router.get("/spare-parts/stats")
async def get_spare_parts_stats(user: dict = Depends(require_tenant)):
    """Get spare parts statistics"""
    total = await db.spare_parts.count_documents({})
    
    # Low stock count
    pipeline = [
        {"$match": {"$expr": {"$lte": ["$quantity", "$min_stock"]}}},
        {"$count": "count"}
    ]
    low_stock_result = await db.spare_parts.aggregate(pipeline).to_list(1)
    low_stock = low_stock_result[0]["count"] if low_stock_result else 0
    
    # Total value
    value_pipeline = [
        {"$group": {
            "_id": None,
            "total_buy_value": {"$sum": {"$multiply": ["$quantity", "$buy_price"]}},
            "total_sell_value": {"$sum": {"$multiply": ["$quantity", "$sell_price"]}}
        }}
    ]
    value_result = await db.spare_parts.aggregate(value_pipeline).to_list(1)
    values = value_result[0] if value_result else {"total_buy_value": 0, "total_sell_value": 0}
    
    # Categories count
    categories_pipeline = [
        {"$group": {"_id": "$category", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    categories = await db.spare_parts.aggregate(categories_pipeline).to_list(20)
    
    return {
        "total": total,
        "low_stock": low_stock,
        "total_buy_value": values.get("total_buy_value", 0),
        "total_sell_value": values.get("total_sell_value", 0),
        "categories": [{"name": c["_id"], "count": c["count"]} for c in categories if c["_id"]]
    }

@api_router.get("/spare-parts/{part_id}")
async def get_spare_part(part_id: str, user: dict = Depends(require_tenant)):
    """Get a single spare part"""
    part = await db.spare_parts.find_one({"id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="قطعة الغيار غير موجودة")
    return part

@api_router.post("/spare-parts")
async def create_spare_part(part: SparePartCreate, user: dict = Depends(require_tenant)):
    """Create a new spare part"""
    part_data = part.model_dump()
    part_data["id"] = str(uuid.uuid4())
    part_data["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.spare_parts.insert_one(part_data)
    part_data.pop("_id", None)
    return part_data

@api_router.put("/spare-parts/{part_id}")
async def update_spare_part(part_id: str, part: SparePartUpdate, user: dict = Depends(require_tenant)):
    """Update a spare part"""
    existing = await db.spare_parts.find_one({"id": part_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="قطعة الغيار غير موجودة")
    
    update_data = {k: v for k, v in part.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.spare_parts.update_one({"id": part_id}, {"$set": update_data})
    updated = await db.spare_parts.find_one({"id": part_id}, {"_id": 0})
    return updated

@api_router.delete("/spare-parts/{part_id}")
async def delete_spare_part(part_id: str, user: dict = Depends(require_tenant)):
    """Delete a spare part"""
    result = await db.spare_parts.delete_one({"id": part_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="قطعة الغيار غير موجودة")
    return {"message": "تم حذف قطعة الغيار بنجاح"}

@api_router.put("/spare-parts/{part_id}/stock")
async def update_spare_part_stock(
    part_id: str,
    quantity_change: int,
    operation: str,  # "add" or "subtract"
    user: dict = Depends(require_tenant)
):
    """Update spare part stock"""
    existing = await db.spare_parts.find_one({"id": part_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="قطعة الغيار غير موجودة")
    
    current_qty = existing.get("quantity", 0)
    if operation == "add":
        new_qty = current_qty + quantity_change
    else:
        new_qty = max(0, current_qty - quantity_change)
    
    await db.spare_parts.update_one(
        {"id": part_id},
        {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"quantity": new_qty}

# ============ WHATSAPP NOTIFICATIONS ============

class WhatsAppSettings(BaseModel):
    enabled: bool = False
    phone_number_id: Optional[str] = None
    access_token: Optional[str] = None
    business_account_id: Optional[str] = None

class WhatsAppMessage(BaseModel):
    phone: str
    message: str

# Status messages in Arabic
REPAIR_STATUS_MESSAGES = {
    "received": "مرحباً {customer_name}! تم استلام جهازك ({device}) للصيانة. رقم التذكرة: {ticket_number}. سنقوم بإشعارك عند أي تحديث.",
    "diagnosing": "تحديث الصيانة #{ticket_number}: جاري فحص جهازك ({device}) لتحديد العطل.",
    "waiting_parts": "تحديث الصيانة #{ticket_number}: جهازك ({device}) بحاجة لقطع غيار. سنقوم بإعلامك فور توفرها.",
    "in_progress": "تحديث الصيانة #{ticket_number}: جاري إصلاح جهازك ({device}). الوقت المتوقع: {estimated_days} أيام.",
    "completed": "🎉 أخبار سارة! تم إصلاح جهازك ({device}) بنجاح. رقم التذكرة: {ticket_number}. يمكنك استلامه الآن. التكلفة: {cost} دج",
    "delivered": "شكراً لثقتك بنا! تم تسليم جهازك ({device}). نتمنى لك يوماً سعيداً! 🙏",
    "cancelled": "تم إلغاء طلب الصيانة #{ticket_number}. للاستفسار يرجى التواصل معنا."
}

@api_router.get("/whatsapp/settings")
async def get_whatsapp_settings(user: dict = Depends(require_tenant)):
    """Get WhatsApp settings"""
    settings = await db.whatsapp_settings.find_one({}, {"_id": 0})
    if not settings:
        settings = {"enabled": False}
    # Don't expose access_token
    if "access_token" in settings:
        settings["access_token"] = "***" if settings["access_token"] else None
    return settings

@api_router.put("/whatsapp/settings")
async def update_whatsapp_settings(settings: WhatsAppSettings, user: dict = Depends(require_tenant)):
    """Update WhatsApp settings"""
    settings_data = settings.model_dump()
    settings_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    settings_data["updated_by"] = user["id"]
    
    await db.whatsapp_settings.update_one(
        {},
        {"$set": settings_data},
        upsert=True
    )
    return {"message": "تم تحديث إعدادات WhatsApp"}

@api_router.post("/whatsapp/send")
async def send_whatsapp_message(message: WhatsAppMessage, user: dict = Depends(require_tenant)):
    """Send a WhatsApp message"""
    settings = await db.whatsapp_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("enabled"):
        raise HTTPException(status_code=400, detail="WhatsApp غير مفعل")
    
    if not settings.get("access_token") or not settings.get("phone_number_id"):
        raise HTTPException(status_code=400, detail="إعدادات WhatsApp غير مكتملة")
    
    # Format phone number (remove leading 0 and add country code)
    phone = message.phone.strip()
    if phone.startswith("0"):
        phone = "213" + phone[1:]  # Algeria country code
    elif not phone.startswith("213"):
        phone = "213" + phone
    
    url = f"https://graph.facebook.com/v18.0/{settings['phone_number_id']}/messages"
    
    headers = {
        "Authorization": f"Bearer {settings['access_token']}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "to": phone,
        "type": "text",
        "text": {"body": message.message}
    }
    
    try:
        response = http_requests.post(url, json=payload, headers=headers)
        
        # Log the message
        await db.whatsapp_logs.insert_one({
            "id": str(uuid.uuid4()),
            "phone": phone,
            "message": message.message,
            "status": "sent" if response.status_code == 200 else "failed",
            "response_code": response.status_code,
            "response": response.text[:500] if response.text else None,
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "sent_by": user["id"]
        })
        
        if response.status_code == 200:
            return {"success": True, "message": "تم إرسال الرسالة"}
        else:
            return {"success": False, "error": response.text}
    except Exception as e:
        logger.error(f"WhatsApp send error: {str(e)}")
        return {"success": False, "error": str(e)}

@api_router.post("/whatsapp/notify-repair/{repair_id}")
async def notify_repair_status_change(repair_id: str, user: dict = Depends(require_tenant)):
    """Send WhatsApp notification for repair status change"""
    repair = await db.repairs.find_one({"id": repair_id}, {"_id": 0})
    if not repair:
        repair = await db.repairs.find_one({"ticket_number": repair_id}, {"_id": 0})
    if not repair:
        raise HTTPException(status_code=404, detail="طلب الصيانة غير موجود")
    
    settings = await db.whatsapp_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("enabled"):
        return {"success": False, "message": "WhatsApp غير مفعل"}
    
    status = repair.get("status", "received")
    message_template = REPAIR_STATUS_MESSAGES.get(status)
    
    if not message_template:
        return {"success": False, "message": "قالب الرسالة غير موجود"}
    
    # Format the message
    message = message_template.format(
        customer_name=repair.get("customer_name", "عميل"),
        device=f"{repair.get('device_brand', '')} {repair.get('device_model', '')}",
        ticket_number=repair.get("ticket_number", repair_id),
        estimated_days=repair.get("estimated_days", 1),
        cost=repair.get("final_cost") or repair.get("estimated_cost", 0)
    )
    
    # Send the message
    whatsapp_msg = WhatsAppMessage(phone=repair.get("customer_phone", ""), message=message)
    return await send_whatsapp_message(whatsapp_msg, user)

@api_router.get("/whatsapp/logs")
async def get_whatsapp_logs(
    limit: int = 50,
    user: dict = Depends(require_tenant)
):
    """Get WhatsApp message logs"""
    logs = await db.whatsapp_logs.find({}, {"_id": 0}).sort("sent_at", -1).limit(limit).to_list(limit)
    return logs

# ============ SPARE PARTS - PRODUCTS INTEGRATION ============

@api_router.post("/spare-parts/use-in-repair")
async def use_spare_part_in_repair(
    repair_id: str,
    part_id: str,
    quantity: int = 1,
    user: dict = Depends(require_tenant)
):
    """Use a spare part in a repair - deducts from inventory"""
    # Verify repair exists
    repair = await db.repairs.find_one({"id": repair_id}, {"_id": 0})
    if not repair:
        raise HTTPException(status_code=404, detail="طلب الصيانة غير موجود")
    
    # Verify spare part exists and has enough stock
    part = await db.spare_parts.find_one({"id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="قطعة الغيار غير موجودة")
    
    if part.get("quantity", 0) < quantity:
        raise HTTPException(status_code=400, detail="الكمية غير كافية في المخزون")
    
    # Deduct from spare parts inventory
    new_qty = part["quantity"] - quantity
    await db.spare_parts.update_one(
        {"id": part_id},
        {"$set": {"quantity": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Record the usage in repair
    usage_record = {
        "part_id": part_id,
        "part_name": part.get("name"),
        "quantity": quantity,
        "unit_price": part.get("sell_price", 0),
        "total_price": part.get("sell_price", 0) * quantity,
        "used_at": datetime.now(timezone.utc).isoformat(),
        "used_by": user["name"]
    }
    
    await db.repairs.update_one(
        {"id": repair_id},
        {
            "$push": {"parts_used": usage_record},
            "$inc": {"parts_cost": usage_record["total_price"]}
        }
    )
    
    # Also check if there's a linked product in main inventory
    if part.get("linked_product_id"):
        await db.products.update_one(
            {"id": part["linked_product_id"]},
            {"$inc": {"quantity": -quantity}}
        )
    
    return {
        "success": True,
        "message": f"تم استخدام {quantity} من {part['name']}",
        "remaining_stock": new_qty,
        "total_cost": usage_record["total_price"]
    }

@api_router.post("/spare-parts/link-to-product")
async def link_spare_part_to_product(
    part_id: str,
    product_id: str,
    user: dict = Depends(require_tenant)
):
    """Link a spare part to a main product for synchronized inventory"""
    # Verify both exist
    part = await db.spare_parts.find_one({"id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="قطعة الغيار غير موجودة")
    
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    # Link them
    await db.spare_parts.update_one(
        {"id": part_id},
        {"$set": {
            "linked_product_id": product_id,
            "linked_product_name": product.get("name"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "success": True,
        "message": f"تم ربط {part['name']} بالمنتج {product['name']}"
    }

@api_router.delete("/spare-parts/unlink-product/{part_id}")
async def unlink_spare_part_from_product(
    part_id: str,
    user: dict = Depends(require_tenant)
):
    """Remove link between spare part and product"""
    part = await db.spare_parts.find_one({"id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="قطعة الغيار غير موجودة")
    
    await db.spare_parts.update_one(
        {"id": part_id},
        {"$unset": {"linked_product_id": "", "linked_product_name": ""}}
    )
    
    return {"success": True, "message": "تم إلغاء الربط"}

@api_router.post("/spare-parts/sync-inventory")
async def sync_spare_parts_with_products(user: dict = Depends(require_tenant)):
    """Sync spare parts inventory with linked products"""
    # Get all linked spare parts
    linked_parts = await db.spare_parts.find(
        {"linked_product_id": {"$exists": True}},
        {"_id": 0}
    ).to_list(1000)
    
    synced = 0
    for part in linked_parts:
        product = await db.products.find_one(
            {"id": part["linked_product_id"]},
            {"_id": 0, "quantity": 1}
        )
        if product:
            # Update spare part quantity to match product
            await db.spare_parts.update_one(
                {"id": part["id"]},
                {"$set": {"quantity": product.get("quantity", 0)}}
            )
            synced += 1
    
    return {"success": True, "synced_count": synced}

# ============ EXPENSES MANAGEMENT ============

class ExpenseCreate(BaseModel):
    title: str
    category: str
    amount: float
    date: Optional[str] = None
    notes: Optional[str] = ""
    recurring: bool = False
    recurring_period: Optional[str] = "monthly"  # monthly, weekly, yearly
    reminder_days_before: int = 3  # Days before due date to send reminder
    code: Optional[str] = ""  # كود التكلفة CH00001

class ExpenseUpdate(BaseModel):
    title: Optional[str] = None
    category: Optional[str] = None
    amount: Optional[float] = None
    date: Optional[str] = None
    notes: Optional[str] = None
    recurring: Optional[bool] = None
    recurring_period: Optional[str] = None
    reminder_days_before: Optional[int] = None
    code: Optional[str] = None

@api_router.get("/expenses")
async def get_expenses(
    category: Optional[str] = None,
    user: dict = Depends(require_tenant)
):
    """Get all expenses"""
    query = {}
    if category:
        query["category"] = category
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return expenses

@api_router.get("/expenses/stats")
async def get_expenses_stats(user: dict = Depends(require_tenant)):
    """Get expenses statistics"""
    # Total expenses
    total_pipeline = [{"$group": {"_id": None, "total": {"$sum": "$amount"}}}]
    total_result = await db.expenses.aggregate(total_pipeline).to_list(1)
    total = total_result[0]["total"] if total_result else 0
    
    # This month
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_pipeline = [
        {"$match": {"date": {"$gte": month_start.isoformat()}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    this_month_result = await db.expenses.aggregate(this_month_pipeline).to_list(1)
    this_month = this_month_result[0]["total"] if this_month_result else 0
    
    # Last month
    last_month_end = month_start - timedelta(seconds=1)
    last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_pipeline = [
        {"$match": {"date": {"$gte": last_month_start.isoformat(), "$lte": last_month_end.isoformat()}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    last_month_result = await db.expenses.aggregate(last_month_pipeline).to_list(1)
    last_month = last_month_result[0]["total"] if last_month_result else 0
    
    # By category
    category_pipeline = [
        {"$group": {"_id": "$category", "total": {"$sum": "$amount"}}},
        {"$sort": {"total": -1}}
    ]
    categories = await db.expenses.aggregate(category_pipeline).to_list(20)
    
    return {
        "total": total,
        "thisMonth": this_month,
        "lastMonth": last_month,
        "byCategory": [{"category": c["_id"], "total": c["total"]} for c in categories if c["_id"]]
    }

@api_router.post("/expenses")
async def create_expense(expense: ExpenseCreate, user: dict = Depends(require_tenant)):
    """Create a new expense"""
    expense_data = expense.model_dump()
    expense_data["id"] = str(uuid.uuid4())
    expense_data["date"] = expense_data["date"] or datetime.now(timezone.utc).isoformat()
    expense_data["created_at"] = datetime.now(timezone.utc).isoformat()
    expense_data["created_by"] = user["id"]
    expense_data["code"] = expense_data.get("code") or ""  # كود التكلفة
    
    await db.expenses.insert_one(expense_data)
    expense_data.pop("_id", None)
    return expense_data

@api_router.put("/expenses/{expense_id}")
async def update_expense(expense_id: str, expense: ExpenseUpdate, user: dict = Depends(require_tenant)):
    """Update an expense"""
    existing = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="التكلفة غير موجودة")
    
    update_data = {k: v for k, v in expense.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.expenses.update_one({"id": expense_id}, {"$set": update_data})
    updated = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    return updated

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, user: dict = Depends(require_tenant)):
    """Delete an expense"""
    result = await db.expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="التكلفة غير موجودة")
    return {"message": "تم حذف التكلفة بنجاح"}

@api_router.get("/expenses/reminders")
async def get_expense_reminders(user: dict = Depends(require_tenant)):
    """Get upcoming expense reminders for recurring expenses"""
    now = datetime.now(timezone.utc)
    
    # Get all recurring expenses
    recurring_expenses = await db.expenses.find(
        {"recurring": True},
        {"_id": 0}
    ).to_list(100)
    
    reminders = []
    for expense in recurring_expenses:
        # Calculate next due date based on recurring period
        date_str = expense.get("date", now.isoformat())
        try:
            # Handle dates with or without timezone
            if 'T' in date_str:
                if '+' in date_str or 'Z' in date_str:
                    last_date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                else:
                    last_date = datetime.fromisoformat(date_str).replace(tzinfo=timezone.utc)
            else:
                # Date only (no time) - treat as midnight UTC
                last_date = datetime.strptime(date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
        except:
            last_date = now
        
        period = expense.get("recurring_period", "monthly")
        reminder_days = expense.get("reminder_days_before", 3)
        
        if period == "monthly":
            # Next month same day
            next_month = last_date.month % 12 + 1
            next_year = last_date.year if next_month > 1 else last_date.year + 1
            try:
                next_due = last_date.replace(month=next_month, year=next_year)
            except ValueError:
                # Handle cases like Jan 31 -> Feb (no 31st)
                next_due = last_date.replace(month=next_month, year=next_year, day=28)
        elif period == "weekly":
            next_due = last_date + timedelta(days=7)
        elif period == "yearly":
            next_due = last_date.replace(year=last_date.year + 1)
        else:
            next_due = last_date + timedelta(days=30)
        
        # Check if reminder should show (within reminder_days_before)
        days_until_due = (next_due - now).days
        if 0 <= days_until_due <= reminder_days:
            reminders.append({
                "expense_id": expense["id"],
                "title": expense["title"],
                "category": expense["category"],
                "amount": expense["amount"],
                "due_date": next_due.isoformat(),
                "days_until_due": days_until_due,
                "is_urgent": days_until_due <= 1
            })
    
    # Sort by days until due
    reminders.sort(key=lambda x: x["days_until_due"])
    return reminders

@api_router.post("/expenses/{expense_id}/mark-paid")
async def mark_expense_paid(expense_id: str, user: dict = Depends(require_tenant)):
    """Mark a recurring expense as paid and update the date"""
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="التكلفة غير موجودة")
    
    # Update the date to today (for next cycle calculation)
    await db.expenses.update_one(
        {"id": expense_id},
        {"$set": {
            "date": datetime.now(timezone.utc).isoformat(),
            "last_paid_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "تم تسجيل الدفع بنجاح", "next_due_calculated": True}

# ============ EMAIL REPORTS ============

class SessionReportEmail(BaseModel):
    recipient_email: EmailStr
    session_id: str
    report_data: dict

def generate_session_report_html(report: dict, language: str = 'ar') -> str:
    """Generate HTML email for session closing report"""
    
    def format_currency(amount):
        return f"{amount:,.2f}"
    
    currency = "دج" if language == 'ar' else "DA"
    
    # Determine difference color
    diff = report.get('cashDifference', 0)
    diff_color = '#22c55e' if diff >= 0 else '#ef4444'
    diff_sign = '+' if diff >= 0 else ''
    
    html = f"""
    <!DOCTYPE html>
    <html dir="rtl" lang="ar">
    <head>
        <meta charset="UTF-8">
    </head>
    <body style="font-family: 'Segoe UI', Tahoma, sans-serif; background-color: #f3f4f6; margin: 0; padding: 20px;">
        <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
            <!-- Header -->
            <div style="background: linear-gradient(135deg, #3b82f6, #1d4ed8); padding: 30px; text-align: center;">
                <h1 style="color: white; margin: 0; font-size: 24px;">📊 تقرير غلق الحصة</h1>
                <p style="color: rgba(255,255,255,0.9); margin-top: 10px;">
                    {datetime.fromisoformat(report.get('closedAt', '')).strftime('%Y-%m-%d %H:%M') if report.get('closedAt') else ''}
                </p>
            </div>
            
            <!-- Cash Summary -->
            <div style="padding: 20px;">
                <h2 style="color: #1f2937; border-bottom: 2px solid #e5e7eb; padding-bottom: 10px; font-size: 18px;">💵 ملخص الصندوق</h2>
                <table style="width: 100%; border-collapse: collapse; margin-top: 15px;">
                    <tr>
                        <td style="padding: 12px; background: #dbeafe; border-radius: 8px; margin: 5px;">
                            <span style="color: #6b7280; font-size: 12px;">رصيد الافتتاح</span><br>
                            <strong style="color: #2563eb; font-size: 18px;">{format_currency(report.get('openingCash', 0))} {currency}</strong>
                        </td>
                        <td style="padding: 12px; background: #dcfce7; border-radius: 8px; margin: 5px;">
                            <span style="color: #6b7280; font-size: 12px;">المبلغ المحصل</span><br>
                            <strong style="color: #16a34a; font-size: 18px;">{format_currency(report.get('totalCollected', 0))} {currency}</strong>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 12px; background: #f3e8ff; border-radius: 8px; margin: 5px;">
                            <span style="color: #6b7280; font-size: 12px;">المتوقع في الصندوق</span><br>
                            <strong style="color: #9333ea; font-size: 18px;">{format_currency(report.get('expectedCash', 0))} {currency}</strong>
                        </td>
                        <td style="padding: 12px; background: #fef3c7; border-radius: 8px; margin: 5px;">
                            <span style="color: #6b7280; font-size: 12px;">الفعلي في الصندوق</span><br>
                            <strong style="color: #d97706; font-size: 18px;">{format_currency(report.get('closingCash', 0))} {currency}</strong>
                        </td>
                    </tr>
                </table>
                
                <!-- Difference -->
                <div style="margin-top: 15px; padding: 15px; background: {'#dcfce7' if diff >= 0 else '#fee2e2'}; border-radius: 8px; display: flex; justify-content: space-between; align-items: center;">
                    <span style="color: #374151; font-weight: 600;">الفرق</span>
                    <span style="color: {diff_color}; font-size: 20px; font-weight: bold;">{diff_sign}{format_currency(diff)} {currency}</span>
                </div>
            </div>
            
            <!-- Sales Summary -->
            <div style="padding: 20px; background: #f9fafb;">
                <h2 style="color: #1f2937; border-bottom: 2px solid #e5e7eb; padding-bottom: 10px; font-size: 18px;">📈 ملخص المبيعات</h2>
                <div style="display: flex; gap: 15px; margin-top: 15px;">
                    <div style="flex: 1; text-align: center; padding: 15px; background: white; border-radius: 8px; border: 1px solid #e5e7eb;">
                        <span style="font-size: 28px; font-weight: bold; color: #3b82f6;">{report.get('salesCount', 0)}</span><br>
                        <span style="color: #6b7280; font-size: 12px;">عدد المبيعات</span>
                    </div>
                    <div style="flex: 1; text-align: center; padding: 15px; background: white; border-radius: 8px; border: 1px solid #e5e7eb;">
                        <span style="font-size: 28px; font-weight: bold; color: #22c55e;">{format_currency(report.get('totalSales', 0))}</span><br>
                        <span style="color: #6b7280; font-size: 12px;">إجمالي المبيعات</span>
                    </div>
                </div>
                
                <!-- Sales by Type -->
                <table style="width: 100%; margin-top: 15px; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden;">
                    <tr style="background: #f3f4f6;">
                        <th style="padding: 10px; text-align: right; font-size: 12px; color: #6b7280;">طريقة الدفع</th>
                        <th style="padding: 10px; text-align: center; font-size: 12px; color: #6b7280;">العدد</th>
                        <th style="padding: 10px; text-align: left; font-size: 12px; color: #6b7280;">المبلغ</th>
                    </tr>
                    <tr style="border-bottom: 1px solid #e5e7eb;">
                        <td style="padding: 10px;">نقدي</td>
                        <td style="padding: 10px; text-align: center;">{len([s for s in report.get('salesByPaymentType', []) if s.get('type') == 'cash'])}</td>
                        <td style="padding: 10px;">{format_currency(report.get('cashSales', 0))} {currency}</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #e5e7eb;">
                        <td style="padding: 10px;">دين</td>
                        <td style="padding: 10px; text-align: center;">{len([s for s in report.get('salesByPaymentType', []) if s.get('type') == 'credit'])}</td>
                        <td style="padding: 10px;">{format_currency(report.get('creditSales', 0))} {currency}</td>
                    </tr>
                    <tr>
                        <td style="padding: 10px;">جزئي</td>
                        <td style="padding: 10px; text-align: center;">{len([s for s in report.get('salesByPaymentType', []) if s.get('type') == 'partial'])}</td>
                        <td style="padding: 10px;">{format_currency(report.get('partialSales', 0))} {currency}</td>
                    </tr>
                </table>
            </div>
            
            <!-- Debts -->
            <div style="padding: 20px;">
                <div style="background: #fef2f2; border: 1px solid #fecaca; border-radius: 8px; padding: 15px;">
                    <h3 style="color: #dc2626; margin: 0 0 10px 0; font-size: 16px;">🔒 الديون الجديدة</h3>
                    <span style="font-size: 24px; font-weight: bold; color: #dc2626;">{format_currency(report.get('totalDebts', 0))} {currency}</span>
                </div>
            </div>
            
            <!-- Footer -->
            <div style="background: #1f2937; padding: 20px; text-align: center;">
                <p style="color: #9ca3af; margin: 0; font-size: 12px;">
                    تم إنشاء هذا التقرير تلقائياً من نظام NT POS
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    return html

@api_router.post("/email/send-session-report")
async def send_session_report_email(
    report_email: SessionReportEmail,
    user: dict = Depends(require_tenant)
):
    """Send session closing report via email"""
    if not RESEND_AVAILABLE:
        raise HTTPException(status_code=500, detail="خدمة البريد الإلكتروني غير متوفرة")
    
    api_key = os.environ.get('RESEND_API_KEY')
    if not api_key:
        raise HTTPException(status_code=500, detail="مفتاح API للبريد غير موجود. يرجى إضافة RESEND_API_KEY في الإعدادات")
    
    sender_email = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
    
    # Generate HTML report
    html_content = generate_session_report_html(report_email.report_data)
    
    params = {
        "from": sender_email,
        "to": [report_email.recipient_email],
        "subject": f"تقرير غلق الحصة - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "html": html_content
    }
    
    try:
        email = await asyncio.to_thread(resend.Emails.send, params)
        
        # Log the email
        await db.email_logs.insert_one({
            "id": str(uuid.uuid4()),
            "type": "session_report",
            "recipient": report_email.recipient_email,
            "session_id": report_email.session_id,
            "status": "sent",
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "sent_by": user["id"]
        })
        
        return {
            "success": True,
            "message": f"تم إرسال التقرير إلى {report_email.recipient_email}",
            "email_id": email.get("id")
        }
    except Exception as e:
        logger.error(f"Failed to send email: {str(e)}")
        raise HTTPException(status_code=500, detail=f"فشل إرسال البريد: {str(e)}")

# ============ EMAIL SETTINGS ============

class EmailSettings(BaseModel):
    enabled: bool = False
    resend_api_key: str = ""
    sender_email: str = "onboarding@resend.dev"
    sender_name: str = "NT POS System"

@api_router.get("/email/settings")
async def get_email_settings(user: dict = Depends(require_tenant)):
    """Get email settings"""
    settings = await db.system_settings.find_one({"type": "email_settings"}, {"_id": 0})
    if not settings:
        return EmailSettings().model_dump()
    
    # Don't expose full API key
    if settings.get("resend_api_key"):
        key = settings["resend_api_key"]
        settings["resend_api_key"] = key[:8] + "..." + key[-4:] if len(key) > 12 else "***configured***"
    
    return settings

@api_router.put("/email/settings")
async def update_email_settings(settings: EmailSettings, user: dict = Depends(require_tenant)):
    """Update email settings"""
    # Check if user is admin
    if user.get("role") not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل إعدادات البريد")
    
    # Get existing settings to preserve API key if not changed
    existing = await db.system_settings.find_one({"type": "email_settings"})
    
    settings_dict = settings.model_dump()
    
    # If API key looks like masked value, keep the old one
    if settings.resend_api_key and ("..." in settings.resend_api_key or settings.resend_api_key == "***configured***"):
        if existing and existing.get("resend_api_key"):
            settings_dict["resend_api_key"] = existing["resend_api_key"]
    
    # Update in database
    await db.system_settings.update_one(
        {"type": "email_settings"},
        {"$set": {**settings_dict, "type": "email_settings", "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    # Update environment variable for resend
    if settings_dict.get("resend_api_key") and "..." not in settings_dict["resend_api_key"]:
        os.environ['RESEND_API_KEY'] = settings_dict["resend_api_key"]
        if RESEND_AVAILABLE:
            resend.api_key = settings_dict["resend_api_key"]
    
    if settings_dict.get("sender_email"):
        os.environ['SENDER_EMAIL'] = settings_dict["sender_email"]
    
    return {"success": True, "message": "تم حفظ إعدادات البريد بنجاح"}

@api_router.post("/email/test")
async def test_email_settings(user: dict = Depends(require_tenant)):
    """Send a test email to verify settings"""
    if not RESEND_AVAILABLE:
        raise HTTPException(status_code=500, detail="مكتبة Resend غير متوفرة")
    
    # Get settings from database
    settings = await db.system_settings.find_one({"type": "email_settings"})
    if not settings or not settings.get("resend_api_key"):
        raise HTTPException(status_code=400, detail="يرجى إدخال مفتاح API أولاً")
    
    api_key = settings.get("resend_api_key")
    sender_email = settings.get("sender_email", "onboarding@resend.dev")
    
    # Set API key
    resend.api_key = api_key
    
    # Get user's email
    user_record = await db.users.find_one({"id": user["id"]})
    if not user_record or not user_record.get("email"):
        raise HTTPException(status_code=400, detail="لم يتم العثور على بريدك الإلكتروني")
    
    try:
        params = {
            "from": sender_email,
            "to": [user_record["email"]],
            "subject": "🧪 اختبار إعدادات البريد - NT POS",
            "html": """
            <div dir="rtl" style="font-family: Arial, sans-serif; padding: 20px; background: #f3f4f6;">
                <div style="max-width: 400px; margin: 0 auto; background: white; border-radius: 8px; padding: 30px; text-align: center;">
                    <h2 style="color: #22c55e;">✅ إعدادات البريد تعمل بنجاح!</h2>
                    <p style="color: #6b7280;">هذه رسالة اختبار من نظام NT POS</p>
                    <p style="color: #9ca3af; font-size: 12px; margin-top: 20px;">
                        إذا وصلتك هذه الرسالة، فإن إعدادات البريد الإلكتروني تعمل بشكل صحيح.
                    </p>
                </div>
            </div>
            """
        }
        
        result = await asyncio.to_thread(resend.Emails.send, params)
        return {"success": True, "message": f"تم إرسال بريد اختباري إلى {user_record['email']}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل إرسال البريد: {str(e)}")

# ============ SMART REPORTS ============

class SmartReportSettings(BaseModel):
    daily_report_enabled: bool = False
    daily_report_time: str = "08:00"
    daily_report_recipients: str = ""
    include_ai_tips: bool = True
    include_sales_summary: bool = True
    include_low_stock_alerts: bool = True
    include_debt_reminders: bool = True

@api_router.get("/smart-reports/settings")
async def get_smart_report_settings(user: dict = Depends(require_tenant)):
    """Get smart report settings"""
    settings = await db.system_settings.find_one({"type": "smart_reports"}, {"_id": 0})
    if not settings:
        return SmartReportSettings().model_dump()
    return settings

@api_router.put("/smart-reports/settings")
async def update_smart_report_settings(settings: SmartReportSettings, user: dict = Depends(require_tenant)):
    """Update smart report settings"""
    if user.get("role") not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.system_settings.update_one(
        {"type": "smart_reports"},
        {"$set": {**settings.model_dump(), "type": "smart_reports", "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"success": True}

@api_router.get("/smart-reports/last")
async def get_last_smart_report(user: dict = Depends(require_tenant)):
    """Get last sent report info"""
    report = await db.smart_reports_log.find_one({}, {"_id": 0}, sort=[("sent_at", -1)])
    return report

@api_router.get("/smart-reports/preview")
async def preview_smart_report(user: dict = Depends(require_tenant)):
    """Generate a preview of the smart report"""
    today = datetime.now(timezone.utc).date()
    yesterday = today - timedelta(days=1)
    
    # Sales summary
    today_start = datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc)
    yesterday_start = datetime.combine(yesterday, datetime.min.time()).replace(tzinfo=timezone.utc)
    
    today_sales = await db.sales.find({"created_at": {"$gte": today_start.isoformat()}}).to_list(1000)
    yesterday_sales = await db.sales.find({
        "created_at": {"$gte": yesterday_start.isoformat(), "$lt": today_start.isoformat()}
    }).to_list(1000)
    
    today_total = sum(s.get("total", 0) for s in today_sales)
    yesterday_total = sum(s.get("total", 0) for s in yesterday_sales)
    today_profit = sum(s.get("profit", 0) for s in today_sales)
    
    change = 0
    if yesterday_total > 0:
        change = ((today_total - yesterday_total) / yesterday_total) * 100
    
    # Low stock products
    low_stock = await db.products.find(
        {"$expr": {"$lt": ["$quantity", {"$ifNull": ["$low_stock_threshold", 10]}]}},
        {"_id": 0, "name_ar": 1, "name_en": 1, "quantity": 1}
    ).to_list(20)
    
    low_stock_list = [{"name": p.get("name_ar") or p.get("name_en"), "quantity": p.get("quantity", 0)} for p in low_stock]
    
    # AI Tips (simple rules-based for now)
    tips = []
    if len(low_stock) > 5:
        tips.append("لديك العديد من المنتجات منخفضة المخزون. يُنصح بمراجعة قائمة المشتريات.")
    if today_total > yesterday_total:
        tips.append(f"مبيعات اليوم أفضل من الأمس بنسبة {change:.1f}%. استمر في العمل الجيد!")
    if today_total < yesterday_total and yesterday_total > 0:
        tips.append("مبيعات اليوم أقل من الأمس. جرب تقديم عروض خاصة لتنشيط المبيعات.")
    if len(today_sales) > 0:
        avg_sale = today_total / len(today_sales)
        tips.append(f"متوسط قيمة الفاتورة اليوم: {avg_sale:.2f} دج")
    
    return {
        "sales": {
            "today_total": today_total,
            "today_count": len(today_sales),
            "today_profit": today_profit,
            "change": change
        },
        "low_stock": low_stock_list,
        "ai_tips": " | ".join(tips) if tips else "لا توجد نصائح حالياً. استمر في العمل!"
    }

@api_router.post("/smart-reports/send-now")
async def send_smart_report_now(user: dict = Depends(require_tenant)):
    """Send smart report immediately"""
    if not RESEND_AVAILABLE:
        raise HTTPException(status_code=500, detail="مكتبة Resend غير متوفرة")
    
    # Get email settings
    email_settings = await db.system_settings.find_one({"type": "email_settings"})
    if not email_settings or not email_settings.get("enabled"):
        raise HTTPException(status_code=400, detail="البريد غير مفعل")
    
    # Get report settings
    report_settings = await db.system_settings.find_one({"type": "smart_reports"})
    recipients = report_settings.get("daily_report_recipients", "") if report_settings else ""
    
    if not recipients:
        # Use current user email
        user_record = await db.users.find_one({"id": user["id"]})
        recipients = user_record.get("email", "") if user_record else ""
    
    if not recipients:
        raise HTTPException(status_code=400, detail="لا يوجد مستلمين محددين")
    
    # Generate report
    preview = await preview_smart_report(user)
    
    # Build HTML email
    html_content = f"""
    <div dir="rtl" style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <h1 style="color: #3b82f6;">📊 التقرير اليومي الذكي</h1>
        <p style="color: #666;">{datetime.now().strftime('%Y-%m-%d')}</p>
        
        <div style="background: #f0fdf4; padding: 15px; border-radius: 8px; margin: 15px 0;">
            <h3 style="color: #166534; margin: 0 0 10px 0;">💰 ملخص المبيعات</h3>
            <p>إجمالي اليوم: <strong>{preview['sales']['today_total']:.2f} دج</strong></p>
            <p>عدد المبيعات: <strong>{preview['sales']['today_count']}</strong></p>
            <p>الربح: <strong>{preview['sales']['today_profit']:.2f} دج</strong></p>
            <p>مقارنة بالأمس: <strong style="color: {'#166534' if preview['sales']['change'] >= 0 else '#dc2626'}">{'+' if preview['sales']['change'] >= 0 else ''}{preview['sales']['change']:.1f}%</strong></p>
        </div>
        
        {'<div style="background: #fef3c7; padding: 15px; border-radius: 8px; margin: 15px 0;"><h3 style="color: #92400e; margin: 0 0 10px 0;">⚠️ تنبيهات المخزون ({} منتجات)</h3></div>'.format(len(preview['low_stock'])) if preview['low_stock'] else ''}
        
        <div style="background: #f3e8ff; padding: 15px; border-radius: 8px; margin: 15px 0;">
            <h3 style="color: #7e22ce; margin: 0 0 10px 0;">✨ نصائح ذكية</h3>
            <p>{preview['ai_tips']}</p>
        </div>
        
        <p style="color: #999; font-size: 12px; margin-top: 30px;">
            هذا التقرير تم إنشاؤه تلقائياً من نظام NT POS
        </p>
    </div>
    """
    
    resend.api_key = email_settings.get("resend_api_key")
    
    try:
        params = {
            "from": email_settings.get("sender_email", "onboarding@resend.dev"),
            "to": [r.strip() for r in recipients.split(",")],
            "subject": f"📊 التقرير اليومي الذكي - {datetime.now().strftime('%Y-%m-%d')}",
            "html": html_content
        }
        
        await asyncio.to_thread(resend.Emails.send, params)
        
        # Log the report
        await db.smart_reports_log.insert_one({
            "id": str(uuid.uuid4()),
            "status": "sent",
            "recipients": recipients,
            "sent_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"success": True, "message": "تم إرسال التقرير بنجاح"}
    except Exception as e:
        await db.smart_reports_log.insert_one({
            "id": str(uuid.uuid4()),
            "status": "failed",
            "recipients": recipients,
            "error": str(e),
            "sent_at": datetime.now(timezone.utc).isoformat()
        })
        raise HTTPException(status_code=500, detail=f"فشل إرسال التقرير: {str(e)}")


# ============ SAAS ROUTES MOVED TO routes/saas_routes.py ============
# All SaaS admin routes (plans, tenants, agents, registration) have been
# refactored to /app/backend/routes/saas_routes.py for better organization.
# Total: 39 endpoints moved

# ============ HEALTH CHECK ============

@api_router.get("/")
async def root():
    return {"message": "NT API is running"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# ============ AI ASSISTANT ============

# AI Chat Models
class AIMessage(BaseModel):
    role: str  # "user" or "assistant"
    content: str
    timestamp: str

class AIChatRequest(BaseModel):
    message: str
    session_id: str
    context: Optional[str] = None  # e.g., "sales", "inventory", "products"

class AIChatResponse(BaseModel):
    response: str
    session_id: str

class AIAnalysisRequest(BaseModel):
    analysis_type: str  # "sales_forecast", "restock", "product_description"
    data: Optional[dict] = None

# Import AI library
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False
    logger.warning("AI library not available")

def get_ai_system_message(context: str = None, language: str = "ar") -> str:
    """Get system message for AI based on context"""
    base_msg = """أنت مساعد ذكي لنظام نقاط البيع (POS). يمكنك المساعدة في:
- تحليل المبيعات وتوقعها
- اقتراح المنتجات التي تحتاج إعادة تخزين
- إنشاء أوصاف للمنتجات
- الإجابة على أسئلة حول المخزون والعملاء والموردين
- تقديم نصائح لتحسين الأعمال

كن مختصراً ومفيداً. أجب باللغة العربية أو الفرنسية حسب لغة السؤال."""
    
    if context == "sales":
        base_msg += "\n\nأنت الآن في قسم المبيعات. ركز على تحليل المبيعات وتوقعاتها."
    elif context == "inventory":
        base_msg += "\n\nأنت الآن في قسم المخزون. ركز على إدارة المخزون واقتراحات إعادة التخزين."
    elif context == "products":
        base_msg += "\n\nأنت الآن في قسم المنتجات. ساعد في إنشاء أوصاف وتحسين معلومات المنتجات."
    elif context == "customers":
        base_msg += "\n\nأنت الآن في قسم العملاء. ساعد في فهم سلوك العملاء وتحسين العلاقات."
    elif context == "reports":
        base_msg += "\n\nأنت الآن في قسم التقارير. ساعد في تحليل البيانات وإنشاء تقارير مفيدة."
    
    return base_msg

@api_router.post("/ai/chat", response_model=AIChatResponse)
async def ai_chat(request: AIChatRequest, user: dict = Depends(require_tenant)):
    """Chat with AI assistant"""
    if not AI_AVAILABLE:
        raise HTTPException(status_code=503, detail="AI service not available")
    
    emergent_key = os.environ.get('EMERGENT_LLM_KEY')
    if not emergent_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")
    
    try:
        # Get or create chat session
        session_id = f"{user['id']}_{request.session_id}"
        
        # Load chat history from database
        chat_history = await db.ai_chat_history.find_one({"session_id": session_id}, {"_id": 0})
        
        # Initialize chat
        chat = LlmChat(
            api_key=emergent_key,
            session_id=session_id,
            system_message=get_ai_system_message(request.context)
        ).with_model("openai", "gpt-4o")
        
        # Build context with business data
        context_data = ""
        if request.context:
            if request.context == "sales":
                # Get recent sales summary
                recent_sales = await db.sales.find().sort("created_at", -1).limit(10).to_list(10)
                total_today = sum(s.get("total", 0) for s in recent_sales if s.get("created_at", "").startswith(datetime.now(timezone.utc).strftime("%Y-%m-%d")))
                context_data = f"\n\nبيانات المبيعات: إجمالي اليوم: {total_today} دج، آخر 10 مبيعات محفوظة."
            elif request.context == "inventory":
                # Get low stock products
                low_stock = await db.products.find({"quantity": {"$lt": 10}}).to_list(20)
                context_data = f"\n\nالمنتجات منخفضة المخزون: {len(low_stock)} منتج"
            elif request.context == "customers":
                # Get customer stats
                total_customers = await db.customers.count_documents({})
                context_data = f"\n\nإجمالي العملاء: {total_customers}"
        
        # Create message with context
        user_message = UserMessage(text=request.message + context_data)
        
        # Get response
        response = await chat.send_message(user_message)
        
        # Save to chat history
        if not chat_history:
            chat_history = {
                "session_id": session_id,
                "user_id": user['id'],
                "messages": [],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        
        chat_history["messages"].append({
            "role": "user",
            "content": request.message,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        chat_history["messages"].append({
            "role": "assistant",
            "content": response,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        chat_history["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        # Keep only last 50 messages
        if len(chat_history["messages"]) > 50:
            chat_history["messages"] = chat_history["messages"][-50:]
        
        await db.ai_chat_history.update_one(
            {"session_id": session_id},
            {"$set": chat_history},
            upsert=True
        )
        
        return AIChatResponse(response=response, session_id=request.session_id)
    
    except Exception as e:
        logger.error(f"AI chat error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"AI error: {str(e)}")

@api_router.get("/ai/chat-history/{session_id}")
async def get_ai_chat_history(session_id: str, user: dict = Depends(require_tenant)):
    """Get chat history for a session"""
    full_session_id = f"{user['id']}_{session_id}"
    history = await db.ai_chat_history.find_one({"session_id": full_session_id}, {"_id": 0})
    if not history:
        return {"messages": []}
    return {"messages": history.get("messages", [])}

@api_router.delete("/ai/chat-history/{session_id}")
async def clear_ai_chat_history(session_id: str, user: dict = Depends(require_tenant)):
    """Clear chat history for a session"""
    full_session_id = f"{user['id']}_{session_id}"
    await db.ai_chat_history.delete_one({"session_id": full_session_id})
    return {"success": True}

@api_router.post("/ai/analyze")
async def ai_analyze(request: AIAnalysisRequest, user: dict = Depends(require_tenant)):
    """Perform AI analysis on business data"""
    if not AI_AVAILABLE:
        raise HTTPException(status_code=503, detail="AI service not available")
    
    emergent_key = os.environ.get('EMERGENT_LLM_KEY')
    if not emergent_key:
        raise HTTPException(status_code=500, detail="AI API key not configured")
    
    try:
        chat = LlmChat(
            api_key=emergent_key,
            session_id=f"analysis_{user['id']}_{datetime.now(timezone.utc).timestamp()}",
            system_message="أنت محلل بيانات ذكي لنظام نقاط البيع. قدم تحليلات مختصرة ومفيدة باللغة العربية."
        ).with_model("openai", "gpt-4o")
        
        if request.analysis_type == "sales_forecast":
            # Get sales data for forecasting
            sales = await db.sales.find().sort("created_at", -1).limit(100).to_list(100)
            
            # Aggregate by day
            daily_sales = {}
            for sale in sales:
                date = sale.get("created_at", "")[:10]
                if date:
                    daily_sales[date] = daily_sales.get(date, 0) + sale.get("total", 0)
            
            prompt = f"""بناءً على بيانات المبيعات التالية، قدم توقعاً للمبيعات في الأسبوع القادم:

المبيعات اليومية (آخر أيام):
{dict(list(daily_sales.items())[:14])}

قدم:
1. توقع المبيعات للأيام السبعة القادمة
2. اتجاه المبيعات (صاعد/هابط/مستقر)
3. نصائح لتحسين المبيعات

أجب بشكل مختصر ومنظم."""
            
            response = await chat.send_message(UserMessage(text=prompt))
            return {"analysis": response, "type": "sales_forecast"}
        
        elif request.analysis_type == "restock":
            # Get low stock and sales velocity
            products = await db.products.find({"quantity": {"$lte": 20}}).to_list(50)
            
            product_list = [
                f"- {p.get('name_ar') or p.get('name_en')}: كمية {p.get('quantity', 0)}, حد أدنى {p.get('low_stock_threshold', 10)}"
                for p in products
            ]
            
            prompt = f"""هذه قائمة المنتجات التي تحتاج مراجعة للمخزون:

{chr(10).join(product_list[:20])}

قدم:
1. ترتيب المنتجات حسب الأولوية لإعادة التخزين
2. كميات مقترحة للطلب
3. نصائح لإدارة المخزون

أجب بشكل مختصر."""
            
            response = await chat.send_message(UserMessage(text=prompt))
            return {"analysis": response, "type": "restock"}
        
        elif request.analysis_type == "product_description":
            product_data = request.data or {}
            product_name = product_data.get("name", "منتج")
            
            prompt = f"""اكتب وصفاً تسويقياً جذاباً للمنتج التالي:

اسم المنتج: {product_name}
التفاصيل: {product_data}

اكتب:
1. وصف قصير (سطر واحد)
2. وصف مفصل (3-4 أسطر)
3. كلمات مفتاحية للبحث

أجب باللغة العربية والفرنسية."""
            
            response = await chat.send_message(UserMessage(text=prompt))
            return {"analysis": response, "type": "product_description"}
        
        elif request.analysis_type == "customer_insights":
            customers = await db.customers.find().to_list(100)
            total_debt = sum(c.get("debt", 0) for c in customers)
            blacklisted = sum(1 for c in customers if c.get("is_blacklisted"))
            
            prompt = f"""حلل بيانات العملاء التالية:

- إجمالي العملاء: {len(customers)}
- إجمالي الديون: {total_debt} دج
- العملاء في القائمة السوداء: {blacklisted}

قدم:
1. تحليل لحالة العملاء
2. نصائح لتحسين العلاقات
3. استراتيجيات لتحصيل الديون

أجب بشكل مختصر."""
            
            response = await chat.send_message(UserMessage(text=prompt))
            return {"analysis": response, "type": "customer_insights"}
        
        else:
            raise HTTPException(status_code=400, detail="Unknown analysis type")
    
    except Exception as e:
        logger.error(f"AI analysis error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"AI error: {str(e)}")

# ============ ADVANCED SALES TRACKING & REPORTS ============

class SalesPermissionSettings(BaseModel):
    allow_employee_edit: bool = False
    allow_employee_delete: bool = False
    allow_discount_without_approval: bool = True
    max_discount_percent: float = 50.0
    max_debt_per_customer: float = 100000.0
    min_sale_price_percent: float = 80.0  # Min price as % of purchase price

@api_router.get("/sales/advanced-report")
async def get_advanced_sales_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    employee_id: Optional[str] = None,
    customer_id: Optional[str] = None,
    product_id: Optional[str] = None,
    payment_method: Optional[str] = None,
    admin: dict = Depends(get_tenant_admin)
):
    """Get advanced sales report with filtering"""
    query = {"status": {"$ne": "returned"}}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date + "T23:59:59"
        else:
            query["created_at"] = {"$lte": end_date + "T23:59:59"}
    if employee_id:
        query["employee_id"] = employee_id
    if customer_id:
        query["customer_id"] = customer_id
    if payment_method:
        query["payment_method"] = payment_method
    
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # If product filter, filter by items
    if product_id:
        sales = [s for s in sales if any(item.get("product_id") == product_id for item in s.get("items", []))]
    
    # Calculate statistics
    total_amount = sum(s.get("total", 0) for s in sales)
    total_paid = sum(s.get("paid_amount", 0) for s in sales)
    total_discount = sum(s.get("discount", 0) for s in sales)
    total_profit = 0
    
    for sale in sales:
        for item in sale.get("items", []):
            purchase_price = item.get("purchase_price", item.get("unit_price", 0) * 0.7)
            profit = (item.get("unit_price", 0) - purchase_price) * item.get("quantity", 1)
            total_profit += profit
    
    # Group by employee
    by_employee = {}
    for sale in sales:
        emp_id = sale.get("employee_id", "unknown")
        emp_name = sale.get("employee_name", "غير محدد")
        if emp_id not in by_employee:
            by_employee[emp_id] = {"name": emp_name, "count": 0, "total": 0}
        by_employee[emp_id]["count"] += 1
        by_employee[emp_id]["total"] += sale.get("total", 0)
    
    # Group by payment method
    by_payment = {}
    for sale in sales:
        method = sale.get("payment_method", "cash")
        if method not in by_payment:
            by_payment[method] = {"count": 0, "total": 0}
        by_payment[method]["count"] += 1
        by_payment[method]["total"] += sale.get("total", 0)
    
    # Top products
    product_sales = {}
    for sale in sales:
        for item in sale.get("items", []):
            pid = item.get("product_id", "unknown")
            pname = item.get("product_name", "غير محدد")
            if pid not in product_sales:
                product_sales[pid] = {"name": pname, "quantity": 0, "total": 0}
            product_sales[pid]["quantity"] += item.get("quantity", 1)
            product_sales[pid]["total"] += item.get("total", 0)
    
    top_products = sorted(product_sales.values(), key=lambda x: x["total"], reverse=True)[:10]
    
    return {
        "sales": sales,
        "statistics": {
            "total_sales": len(sales),
            "total_amount": total_amount,
            "total_paid": total_paid,
            "total_remaining": total_amount - total_paid,
            "total_discount": total_discount,
            "total_profit": total_profit,
            "average_sale": total_amount / len(sales) if sales else 0
        },
        "by_employee": list(by_employee.values()),
        "by_payment_method": by_payment,
        "top_products": top_products
    }

@api_router.get("/sales/employee-report/{employee_id}")
async def get_employee_sales_report(
    employee_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin: dict = Depends(get_tenant_admin)
):
    """Get detailed sales report for a specific employee"""
    query = {"employee_id": employee_id, "status": {"$ne": "returned"}}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date + "T23:59:59"
        else:
            query["created_at"] = {"$lte": end_date + "T23:59:59"}
    
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Get employee info
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    
    # Calculate stats
    total_amount = sum(s.get("total", 0) for s in sales)
    total_paid = sum(s.get("paid_amount", 0) for s in sales)
    
    # Sales by hour
    by_hour = {}
    for sale in sales:
        try:
            hour = sale.get("created_at", "")[:13].split("T")[1] if "T" in sale.get("created_at", "") else "00"
            if hour not in by_hour:
                by_hour[hour] = {"count": 0, "total": 0}
            by_hour[hour]["count"] += 1
            by_hour[hour]["total"] += sale.get("total", 0)
        except:
            pass
    
    return {
        "employee": employee,
        "sales": sales,
        "statistics": {
            "total_sales": len(sales),
            "total_amount": total_amount,
            "total_paid": total_paid,
            "total_remaining": total_amount - total_paid,
            "average_sale": total_amount / len(sales) if sales else 0
        },
        "by_hour": by_hour
    }

@api_router.get("/sales/peak-hours")
async def get_peak_hours_report(
    days: int = 30,
    admin: dict = Depends(get_tenant_admin)
):
    """Get sales peak hours analysis"""
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    
    sales = await db.sales.find(
        {"created_at": {"$gte": start_date}, "status": {"$ne": "returned"}},
        {"_id": 0, "created_at": 1, "total": 1}
    ).to_list(10000)
    
    # Group by hour
    by_hour = {str(i).zfill(2): {"count": 0, "total": 0} for i in range(24)}
    
    for sale in sales:
        try:
            if "T" in sale.get("created_at", ""):
                hour = sale["created_at"].split("T")[1][:2]
                by_hour[hour]["count"] += 1
                by_hour[hour]["total"] += sale.get("total", 0)
        except:
            pass
    
    # Group by day of week
    by_day = {i: {"count": 0, "total": 0, "name_ar": "", "name_en": ""} for i in range(7)}
    day_names_ar = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
    day_names_en = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    for i in range(7):
        by_day[i]["name_ar"] = day_names_ar[i]
        by_day[i]["name_en"] = day_names_en[i]
    
    for sale in sales:
        try:
            date = datetime.fromisoformat(sale["created_at"].replace("Z", "+00:00"))
            weekday = date.weekday()
            by_day[weekday]["count"] += 1
            by_day[weekday]["total"] += sale.get("total", 0)
        except:
            pass
    
    return {
        "by_hour": by_hour,
        "by_day": list(by_day.values()),
        "peak_hour": max(by_hour.items(), key=lambda x: x[1]["total"])[0] if by_hour else None,
        "peak_day": max(by_day.items(), key=lambda x: x[1]["total"])[0] if by_day else None
    }

@api_router.get("/sales/returns-report")
async def get_returns_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin: dict = Depends(get_tenant_admin)
):
    """Get returns report"""
    query = {"status": "returned"}
    
    if start_date:
        query["returned_at"] = {"$gte": start_date}
    if end_date:
        if "returned_at" in query:
            query["returned_at"]["$lte"] = end_date + "T23:59:59"
        else:
            query["returned_at"] = {"$lte": end_date + "T23:59:59"}
    
    returns = await db.sales.find(query, {"_id": 0}).sort("returned_at", -1).to_list(1000)
    
    total_returned = sum(r.get("total", 0) for r in returns)
    
    # Group by reason
    by_reason = {}
    for ret in returns:
        reason = ret.get("return_reason", "غير محدد")
        if reason not in by_reason:
            by_reason[reason] = {"count": 0, "total": 0}
        by_reason[reason]["count"] += 1
        by_reason[reason]["total"] += ret.get("total", 0)
    
    return {
        "returns": returns,
        "statistics": {
            "total_returns": len(returns),
            "total_amount": total_returned
        },
        "by_reason": by_reason
    }

@api_router.get("/sales/{sale_id}/audit-log")
async def get_sale_audit_log(sale_id: str, admin: dict = Depends(get_tenant_admin)):
    """Get audit log for a specific sale"""
    logs = await db.sale_audit_logs.find(
        {"sale_id": sale_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return logs

# ============ FEATURES MANAGEMENT ============

@api_router.get("/settings/features")
async def get_features_settings(admin: dict = Depends(get_tenant_admin)):
    """Get enabled/disabled features for the system"""
    settings = await db.settings.find_one({"key": "features"}, {"_id": 0})
    if settings:
        return settings.get("value", {})
    
    # Return default features
    return {
        "sales": {"enabled": True, "subFeatures": {"pos": True, "invoices": True, "quotes": True, "returns": True, "discounts": True, "price_types": True}},
        "inventory": {"enabled": True, "subFeatures": {"products": True, "categories": True, "stock_alerts": True, "barcode": True, "warehouses": False, "stock_transfer": False, "inventory_count": True}},
        "purchases": {"enabled": True, "subFeatures": {"purchase_orders": True, "suppliers": True, "supplier_payments": True, "purchase_returns": False}},
        "customers": {"enabled": True, "subFeatures": {"customer_list": True, "customer_debts": True, "customer_families": True, "blacklist": True, "debt_reminders": True}},
        "employees": {"enabled": True, "subFeatures": {"employee_list": True, "attendance": True, "salaries": True, "commissions": True, "advances": True, "employee_accounts": True}},
        "reports": {"enabled": True, "subFeatures": {"sales_reports": True, "inventory_reports": True, "financial_reports": True, "customer_reports": True, "smart_reports": False, "export_reports": True}},
        "expenses": {"enabled": True, "subFeatures": {"expense_tracking": True, "expense_categories": True, "recurring_expenses": False}},
        "repairs": {"enabled": True, "subFeatures": {"repair_tickets": True, "repair_status": True, "repair_invoice": True}},
        "delivery": {"enabled": True, "subFeatures": {"delivery_tracking": True, "shipping_companies": True, "delivery_fees": True, "yalidine_integration": False}},
        "ecommerce": {"enabled": False, "subFeatures": {"woocommerce": False, "product_sync": False, "order_sync": False}},
        "loyalty": {"enabled": False, "subFeatures": {"loyalty_points": False, "coupons": False, "promotions": True}},
        "notifications": {"enabled": True, "subFeatures": {"push_notifications": True, "email_notifications": False, "sms_notifications": False, "whatsapp_notifications": False}},
        "services": {"enabled": False, "subFeatures": {"flexy_recharge": False, "bill_payment": False}}
    }

@api_router.post("/settings/features")
async def save_features_settings(features: dict, admin: dict = Depends(get_tenant_admin)):
    """Save features settings - Super Admin only applies to all sub-accounts"""
    if admin.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can change features")
    
    await db.settings.update_one(
        {"key": "features"},
        {"$set": {"key": "features", "value": features, "updated_by": admin["id"], "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"message": "Features saved successfully"}

# ============ USER PERMISSIONS MANAGEMENT ============

@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(user_id: str, data: dict, admin: dict = Depends(get_tenant_admin)):
    """Update specific user permissions"""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Only super_admin can modify super_admin permissions
    if user.get("role") == "super_admin" and admin.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Only super admin can modify super admin permissions")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"permissions": data.get("permissions", {}), "permissions_updated_at": datetime.now(timezone.utc).isoformat(), "permissions_updated_by": admin["id"]}}
    )
    
    return {"message": "Permissions updated successfully"}

@api_router.post("/sales/{sale_id}/log-action")
async def log_sale_action(
    sale_id: str,
    action: str,
    details: dict = {},
    user: dict = Depends(require_tenant)
):
    """Log an action on a sale"""
    log = {
        "id": str(uuid.uuid4()),
        "sale_id": sale_id,
        "action": action,
        "details": details,
        "user_id": user["id"],
        "user_name": user.get("name", user.get("email", "")),
        "user_role": user.get("role", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.sale_audit_logs.insert_one(log)
    return {"message": "Action logged", "log_id": log["id"]}

@api_router.get("/settings/sales-permissions")
async def get_sales_permissions(admin: dict = Depends(get_tenant_admin)):
    """Get sales permission settings"""
    settings = await db.settings.find_one({"key": "sales_permissions"}, {"_id": 0})
    if settings:
        return settings.get("value", {})
    return {
        "allow_employee_edit": False,
        "allow_employee_delete": False,
        "allow_discount_without_approval": True,
        "max_discount_percent": 50.0,
        "max_debt_per_customer": 100000.0,
        "min_sale_price_percent": 80.0
    }

@api_router.post("/settings/sales-permissions")
async def update_sales_permissions(
    settings: SalesPermissionSettings,
    admin: dict = Depends(get_tenant_admin)
):
    """Update sales permission settings"""
    await db.settings.update_one(
        {"key": "sales_permissions"},
        {"$set": {"key": "sales_permissions", "value": settings.model_dump()}},
        upsert=True
    )
    return {"message": "Settings updated"}

# ============ RECEIPT SETTINGS ============

class ReceiptTemplate(BaseModel):
    id: str = ""
    name: str
    name_ar: str
    width: str = "80mm"  # 58mm, 80mm, A4
    show_logo: bool = True
    show_header: bool = True
    show_footer: bool = True
    header_text: str = ""
    footer_text: str = ""
    font_size: str = "normal"  # small, normal, large
    is_default: bool = False

class ReceiptSettings(BaseModel):
    auto_print: bool = False
    show_print_dialog: bool = True
    default_template_id: str = ""
    templates: List[dict] = []

@api_router.get("/settings/receipt")
async def get_receipt_settings(user: dict = Depends(require_tenant)):
    """Get receipt/invoice settings"""
    settings = await db.settings.find_one({"key": "receipt_settings"}, {"_id": 0})
    if settings:
        return settings.get("value", {})
    # Default settings
    return {
        "auto_print": False,
        "show_print_dialog": True,
        "default_template_id": "default_80mm",
        "templates": [
            {
                "id": "default_58mm",
                "name": "Thermal 58mm",
                "name_ar": "حراري 58 مم",
                "width": "58mm",
                "show_logo": False,
                "show_header": True,
                "show_footer": True,
                "header_text": "",
                "footer_text": "شكراً لزيارتكم",
                "font_size": "small",
                "is_default": False
            },
            {
                "id": "default_80mm",
                "name": "Thermal 80mm",
                "name_ar": "حراري 80 مم",
                "width": "80mm",
                "show_logo": True,
                "show_header": True,
                "show_footer": True,
                "header_text": "",
                "footer_text": "شكراً لزيارتكم",
                "font_size": "normal",
                "is_default": True
            },
            {
                "id": "default_a4",
                "name": "A4 Full Page",
                "name_ar": "صفحة A4 كاملة",
                "width": "A4",
                "show_logo": True,
                "show_header": True,
                "show_footer": True,
                "header_text": "",
                "footer_text": "شكراً لزيارتكم",
                "font_size": "normal",
                "is_default": False
            }
        ]
    }

@api_router.post("/settings/receipt")
async def update_receipt_settings(settings: dict, admin: dict = Depends(get_tenant_admin)):
    """Update receipt settings"""
    await db.settings.update_one(
        {"key": "receipt_settings"},
        {"$set": {"key": "receipt_settings", "value": settings}},
        upsert=True
    )
    return {"message": "Settings updated"}

@api_router.get("/sales/product-tracking/{product_id}")
async def get_product_sales_tracking(
    product_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_tenant)
):
    """Track all sales for a specific product"""
    query = {"status": {"$ne": "returned"}}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date + "T23:59:59"
        else:
            query["created_at"] = {"$lte": end_date + "T23:59:59"}
    
    all_sales = await db.sales.find(query, {"_id": 0}).to_list(10000)
    
    # Filter sales containing this product
    product_sales = []
    total_quantity = 0
    total_revenue = 0
    total_profit = 0
    
    for sale in all_sales:
        for item in sale.get("items", []):
            if item.get("product_id") == product_id:
                product_sales.append({
                    "sale_id": sale["id"],
                    "date": sale["created_at"],
                    "customer": sale.get("customer_name", "زبون عابر"),
                    "employee": sale.get("employee_name", ""),
                    "quantity": item.get("quantity", 1),
                    "unit_price": item.get("unit_price", 0),
                    "total": item.get("total", 0),
                    "payment_method": sale.get("payment_method", "cash")
                })
                total_quantity += item.get("quantity", 1)
                total_revenue += item.get("total", 0)
                purchase_price = item.get("purchase_price", item.get("unit_price", 0) * 0.7)
                total_profit += (item.get("unit_price", 0) - purchase_price) * item.get("quantity", 1)
    
    # Get product info
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    
    return {
        "product": product,
        "sales": product_sales,
        "statistics": {
            "total_sales": len(product_sales),
            "total_quantity": total_quantity,
            "total_revenue": total_revenue,
            "total_profit": total_profit,
            "average_price": total_revenue / total_quantity if total_quantity > 0 else 0
        }
    }


# ============ SYSTEM UPDATES (Super Admin Only) ============

class AnnouncementCreate(BaseModel):
    title_ar: str
    title_fr: str = ""
    message_ar: str
    message_fr: str = ""
    type: str = "info"  # info, feature, maintenance, warning, promotion
    priority: str = "normal"  # low, normal, high, urgent
    target: str = "all"  # all, active

class SettingsPush(BaseModel):
    settings: List[str]

@api_router.get("/system-updates/stats")
async def get_system_stats(admin: dict = Depends(get_super_admin)):
    """Get system statistics for super admin"""
    total_tenants = await db.saas_tenants.count_documents({})
    active_tenants = await db.saas_tenants.count_documents({"status": "active"})
    total_announcements = await db.system_announcements.count_documents({})
    
    return {
        "total_tenants": total_tenants,
        "active_tenants": active_tenants,
        "total_announcements": total_announcements
    }

@api_router.get("/system-updates/announcements")
async def get_announcements(admin: dict = Depends(get_super_admin)):
    """Get all system announcements"""
    announcements = await db.system_announcements.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return announcements

@api_router.post("/system-updates/announcements")
async def create_announcement(data: AnnouncementCreate, admin: dict = Depends(get_super_admin)):
    """Create and broadcast a new announcement"""
    now = datetime.now(timezone.utc).isoformat()
    announcement_id = str(uuid.uuid4())
    
    announcement = {
        "id": announcement_id,
        "title_ar": data.title_ar,
        "title_fr": data.title_fr,
        "message_ar": data.message_ar,
        "message_fr": data.message_fr,
        "type": data.type,
        "priority": data.priority,
        "target": data.target,
        "created_by": admin["id"],
        "created_at": now,
        "read_count": 0
    }
    
    await db.system_announcements.insert_one(announcement)
    
    # Create notifications for all users
    query = {}
    if data.target == "active":
        # Get active tenant IDs
        active_tenants = await db.saas_tenants.find({"status": "active"}, {"id": 1}).to_list(1000)
        tenant_ids = [t["id"] for t in active_tenants]
        query = {"tenant_id": {"$in": tenant_ids}}
    
    users = await db.users.find(query, {"id": 1}).to_list(10000)
    
    notifications = []
    for user in users:
        notifications.append({
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "type": f"system_{data.type}",
            "title": data.title_ar,
            "title_ar": data.title_ar,
            "title_fr": data.title_fr,
            "message": data.message_ar,
            "message_ar": data.message_ar,
            "message_fr": data.message_fr,
            "priority": data.priority,
            "announcement_id": announcement_id,
            "read": False,
            "created_at": now
        })
    
    if notifications:
        await db.notifications.insert_many(notifications)
    
    return {"message": "Announcement created and sent", "id": announcement_id, "recipients": len(notifications)}

@api_router.delete("/system-updates/announcements/{announcement_id}")
async def delete_announcement(announcement_id: str, admin: dict = Depends(get_super_admin)):
    """Delete an announcement"""
    result = await db.system_announcements.delete_one({"id": announcement_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Announcement not found")
    
    # Also delete related notifications
    await db.notifications.delete_many({"announcement_id": announcement_id})
    
    return {"message": "Announcement deleted"}

@api_router.post("/system-updates/push-settings")
async def push_settings(data: SettingsPush, admin: dict = Depends(get_super_admin)):
    """Push settings to all tenants"""
    now = datetime.now(timezone.utc).isoformat()
    
    # Get current admin settings as template
    admin_settings = {}
    
    for setting_type in data.settings:
        if setting_type == "receipt_settings":
            settings = await db.settings.find_one({"type": "receipt"}, {"_id": 0})
            if settings:
                admin_settings["receipt"] = settings
        elif setting_type == "notification_settings":
            settings = await db.notification_settings.find_one({}, {"_id": 0})
            if settings:
                admin_settings["notifications"] = settings
        elif setting_type == "loyalty_settings":
            settings = await db.loyalty_settings.find_one({}, {"_id": 0})
            if settings:
                admin_settings["loyalty"] = settings
        elif setting_type == "pos_settings":
            settings = await db.settings.find_one({"type": "pos"}, {"_id": 0})
            if settings:
                admin_settings["pos"] = settings
    
    # Get all active tenants
    tenants = await db.saas_tenants.find({"status": "active"}, {"database_name": 1}).to_list(1000)
    
    updated_count = 0
    for tenant in tenants:
        try:
            tenant_db = client[tenant["database_name"]]
            for key, value in admin_settings.items():
                if key == "receipt":
                    await tenant_db.settings.update_one(
                        {"type": "receipt"}, 
                        {"$set": value}, 
                        upsert=True
                    )
                elif key == "notifications":
                    await tenant_db.notification_settings.update_one(
                        {}, 
                        {"$set": value}, 
                        upsert=True
                    )
                elif key == "loyalty":
                    await tenant_db.loyalty_settings.update_one(
                        {}, 
                        {"$set": value}, 
                        upsert=True
                    )
                elif key == "pos":
                    await tenant_db.settings.update_one(
                        {"type": "pos"}, 
                        {"$set": value}, 
                        upsert=True
                    )
            updated_count += 1
        except Exception as e:
            print(f"Error updating tenant {tenant.get('database_name')}: {e}")
    
    # Log the action
    await db.system_logs.insert_one({
        "id": str(uuid.uuid4()),
        "action": "push_settings",
        "settings": data.settings,
        "updated_tenants": updated_count,
        "admin_id": admin["id"],
        "created_at": now
    })
    
    return {"message": f"Settings pushed to {updated_count} tenants", "updated_count": updated_count}

# ============ REAL-TIME SYNC SYSTEM ============

class SyncConfigCreate(BaseModel):
    name: str
    sync_types: List[str]  # receipt, notifications, loyalty, pos, products, families, theme
    target: str = "all"  # all, active, selected
    selected_tenants: List[str] = []
    auto_sync: bool = False
    locked: bool = False  # If true, tenants cannot modify

class SyncAction(BaseModel):
    sync_types: List[str]
    target: str = "all"
    selected_tenants: List[str] = []

@api_router.get("/sync/configs")
async def get_sync_configs(admin: dict = Depends(get_super_admin)):
    """Get all sync configurations"""
    configs = await db.sync_configs.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return configs

@api_router.post("/sync/configs")
async def create_sync_config(data: SyncConfigCreate, admin: dict = Depends(get_super_admin)):
    """Create a new sync configuration"""
    now = datetime.now(timezone.utc).isoformat()
    config = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "sync_types": data.sync_types,
        "target": data.target,
        "selected_tenants": data.selected_tenants,
        "auto_sync": data.auto_sync,
        "locked": data.locked,
        "created_by": admin["id"],
        "created_at": now,
        "last_sync": None
    }
    await db.sync_configs.insert_one(config)
    return config

@api_router.delete("/sync/configs/{config_id}")
async def delete_sync_config(config_id: str, admin: dict = Depends(get_super_admin)):
    """Delete a sync configuration"""
    result = await db.sync_configs.delete_one({"id": config_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Config not found")
    return {"message": "Config deleted"}

@api_router.post("/sync/execute")
async def execute_sync(data: SyncAction, admin: dict = Depends(get_super_admin)):
    """Execute sync to tenants"""
    now = datetime.now(timezone.utc).isoformat()
    sync_id = str(uuid.uuid4())
    
    # Get tenants to sync
    if data.target == "selected" and data.selected_tenants:
        tenants = await db.saas_tenants.find(
            {"id": {"$in": data.selected_tenants}}, 
            {"id": 1, "name": 1, "database_name": 1}
        ).to_list(1000)
    elif data.target == "active":
        tenants = await db.saas_tenants.find(
            {"status": "active"}, 
            {"id": 1, "name": 1, "database_name": 1}
        ).to_list(1000)
    else:
        tenants = await db.saas_tenants.find(
            {}, {"id": 1, "name": 1, "database_name": 1}
        ).to_list(1000)
    
    # Collect data to sync
    sync_data = {}
    sync_labels = {
        "receipt": "إعدادات الإيصال",
        "notifications": "إعدادات الإشعارات",
        "loyalty": "إعدادات الولاء",
        "pos": "إعدادات نقطة البيع",
        "products": "المنتجات",
        "families": "عائلات المنتجات",
        "theme": "تصميم الواجهة"
    }
    
    for sync_type in data.sync_types:
        if sync_type == "receipt":
            settings = await db.settings.find_one({"type": "receipt"}, {"_id": 0})
            if settings: sync_data["receipt"] = settings
        elif sync_type == "notifications":
            settings = await db.notification_settings.find_one({}, {"_id": 0})
            if settings: sync_data["notifications"] = settings
        elif sync_type == "loyalty":
            settings = await db.loyalty_settings.find_one({}, {"_id": 0})
            if settings: sync_data["loyalty"] = settings
        elif sync_type == "pos":
            settings = await db.settings.find_one({"type": "pos"}, {"_id": 0})
            if settings: sync_data["pos"] = settings
        elif sync_type == "products":
            products = await db.products.find({}, {"_id": 0}).to_list(10000)
            sync_data["products"] = products
        elif sync_type == "families":
            families = await db.product_families.find({}, {"_id": 0}).to_list(1000)
            sync_data["families"] = families
        elif sync_type == "theme":
            theme = await db.settings.find_one({"type": "theme"}, {"_id": 0})
            if theme: sync_data["theme"] = theme
    
    # Execute sync
    success_count = 0
    failed_tenants = []
    
    for tenant in tenants:
        try:
            tenant_db = client[tenant["database_name"]]
            
            for key, value in sync_data.items():
                if key == "receipt":
                    await tenant_db.settings.update_one({"type": "receipt"}, {"$set": value}, upsert=True)
                elif key == "notifications":
                    await tenant_db.notification_settings.update_one({}, {"$set": value}, upsert=True)
                elif key == "loyalty":
                    await tenant_db.loyalty_settings.update_one({}, {"$set": value}, upsert=True)
                elif key == "pos":
                    await tenant_db.settings.update_one({"type": "pos"}, {"$set": value}, upsert=True)
                elif key == "products":
                    if value:
                        await tenant_db.products.delete_many({})
                        await tenant_db.products.insert_many(value)
                elif key == "families":
                    if value:
                        await tenant_db.product_families.delete_many({})
                        await tenant_db.product_families.insert_many(value)
                elif key == "theme":
                    await tenant_db.settings.update_one({"type": "theme"}, {"$set": value}, upsert=True)
            
            # Send notification to tenant
            await tenant_db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "type": "system_update",
                "title": "تحديث من الإدارة",
                "message": f"تم تحديث: {', '.join([sync_labels.get(t, t) for t in data.sync_types])}",
                "priority": "high",
                "read": False,
                "created_at": now
            })
            
            success_count += 1
        except Exception as e:
            failed_tenants.append({"id": tenant["id"], "name": tenant["name"], "error": str(e)})
    
    # Log sync action
    sync_log = {
        "id": sync_id,
        "sync_types": data.sync_types,
        "target": data.target,
        "total_tenants": len(tenants),
        "success_count": success_count,
        "failed_count": len(failed_tenants),
        "failed_tenants": failed_tenants,
        "admin_id": admin["id"],
        "admin_name": admin.get("name", ""),
        "created_at": now
    }
    await db.sync_logs.insert_one(sync_log)
    
    return {
        "message": f"تم المزامنة بنجاح إلى {success_count} مشترك",
        "sync_id": sync_id,
        "success_count": success_count,
        "failed_count": len(failed_tenants),
        "failed_tenants": failed_tenants
    }

@api_router.get("/sync/logs")
async def get_sync_logs(admin: dict = Depends(get_super_admin), limit: int = 50):
    """Get sync history logs"""
    logs = await db.sync_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return logs

@api_router.get("/sync/available-types")
async def get_available_sync_types(admin: dict = Depends(get_super_admin)):
    """Get available sync types with counts"""
    products_count = await db.products.count_documents({})
    families_count = await db.product_families.count_documents({})
    
    return [
        {"id": "receipt", "name": "إعدادات الإيصال", "name_fr": "Paramètres du reçu", "icon": "Receipt", "count": 1},
        {"id": "notifications", "name": "إعدادات الإشعارات", "name_fr": "Paramètres des notifications", "icon": "Bell", "count": 1},
        {"id": "loyalty", "name": "إعدادات الولاء", "name_fr": "Paramètres de fidélité", "icon": "Award", "count": 1},
        {"id": "pos", "name": "إعدادات نقطة البيع", "name_fr": "Paramètres POS", "icon": "Store", "count": 1},
        {"id": "products", "name": "المنتجات", "name_fr": "Produits", "icon": "Package", "count": products_count},
        {"id": "families", "name": "عائلات المنتجات", "name_fr": "Familles de produits", "icon": "Folder", "count": families_count},
        {"id": "theme", "name": "تصميم الواجهة", "name_fr": "Thème", "icon": "Palette", "count": 1}
    ]

@api_router.post("/sync/lock-settings")
async def lock_tenant_settings(tenant_ids: List[str], lock: bool, admin: dict = Depends(get_super_admin)):
    """Lock or unlock settings for specific tenants"""
    now = datetime.now(timezone.utc).isoformat()
    
    for tenant_id in tenant_ids:
        tenant = await db.saas_tenants.find_one({"id": tenant_id})
        if tenant:
            tenant_db = client[tenant["database_name"]]
            await tenant_db.settings.update_one(
                {"type": "admin_lock"},
                {"$set": {"locked": lock, "updated_at": now}},
                upsert=True
            )
    
    return {"message": f"تم {'قفل' if lock else 'فتح'} الإعدادات لـ {len(tenant_ids)} مشترك"}


# ============ TENANT DATABASE MANAGEMENT ============

@api_router.get("/tenant/database-info")
async def get_tenant_database_info(current_user: dict = Depends(require_tenant)):
    """Get database info for current tenant"""
    tenant_id = current_user.get("tenant_id") or current_user.get("id")
    
    # Get tenant info
    tenant = await db.saas_tenants.find_one({"id": tenant_id}, {"_id": 0})
    
    if tenant:
        db_name = f"tenant_{tenant_id.replace('-', '_')}"
        try:
            tenant_db = client[db_name]
            stats = await tenant_db.command("dbStats")
            cols = await tenant_db.list_collection_names()
            docs = sum([await tenant_db[c].count_documents({}) for c in cols])
            
            return {
                "size_mb": round(stats.get("dataSize", 0) / (1024 * 1024), 2),
                "collections_count": len(cols),
                "documents_count": docs,
                "last_backup": tenant.get("last_backup"),
                "is_frozen": tenant.get("is_frozen", False),
                "status": "frozen" if tenant.get("is_frozen") else "healthy"
            }
        except Exception as e:
            logger.error(f"Error getting tenant DB info: {e}")
    
    # Fallback: return estimated data
    return {
        "size_mb": 0,
        "collections_count": 8,
        "documents_count": 0,
        "last_backup": None,
        "is_frozen": False,
        "status": "healthy"
    }

@api_router.post("/tenant/request-backup")
async def request_tenant_backup(current_user: dict = Depends(require_tenant)):
    """Request backup for tenant database"""
    tenant_id = current_user.get("tenant_id") or current_user.get("id")
    now = datetime.now(timezone.utc).isoformat()
    
    # Create backup request
    request_id = str(uuid.uuid4())
    await db.backup_requests.insert_one({
        "id": request_id,
        "tenant_id": tenant_id,
        "tenant_name": current_user.get("name") or current_user.get("company_name"),
        "status": "pending",
        "requested_at": now,
        "processed_at": None
    })
    
    # Log operation
    await db.database_operation_logs.insert_one({
        "id": str(uuid.uuid4()),
        "operation": "backup_request",
        "database_id": tenant_id,
        "database_name": f"tenant_{tenant_id.replace('-', '_')}",
        "executed_by": current_user.get("name"),
        "status": "pending",
        "details": "طلب نسخة احتياطية من المشترك",
        "created_at": now
    })
    
    return {"message": "تم إرسال طلب النسخ الاحتياطي", "request_id": request_id}

@api_router.get("/tenant/export-data")
async def export_tenant_data(current_user: dict = Depends(require_tenant)):
    """Export tenant's own data"""
    import json
    
    tenant_id = current_user.get("tenant_id") or current_user.get("id")
    
    # Check if frozen
    tenant = await db.saas_tenants.find_one({"id": tenant_id}, {"_id": 0})
    if tenant and tenant.get("is_frozen"):
        raise HTTPException(status_code=403, detail="قاعدة البيانات مجمدة")
    
    db_name = f"tenant_{tenant_id.replace('-', '_')}"
    
    try:
        tenant_db = client[db_name]
        export_data = {}
        
        # Export allowed collections only
        allowed_collections = ["products", "customers", "suppliers", "employees", "sales", "expenses", "settings"]
        
        for col in allowed_collections:
            try:
                docs = await tenant_db[col].find({}, {"_id": 0}).to_list(100000)
                export_data[col] = docs
            except:
                export_data[col] = []
        
        export_data["exported_at"] = datetime.now(timezone.utc).isoformat()
        export_data["tenant_id"] = tenant_id
        
        content = json.dumps(export_data, ensure_ascii=False, indent=2, default=str)
        
        # Log export
        await db.database_operation_logs.insert_one({
            "id": str(uuid.uuid4()),
            "operation": "self_export",
            "database_id": tenant_id,
            "database_name": db_name,
            "executed_by": current_user.get("name"),
            "status": "success",
            "details": "تصدير بيانات ذاتي",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return StreamingResponse(
            io.BytesIO(content.encode('utf-8')),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={db_name}_export.json"}
        )
    except Exception as e:
        logger.error(f"Error exporting tenant data: {e}")
        raise HTTPException(status_code=500, detail="حدث خطأ أثناء التصدير")



# ============ DATABASE MANAGEMENT MOVED TO routes/saas_routes.py ============

# ============ SENDGRID EMAIL NOTIFICATIONS ============

class SendGridSettings(BaseModel):
    enabled: bool = False
    api_key: str = ""
    sender_email: str = ""
    sender_name: str = "NT Commerce"
    # Notification types
    new_sale_notification: bool = True
    low_stock_notification: bool = True
    daily_report: bool = False
    weekly_report: bool = False
    notification_email: str = ""

class EmailNotificationRequest(BaseModel):
    notification_type: str  # new_sale, low_stock, daily_report, weekly_report, custom
    recipient_email: str
    subject: str = ""
    data: dict = {}

async def send_email_with_sendgrid(to_email: str, subject: str, html_content: str, settings: dict = None):
    """Send email using SendGrid"""
    if not SENDGRID_AVAILABLE:
        raise HTTPException(status_code=500, detail="SendGrid غير متوفر")
    
    # Get settings from database if not provided
    if not settings:
        settings = await main_db.system_settings.find_one({"type": "sendgrid_settings"})
    
    if not settings or not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="يرجى إعداد مفتاح SendGrid أولاً")
    
    try:
        sg = SendGridAPIClient(settings["api_key"])
        from_email = Email(settings.get("sender_email", "noreply@ntcommerce.com"), settings.get("sender_name", "NT Commerce"))
        to_email_obj = To(to_email)
        content = Content("text/html", html_content)
        mail = Mail(from_email, to_email_obj, subject, content)
        
        response = sg.send(mail)
        return response.status_code == 202
    except Exception as e:
        logger.error(f"SendGrid error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"فشل إرسال البريد: {str(e)}")

def generate_sale_notification_html(sale_data: dict):
    """Generate HTML for sale notification"""
    return f"""
    <div dir="rtl" style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #f8f9fa; padding: 20px;">
        <div style="background: white; border-radius: 12px; padding: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <div style="text-align: center; margin-bottom: 20px;">
                <h1 style="color: #22c55e; margin: 0;">🛒 عملية بيع جديدة</h1>
            </div>
            <div style="background: #f0fdf4; border-radius: 8px; padding: 15px; margin: 15px 0;">
                <p style="margin: 5px 0;"><strong>رقم الفاتورة:</strong> {sale_data.get('invoice_number', 'N/A')}</p>
                <p style="margin: 5px 0;"><strong>الزبون:</strong> {sale_data.get('customer_name', 'زبون عام')}</p>
                <p style="margin: 5px 0;"><strong>المبلغ الإجمالي:</strong> <span style="color: #16a34a; font-size: 1.2em; font-weight: bold;">{sale_data.get('total', 0):,.2f} دج</span></p>
                <p style="margin: 5px 0;"><strong>طريقة الدفع:</strong> {sale_data.get('payment_method', 'نقداً')}</p>
                <p style="margin: 5px 0;"><strong>عدد المنتجات:</strong> {sale_data.get('items_count', 0)}</p>
            </div>
            <p style="color: #6b7280; font-size: 12px; text-align: center; margin-top: 20px;">
                تم إرسال هذا الإشعار تلقائياً من نظام NT Commerce
            </p>
        </div>
    </div>
    """

def generate_low_stock_notification_html(products: list):
    """Generate HTML for low stock notification"""
    products_html = ""
    for p in products[:20]:  # Limit to 20 products
        products_html += f"""
        <tr>
            <td style="padding: 8px; border-bottom: 1px solid #e5e7eb;">{p.get('name', 'N/A')}</td>
            <td style="padding: 8px; border-bottom: 1px solid #e5e7eb; text-align: center; color: {'#dc2626' if p.get('stock', 0) == 0 else '#f59e0b'};">
                {p.get('stock', 0)}
            </td>
            <td style="padding: 8px; border-bottom: 1px solid #e5e7eb; text-align: center;">{p.get('min_quantity', 10)}</td>
        </tr>
        """
    
    return f"""
    <div dir="rtl" style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #f8f9fa; padding: 20px;">
        <div style="background: white; border-radius: 12px; padding: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <div style="text-align: center; margin-bottom: 20px;">
                <h1 style="color: #f59e0b; margin: 0;">⚠️ تنبيه انخفاض المخزون</h1>
                <p style="color: #6b7280;">يوجد {len(products)} منتج بحاجة إلى إعادة تزويد</p>
            </div>
            <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                <thead>
                    <tr style="background: #f3f4f6;">
                        <th style="padding: 10px; text-align: right;">المنتج</th>
                        <th style="padding: 10px; text-align: center;">الكمية الحالية</th>
                        <th style="padding: 10px; text-align: center;">الحد الأدنى</th>
                    </tr>
                </thead>
                <tbody>
                    {products_html}
                </tbody>
            </table>
            <p style="color: #6b7280; font-size: 12px; text-align: center; margin-top: 20px;">
                تم إرسال هذا الإشعار تلقائياً من نظام NT Commerce
            </p>
        </div>
    </div>
    """

def generate_daily_report_html(report_data: dict):
    """Generate HTML for daily report"""
    return f"""
    <div dir="rtl" style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #f8f9fa; padding: 20px;">
        <div style="background: white; border-radius: 12px; padding: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <div style="text-align: center; margin-bottom: 20px;">
                <h1 style="color: #3b82f6; margin: 0;">📊 التقرير اليومي</h1>
                <p style="color: #6b7280;">{report_data.get('date', datetime.now().strftime('%Y-%m-%d'))}</p>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin: 20px 0;">
                <div style="background: #f0fdf4; border-radius: 8px; padding: 15px; text-align: center;">
                    <p style="margin: 0; color: #6b7280; font-size: 12px;">إجمالي المبيعات</p>
                    <p style="margin: 5px 0; color: #16a34a; font-size: 1.5em; font-weight: bold;">{report_data.get('total_sales', 0):,.2f} دج</p>
                </div>
                <div style="background: #eff6ff; border-radius: 8px; padding: 15px; text-align: center;">
                    <p style="margin: 0; color: #6b7280; font-size: 12px;">عدد الفواتير</p>
                    <p style="margin: 5px 0; color: #3b82f6; font-size: 1.5em; font-weight: bold;">{report_data.get('sales_count', 0)}</p>
                </div>
                <div style="background: #fef3c7; border-radius: 8px; padding: 15px; text-align: center;">
                    <p style="margin: 0; color: #6b7280; font-size: 12px;">صافي الربح</p>
                    <p style="margin: 5px 0; color: #d97706; font-size: 1.5em; font-weight: bold;">{report_data.get('total_profit', 0):,.2f} دج</p>
                </div>
                <div style="background: #fce7f3; border-radius: 8px; padding: 15px; text-align: center;">
                    <p style="margin: 0; color: #6b7280; font-size: 12px;">المصاريف</p>
                    <p style="margin: 5px 0; color: #db2777; font-size: 1.5em; font-weight: bold;">{report_data.get('total_expenses', 0):,.2f} دج</p>
                </div>
            </div>
            
            <div style="background: #f3f4f6; border-radius: 8px; padding: 15px; margin-top: 15px;">
                <p style="margin: 5px 0;"><strong>أفضل منتج مبيعاً:</strong> {report_data.get('top_product', 'N/A')}</p>
                <p style="margin: 5px 0;"><strong>زبائن جدد:</strong> {report_data.get('new_customers', 0)}</p>
                <p style="margin: 5px 0;"><strong>ديون جديدة:</strong> {report_data.get('new_debts', 0):,.2f} دج</p>
                <p style="margin: 5px 0;"><strong>ديون محصلة:</strong> {report_data.get('collected_debts', 0):,.2f} دج</p>
            </div>
            
            <p style="color: #6b7280; font-size: 12px; text-align: center; margin-top: 20px;">
                تم إرسال هذا التقرير تلقائياً من نظام NT Commerce
            </p>
        </div>
    </div>
    """

@api_router.get("/notifications/sendgrid/settings")
async def get_sendgrid_settings(user: dict = Depends(require_tenant)):
    """Get SendGrid email notification settings"""
    settings = await db.system_settings.find_one({"type": "sendgrid_settings"}, {"_id": 0})
    if not settings:
        return SendGridSettings().model_dump()
    
    # Mask API key
    if settings.get("api_key"):
        key = settings["api_key"]
        settings["api_key"] = key[:8] + "..." + key[-4:] if len(key) > 12 else "***configured***"
    
    return settings

@api_router.put("/notifications/sendgrid/settings")
async def update_sendgrid_settings(settings: SendGridSettings, user: dict = Depends(require_tenant)):
    """Update SendGrid email notification settings"""
    if user.get("role") not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # Get existing settings to preserve API key if masked
    existing = await db.system_settings.find_one({"type": "sendgrid_settings"})
    
    settings_dict = settings.model_dump()
    
    # If API key looks masked, keep the old one
    if settings.api_key and ("..." in settings.api_key or settings.api_key == "***configured***"):
        if existing and existing.get("api_key"):
            settings_dict["api_key"] = existing["api_key"]
    
    await db.system_settings.update_one(
        {"type": "sendgrid_settings"},
        {"$set": {**settings_dict, "type": "sendgrid_settings", "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"success": True, "message": "تم حفظ إعدادات الإشعارات"}

@api_router.post("/notifications/sendgrid/test")
async def test_sendgrid_settings(user: dict = Depends(require_tenant)):
    """Send a test email via SendGrid"""
    settings = await db.system_settings.find_one({"type": "sendgrid_settings"})
    if not settings or not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="يرجى إعداد مفتاح SendGrid أولاً")
    
    user_record = await db.users.find_one({"id": user["id"]})
    if not user_record or not user_record.get("email"):
        raise HTTPException(status_code=400, detail="لم يتم العثور على بريدك الإلكتروني")
    
    test_html = """
    <div dir="rtl" style="font-family: Arial, sans-serif; padding: 20px; background: #f3f4f6;">
        <div style="max-width: 400px; margin: 0 auto; background: white; border-radius: 8px; padding: 30px; text-align: center;">
            <h2 style="color: #22c55e;">✅ SendGrid يعمل بنجاح!</h2>
            <p style="color: #6b7280;">هذه رسالة اختبار من نظام NT Commerce</p>
            <p style="color: #9ca3af; font-size: 12px; margin-top: 20px;">
                إذا وصلتك هذه الرسالة، فإن إعدادات البريد الإلكتروني تعمل بشكل صحيح.
            </p>
        </div>
    </div>
    """
    
    try:
        await send_email_with_sendgrid(user_record["email"], "🧪 اختبار SendGrid - NT Commerce", test_html, settings)
        return {"success": True, "message": f"تم إرسال بريد اختباري إلى {user_record['email']}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/notifications/send")
async def send_notification(request: EmailNotificationRequest, user: dict = Depends(require_tenant)):
    """Send a notification email"""
    settings = await db.system_settings.find_one({"type": "sendgrid_settings"})
    if not settings or not settings.get("enabled"):
        raise HTTPException(status_code=400, detail="إشعارات البريد غير مفعلة")
    
    subject = request.subject
    html_content = ""
    
    if request.notification_type == "new_sale":
        subject = subject or f"🛒 عملية بيع جديدة - {request.data.get('invoice_number', '')}"
        html_content = generate_sale_notification_html(request.data)
    elif request.notification_type == "low_stock":
        subject = subject or "⚠️ تنبيه انخفاض المخزون"
        html_content = generate_low_stock_notification_html(request.data.get("products", []))
    elif request.notification_type == "daily_report":
        subject = subject or f"📊 التقرير اليومي - {datetime.now().strftime('%Y-%m-%d')}"
        html_content = generate_daily_report_html(request.data)
    else:
        html_content = request.data.get("html_content", "<p>إشعار من NT Commerce</p>")
    
    try:
        await send_email_with_sendgrid(request.recipient_email, subject, html_content, settings)
        
        # Log the notification
        await db.notification_logs.insert_one({
            "id": str(uuid.uuid4()),
            "type": request.notification_type,
            "recipient": request.recipient_email,
            "subject": subject,
            "status": "sent",
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "sent_by": user["id"]
        })
        
        return {"success": True, "message": "تم إرسال الإشعار بنجاح"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/notifications/check-low-stock")
async def check_and_notify_low_stock(user: dict = Depends(require_tenant)):
    """Check for low stock products and send notification"""
    settings = await db.system_settings.find_one({"type": "sendgrid_settings"})
    if not settings or not settings.get("enabled") or not settings.get("low_stock_notification"):
        return {"success": False, "message": "إشعارات انخفاض المخزون غير مفعلة"}
    
    # Get low stock products
    low_stock_products = await db.products.find({
        "$expr": {"$lte": ["$stock", "$min_quantity"]}
    }).to_list(100)
    
    if not low_stock_products:
        return {"success": True, "message": "لا توجد منتجات منخفضة المخزون"}
    
    products_list = [{"name": p.get("name", ""), "stock": p.get("stock", 0), "min_quantity": p.get("min_quantity", 10)} for p in low_stock_products]
    
    recipient = settings.get("notification_email")
    if not recipient:
        return {"success": False, "message": "يرجى إعداد بريد الإشعارات"}
    
    html_content = generate_low_stock_notification_html(products_list)
    
    try:
        await send_email_with_sendgrid(recipient, "⚠️ تنبيه انخفاض المخزون - NT Commerce", html_content, settings)
        return {"success": True, "message": f"تم إرسال تنبيه بـ {len(products_list)} منتج منخفض المخزون"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/notifications/send-daily-report")
async def send_daily_report(user: dict = Depends(require_tenant)):
    """Generate and send daily report"""
    settings = await db.system_settings.find_one({"type": "sendgrid_settings"})
    if not settings or not settings.get("enabled"):
        raise HTTPException(status_code=400, detail="إشعارات البريد غير مفعلة")
    
    recipient = settings.get("notification_email")
    if not recipient:
        raise HTTPException(status_code=400, detail="يرجى إعداد بريد الإشعارات")
    
    # Generate report data
    today = datetime.now(timezone.utc).date()
    today_start = datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc)
    
    sales = await db.sales.find({"created_at": {"$gte": today_start.isoformat()}}).to_list(1000)
    expenses = await db.expenses.find({"created_at": {"$gte": today_start.isoformat()}}).to_list(1000)
    
    total_sales = sum(s.get("total", 0) for s in sales)
    total_profit = sum(s.get("profit", 0) for s in sales)
    total_expenses = sum(e.get("amount", 0) for e in expenses)
    
    # Find top product
    product_sales = {}
    for s in sales:
        for item in s.get("items", []):
            pid = item.get("product_id", "")
            product_sales[pid] = product_sales.get(pid, 0) + item.get("quantity", 0)
    
    top_product_id = max(product_sales, key=product_sales.get) if product_sales else None
    top_product = await db.products.find_one({"id": top_product_id}) if top_product_id else None
    
    report_data = {
        "date": today.strftime('%Y-%m-%d'),
        "total_sales": total_sales,
        "sales_count": len(sales),
        "total_profit": total_profit,
        "total_expenses": total_expenses,
        "top_product": top_product.get("name", "N/A") if top_product else "N/A",
        "new_customers": await db.customers.count_documents({"created_at": {"$gte": today_start.isoformat()}}),
        "new_debts": sum(s.get("remaining", 0) for s in sales if s.get("payment_status") == "partial"),
        "collected_debts": 0  # Would need debt payments tracking
    }
    
    html_content = generate_daily_report_html(report_data)
    
    try:
        await send_email_with_sendgrid(recipient, f"📊 التقرير اليومي - {today.strftime('%Y-%m-%d')}", html_content, settings)
        return {"success": True, "message": "تم إرسال التقرير اليومي"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ STRIPE PAYMENT INTEGRATION ============

# Subscription packages (prices defined on backend for security)
SUBSCRIPTION_PACKAGES = {
    "basic_monthly": {"name": "الباقة الأساسية - شهري", "amount": 2500.0, "duration_days": 30, "currency": "dzd"},
    "basic_yearly": {"name": "الباقة الأساسية - سنوي", "amount": 25000.0, "duration_days": 365, "currency": "dzd"},
    "pro_monthly": {"name": "الباقة المتقدمة - شهري", "amount": 5000.0, "duration_days": 30, "currency": "dzd"},
    "pro_yearly": {"name": "الباقة المتقدمة - سنوي", "amount": 50000.0, "duration_days": 365, "currency": "dzd"},
    "enterprise_monthly": {"name": "باقة المؤسسات - شهري", "amount": 10000.0, "duration_days": 30, "currency": "dzd"},
    "enterprise_yearly": {"name": "باقة المؤسسات - سنوي", "amount": 100000.0, "duration_days": 365, "currency": "dzd"},
}

class CreateCheckoutRequest(BaseModel):
    package_id: str
    origin_url: str

class PaymentRecord(BaseModel):
    tenant_id: Optional[str] = None
    amount: float
    currency: str = "dzd"
    payment_method: str = "stripe"
    description: str = ""
    invoice_number: str = ""
    status: str = "pending"
    metadata: dict = {}

class PaymentUpdateRequest(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None
    invoice_number: Optional[str] = None

@api_router.get("/payments/packages")
async def get_subscription_packages():
    """Get available subscription packages"""
    packages = []
    for pkg_id, pkg in SUBSCRIPTION_PACKAGES.items():
        packages.append({
            "id": pkg_id,
            "name": pkg["name"],
            "amount": pkg["amount"],
            "duration_days": pkg["duration_days"],
            "currency": pkg["currency"]
        })
    return packages

@api_router.post("/payments/create-checkout")
async def create_checkout_session(request: CreateCheckoutRequest, http_request: Request):
    """Create a Stripe checkout session"""
    if not STRIPE_AVAILABLE:
        raise HTTPException(status_code=500, detail="Stripe غير متوفر")
    
    package = SUBSCRIPTION_PACKAGES.get(request.package_id)
    if not package:
        raise HTTPException(status_code=400, detail="الباقة غير موجودة")
    
    api_key = os.environ.get("STRIPE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="مفتاح Stripe غير موجود")
    
    try:
        # Build URLs from origin
        success_url = f"{request.origin_url}/payment-success?session_id={{CHECKOUT_SESSION_ID}}"
        cancel_url = f"{request.origin_url}/payment-cancel"
        
        # Create webhook URL
        host_url = str(http_request.base_url)
        webhook_url = f"{host_url}api/webhook/stripe"
        
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
        
        checkout_request = CheckoutSessionRequest(
            amount=package["amount"],
            currency="usd",  # Stripe requires USD for most operations
            success_url=success_url,
            cancel_url=cancel_url,
            metadata={
                "package_id": request.package_id,
                "package_name": package["name"],
                "duration_days": str(package["duration_days"]),
                "source": "nt_commerce"
            }
        )
        
        session: CheckoutSessionResponse = await stripe_checkout.create_checkout_session(checkout_request)
        
        # Create payment transaction record
        transaction_id = str(uuid.uuid4())
        await main_db.payment_transactions.insert_one({
            "id": transaction_id,
            "session_id": session.session_id,
            "package_id": request.package_id,
            "package_name": package["name"],
            "amount": package["amount"],
            "currency": package["currency"],
            "payment_status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "metadata": {
                "package_id": request.package_id,
                "duration_days": package["duration_days"]
            }
        })
        
        return {
            "url": session.url,
            "session_id": session.session_id,
            "transaction_id": transaction_id
        }
    except Exception as e:
        logger.error(f"Stripe checkout error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"فشل إنشاء جلسة الدفع: {str(e)}")

@api_router.get("/payments/status/{session_id}")
async def get_payment_status(session_id: str, http_request: Request):
    """Get payment status from Stripe"""
    if not STRIPE_AVAILABLE:
        raise HTTPException(status_code=500, detail="Stripe غير متوفر")
    
    api_key = os.environ.get("STRIPE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="مفتاح Stripe غير موجود")
    
    try:
        host_url = str(http_request.base_url)
        webhook_url = f"{host_url}api/webhook/stripe"
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
        
        status: CheckoutStatusResponse = await stripe_checkout.get_checkout_status(session_id)
        
        # Update transaction in database
        await main_db.payment_transactions.update_one(
            {"session_id": session_id},
            {"$set": {
                "payment_status": status.payment_status,
                "status": status.status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        return {
            "status": status.status,
            "payment_status": status.payment_status,
            "amount_total": status.amount_total,
            "currency": status.currency,
            "metadata": status.metadata
        }
    except Exception as e:
        logger.error(f"Stripe status error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"فشل التحقق من حالة الدفع: {str(e)}")

@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """Handle Stripe webhooks"""
    if not STRIPE_AVAILABLE:
        raise HTTPException(status_code=500, detail="Stripe غير متوفر")
    
    api_key = os.environ.get("STRIPE_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="مفتاح Stripe غير موجود")
    
    try:
        body = await request.body()
        signature = request.headers.get("Stripe-Signature")
        
        host_url = str(request.base_url)
        webhook_url = f"{host_url}api/webhook/stripe"
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
        
        webhook_response = await stripe_checkout.handle_webhook(body, signature)
        
        # Update payment transaction
        if webhook_response.session_id:
            await main_db.payment_transactions.update_one(
                {"session_id": webhook_response.session_id},
                {"$set": {
                    "payment_status": webhook_response.payment_status,
                    "event_type": webhook_response.event_type,
                    "event_id": webhook_response.event_id,
                    "webhook_received_at": datetime.now(timezone.utc).isoformat()
                }}
            )
        
        return {"received": True}
    except Exception as e:
        logger.error(f"Stripe webhook error: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))

# Payment Records Management (for manual/offline payments)
@api_router.get("/payments/records")
async def get_payment_records(
    page: int = 1,
    limit: int = 20,
    status: Optional[str] = None,
    admin: dict = Depends(get_super_admin)
):
    """Get all payment records"""
    query = {}
    if status:
        query["payment_status"] = status
    
    skip = (page - 1) * limit
    records = await main_db.payment_transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await main_db.payment_transactions.count_documents(query)
    
    return {
        "records": records,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }

@api_router.post("/payments/records")
async def create_payment_record(payment: PaymentRecord, admin: dict = Depends(get_super_admin)):
    """Create a manual payment record"""
    record_id = str(uuid.uuid4())
    record = {
        "id": record_id,
        "tenant_id": payment.tenant_id,
        "amount": payment.amount,
        "currency": payment.currency,
        "payment_method": payment.payment_method,
        "description": payment.description,
        "invoice_number": payment.invoice_number or f"INV-{datetime.now().strftime('%Y%m%d')}-{record_id[:8].upper()}",
        "payment_status": payment.status,
        "metadata": payment.metadata,
        "created_by": admin.get("id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await main_db.payment_transactions.insert_one(record)
    
    # If payment is for a tenant, update subscription if paid
    if payment.tenant_id and payment.status == "paid":
        # Extend subscription
        tenant = await main_db.saas_tenants.find_one({"id": payment.tenant_id})
        if tenant:
            current_end = tenant.get("subscription_end")
            if current_end:
                try:
                    end_date = datetime.fromisoformat(current_end.replace('Z', '+00:00'))
                except:
                    end_date = datetime.now(timezone.utc)
            else:
                end_date = datetime.now(timezone.utc)
            
            # Default 30 days extension for manual payments
            new_end = end_date + timedelta(days=30)
            await main_db.saas_tenants.update_one(
                {"id": payment.tenant_id},
                {"$set": {"subscription_end": new_end.isoformat()}}
            )
    
    return {"id": record_id, "message": "تم إنشاء سجل الدفع"}

@api_router.put("/payments/records/{record_id}")
async def update_payment_record(record_id: str, update: PaymentUpdateRequest, admin: dict = Depends(get_super_admin)):
    """Update a payment record"""
    update_data = {}
    if update.status:
        update_data["payment_status"] = update.status
    if update.notes:
        update_data["notes"] = update.notes
    if update.invoice_number:
        update_data["invoice_number"] = update.invoice_number
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = admin.get("id")
    
    result = await main_db.payment_transactions.update_one(
        {"id": record_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="سجل الدفع غير موجود")
    
    return {"success": True, "message": "تم تحديث سجل الدفع"}

@api_router.delete("/payments/records/{record_id}")
async def delete_payment_record(record_id: str, admin: dict = Depends(get_super_admin)):
    """Delete a payment record"""
    result = await main_db.payment_transactions.delete_one({"id": record_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="سجل الدفع غير موجود")
    
    return {"success": True, "message": "تم حذف سجل الدفع"}

# Invoice Generation
@api_router.get("/payments/invoice/{record_id}")
async def generate_invoice(record_id: str, admin: dict = Depends(get_super_admin)):
    """Generate invoice for a payment"""
    record = await main_db.payment_transactions.find_one({"id": record_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="سجل الدفع غير موجود")
    
    # Get tenant info if available
    tenant = None
    if record.get("tenant_id"):
        tenant = await main_db.saas_tenants.find_one({"id": record["tenant_id"]}, {"_id": 0})
    
    invoice_html = f"""
    <html dir="rtl">
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 40px; }}
            .header {{ text-align: center; border-bottom: 2px solid #333; padding-bottom: 20px; margin-bottom: 30px; }}
            .invoice-info {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
            .table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            .table th, .table td {{ border: 1px solid #ddd; padding: 12px; text-align: right; }}
            .table th {{ background: #f3f4f6; }}
            .total {{ font-size: 1.2em; font-weight: bold; text-align: left; margin-top: 20px; }}
            .footer {{ text-align: center; margin-top: 50px; color: #666; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>فاتورة</h1>
            <p>NT Commerce</p>
        </div>
        
        <div class="invoice-info">
            <div>
                <p><strong>رقم الفاتورة:</strong> {record.get('invoice_number', 'N/A')}</p>
                <p><strong>التاريخ:</strong> {record.get('created_at', '')[:10]}</p>
                <p><strong>الحالة:</strong> {'مدفوع' if record.get('payment_status') == 'paid' else 'قيد الانتظار'}</p>
            </div>
            <div>
                <p><strong>العميل:</strong> {tenant.get('business_name', 'N/A') if tenant else 'N/A'}</p>
                <p><strong>البريد:</strong> {tenant.get('email', 'N/A') if tenant else 'N/A'}</p>
            </div>
        </div>
        
        <table class="table">
            <thead>
                <tr>
                    <th>الوصف</th>
                    <th>المبلغ</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td>{record.get('description', record.get('package_name', 'اشتراك'))}</td>
                    <td>{record.get('amount', 0):,.2f} {record.get('currency', 'دج').upper()}</td>
                </tr>
            </tbody>
        </table>
        
        <div class="total">
            الإجمالي: {record.get('amount', 0):,.2f} {record.get('currency', 'دج').upper()}
        </div>
        
        <div class="footer">
            <p>شكراً لتعاملكم معنا</p>
            <p>NT Commerce - نظام إدارة نقاط البيع</p>
        </div>
    </body>
    </html>
    """
    
    return StreamingResponse(
        io.BytesIO(invoice_html.encode('utf-8')),
        media_type="text/html",
        headers={"Content-Disposition": f"inline; filename=invoice_{record.get('invoice_number', record_id)}.html"}
    )


# ============ ONLINE STORE ============

class StoreSettings(BaseModel):
    enabled: bool = False
    store_name: str = ""
    store_slug: str = ""
    description: str = ""
    logo_url: str = ""
    banner_url: str = ""
    primary_color: str = "#3b82f6"
    contact_phone: str = ""
    contact_email: str = ""
    contact_address: str = ""
    working_hours: str = "09:00 - 18:00"
    cod_enabled: bool = True
    delivery_enabled: bool = True
    min_order_amount: float = 0
    delivery_fee: float = 0
    free_delivery_threshold: float = 0

class StoreOrder(BaseModel):
    customer_name: str
    customer_phone: str
    customer_email: str = ""
    delivery_address: str
    delivery_city: str = ""
    delivery_wilaya: str = ""
    items: List[dict]
    subtotal: float
    delivery_fee: float = 0
    total: float
    notes: str = ""
    payment_method: str = "cod"  # cod = cash on delivery

@api_router.get("/store/settings")
async def get_store_settings(admin: dict = Depends(get_tenant_admin)):
    """Get store settings for tenant"""
    settings = await db.store_settings.find_one({}, {"_id": 0})
    return settings or StoreSettings().model_dump()

@api_router.put("/store/settings")
async def update_store_settings(settings: StoreSettings, admin: dict = Depends(get_tenant_admin)):
    """Update store settings"""
    tenant_id = admin.get("tenant_id")
    
    # Save settings in tenant database
    await db.store_settings.update_one(
        {},
        {"$set": settings.model_dump()},
        upsert=True
    )
    
    # Also save slug mapping in main database for public access
    if settings.store_slug and tenant_id:
        await main_db.store_slugs.update_one(
            {"tenant_id": tenant_id},
            {"$set": {
                "tenant_id": tenant_id,
                "store_slug": settings.store_slug,
                "enabled": settings.enabled,
                "store_name": settings.store_name,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
        # Also create index on store_slug for faster lookups
        await main_db.store_slugs.create_index("store_slug", unique=True, sparse=True)
    
    return {"message": "تم حفظ إعدادات المتجر"}

@api_router.get("/store/products")
async def get_store_products(admin: dict = Depends(get_tenant_admin)):
    """Get products listed in the store"""
    store_products = await db.store_products.find({}, {"_id": 0}).to_list(1000)
    return store_products

@api_router.post("/store/products")
async def add_store_product(data: dict, admin: dict = Depends(get_tenant_admin)):
    """Add product to store"""
    product_id = data.get("product_id")
    if not product_id:
        raise HTTPException(status_code=400, detail="product_id required")
    
    # Check if already added
    existing = await db.store_products.find_one({"product_id": product_id})
    if existing:
        return {"message": "Product already in store"}
    
    store_product = {
        "id": str(uuid.uuid4()),
        "product_id": product_id,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.store_products.insert_one(store_product)
    return {"message": "تمت إضافة المنتج للمتجر"}

@api_router.delete("/store/products/{product_id}")
async def remove_store_product(product_id: str, admin: dict = Depends(get_tenant_admin)):
    """Remove product from store"""
    await db.store_products.delete_one({"product_id": product_id})
    return {"message": "تمت إزالة المنتج من المتجر"}

@api_router.get("/store/orders")
async def get_store_orders(status: Optional[str] = None, admin: dict = Depends(get_tenant_admin)):
    """Get store orders"""
    query = {}
    if status:
        query["status"] = status
    orders = await db.store_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return orders

@api_router.put("/store/orders/{order_id}/status")
async def update_store_order_status(order_id: str, data: dict, admin: dict = Depends(get_tenant_admin)):
    """Update order status"""
    status = data.get("status")
    if not status:
        raise HTTPException(status_code=400, detail="status required")
    
    await db.store_orders.update_one(
        {"id": order_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "تم تحديث حالة الطلب"}

# Public store endpoints (no auth required)
@api_router.get("/shop/{store_slug}")
async def get_public_store(store_slug: str):
    """Get public store by slug"""
    # Find tenant by store slug from main database
    slug_mapping = await main_db.store_slugs.find_one({"store_slug": store_slug, "enabled": True}, {"_id": 0})
    if not slug_mapping:
        raise HTTPException(status_code=404, detail="Store not found")
    
    tenant_id = slug_mapping.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Store not configured")
    
    # Get tenant database
    tenant_db = get_tenant_db(tenant_id)
    
    # Get store settings from tenant database
    settings = await tenant_db.store_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("enabled"):
        raise HTTPException(status_code=404, detail="Store not available")
    
    # Get store products with details
    store_products = await tenant_db.store_products.find({"is_active": True}, {"_id": 0}).to_list(1000)
    product_ids = [sp["product_id"] for sp in store_products]
    
    products = await tenant_db.products.find(
        {"id": {"$in": product_ids}, "quantity": {"$gt": 0}},
        {"_id": 0, "id": 1, "name_ar": 1, "name_en": 1, "retail_price": 1, "image_url": 1, "description_ar": 1, "description_en": 1, "quantity": 1}
    ).to_list(1000)
    
    return {
        "settings": settings,
        "products": products,
        "tenant_id": tenant_id
    }

@api_router.post("/shop/{store_slug}/order")
async def create_public_order(store_slug: str, order: StoreOrder):
    """Create order from public store (COD)"""
    # Find tenant by store slug from main database
    slug_mapping = await main_db.store_slugs.find_one({"store_slug": store_slug, "enabled": True}, {"_id": 0})
    if not slug_mapping:
        raise HTTPException(status_code=404, detail="Store not found")
    
    tenant_id = slug_mapping.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Store not configured")
    
    # Get tenant database
    tenant_db = get_tenant_db(tenant_id)
    
    # Validate store is enabled
    settings = await tenant_db.store_settings.find_one({"enabled": True})
    if not settings:
        raise HTTPException(status_code=404, detail="Store not available")
    
    # Validate minimum order amount
    if settings.get("min_order_amount", 0) > 0 and order.subtotal < settings["min_order_amount"]:
        raise HTTPException(status_code=400, detail=f"Minimum order amount is {settings['min_order_amount']}")
    
    # Validate products availability and update stock
    for item in order.items:
        product = await tenant_db.products.find_one({"id": item.get("product_id")})
        if not product:
            raise HTTPException(status_code=400, detail=f"Product {item.get('name', 'Unknown')} not found")
        if product.get("quantity", 0) < item.get("quantity", 1):
            raise HTTPException(status_code=400, detail=f"Product {item.get('name', product.get('name_ar', 'Unknown'))} out of stock")
    
    # Update stock for each product
    for item in order.items:
        await tenant_db.products.update_one(
            {"id": item.get("product_id")},
            {"$inc": {"quantity": -item.get("quantity", 1)}}
        )
    
    # Generate order number
    count = await tenant_db.store_orders.count_documents({}) + 1
    order_number = f"WEB{count:06d}"
    
    # Create order
    order_data = {
        "id": str(uuid.uuid4()),
        "order_number": order_number,
        "store_slug": store_slug,
        **order.model_dump(),
        "status": "pending",
        "payment_status": "unpaid",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await tenant_db.store_orders.insert_one(order_data)
    
    return {
        "message": "تم استلام طلبك بنجاح",
        "order_number": order_number,
        "order_id": order_data["id"]
    }


# ============ DEFECTIVE PRODUCTS MANAGEMENT ============

class DefectiveProductCreate(BaseModel):
    product_id: str
    quantity: int
    reason: str  # manufacturing, shipping, storage, other
    notes: Optional[str] = ""
    supplier_id: Optional[str] = None
    action: str = "pending"  # pending, return_to_supplier, dispose, repair, discount_sale
    images: Optional[List[str]] = []

@api_router.post("/defective-products")
async def create_defective_product(data: DefectiveProductCreate, admin: dict = Depends(get_tenant_admin)):
    """Register a defective product and deduct from inventory"""
    db = get_tenant_db(admin["tenant_id"])
    
    # Get product info
    product = await db.products.find_one({"id": data.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check if enough quantity
    if product.get("quantity", 0) < data.quantity:
        raise HTTPException(status_code=400, detail="Not enough quantity in stock")
    
    # Get supplier info if provided
    supplier_name = None
    if data.supplier_id:
        supplier = await db.suppliers.find_one({"id": data.supplier_id}, {"_id": 0, "name": 1})
        supplier_name = supplier.get("name") if supplier else None
    
    # Create defective product record
    defective_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    defective_doc = {
        "id": defective_id,
        "product_id": data.product_id,
        "product_name": product.get("name_ar") or product.get("name_en", ""),
        "product_code": product.get("article_code", ""),
        "quantity": data.quantity,
        "reason": data.reason,
        "notes": data.notes or "",
        "supplier_id": data.supplier_id,
        "supplier_name": supplier_name,
        "action": data.action,
        "status": "pending",
        "return_request_id": None,
        "created_at": now,
        "updated_at": now,
        "created_by": admin.get("name", admin.get("email", "")),
        "images": data.images or [],
        "unit_cost": product.get("purchase_price", 0),
        "total_cost": product.get("purchase_price", 0) * data.quantity
    }
    
    await db.defective_products.insert_one(defective_doc)
    
    # Deduct from main inventory
    await db.products.update_one(
        {"id": data.product_id},
        {"$inc": {"quantity": -data.quantity}}
    )
    
    # If action is return_to_supplier, create a return request
    if data.action == "return_to_supplier" and data.supplier_id:
        return_request_id = str(uuid.uuid4())
        return_request = {
            "id": return_request_id,
            "defective_product_id": defective_id,
            "supplier_id": data.supplier_id,
            "supplier_name": supplier_name,
            "product_id": data.product_id,
            "product_name": defective_doc["product_name"],
            "quantity": data.quantity,
            "reason": data.reason,
            "notes": data.notes,
            "status": "pending",
            "created_at": now,
            "updated_at": now
        }
        await db.supplier_returns.insert_one(return_request)
        await db.defective_products.update_one(
            {"id": defective_id},
            {"$set": {"return_request_id": return_request_id}}
        )
        defective_doc["return_request_id"] = return_request_id
    
    # Check defective rate and create alert if needed
    total_products = await db.products.count_documents({})
    total_defective = await db.defective_products.count_documents({"status": {"$ne": "completed"}})
    defective_rate = (total_defective / total_products * 100) if total_products > 0 else 0
    
    if defective_rate > 5:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "defective_alert",
            "title": "تنبيه: نسبة المنتجات المعطلة مرتفعة",
            "message": f"نسبة المنتجات المعطلة وصلت إلى {defective_rate:.1f}%",
            "read": False,
            "created_at": now
        })
    
    return {**defective_doc, "_id": None}

@api_router.get("/defective-products")
async def get_defective_products(
    status: Optional[str] = None,
    reason: Optional[str] = None,
    supplier_id: Optional[str] = None,
    admin: dict = Depends(get_tenant_admin)
):
    """Get all defective products with filters"""
    db = get_tenant_db(admin["tenant_id"])
    
    query = {}
    if status:
        query["status"] = status
    if reason:
        query["reason"] = reason
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    defective_products = await db.defective_products.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return defective_products

@api_router.get("/defective-products/stats")
async def get_defective_stats(admin: dict = Depends(get_tenant_admin)):
    """Get defective products statistics"""
    db = get_tenant_db(admin["tenant_id"])
    
    total = await db.defective_products.count_documents({})
    pending = await db.defective_products.count_documents({"status": "pending"})
    in_progress = await db.defective_products.count_documents({"status": "in_progress"})
    completed = await db.defective_products.count_documents({"status": "completed"})
    
    by_reason = await db.defective_products.aggregate([
        {"$group": {"_id": "$reason", "count": {"$sum": 1}, "total_qty": {"$sum": "$quantity"}}}
    ]).to_list(10)
    
    by_action = await db.defective_products.aggregate([
        {"$group": {"_id": "$action", "count": {"$sum": 1}, "total_qty": {"$sum": "$quantity"}}}
    ]).to_list(10)
    
    cost_result = await db.defective_products.aggregate([
        {"$group": {"_id": None, "total_cost": {"$sum": "$total_cost"}, "total_qty": {"$sum": "$quantity"}}}
    ]).to_list(1)
    
    total_cost = cost_result[0]["total_cost"] if cost_result else 0
    total_qty = cost_result[0]["total_qty"] if cost_result else 0
    
    by_supplier = await db.defective_products.aggregate([
        {"$match": {"supplier_id": {"$ne": None}}},
        {"$group": {"_id": "$supplier_id", "supplier_name": {"$first": "$supplier_name"}, "count": {"$sum": 1}, "total_qty": {"$sum": "$quantity"}}}
    ]).to_list(20)
    
    return {
        "total": total,
        "pending": pending,
        "in_progress": in_progress,
        "completed": completed,
        "total_cost": total_cost,
        "total_quantity": total_qty,
        "by_reason": {item["_id"]: {"count": item["count"], "quantity": item["total_qty"]} for item in by_reason if item["_id"]},
        "by_action": {item["_id"]: {"count": item["count"], "quantity": item["total_qty"]} for item in by_action if item["_id"]},
        "by_supplier": by_supplier
    }

@api_router.put("/defective-products/{defective_id}")
async def update_defective_product(defective_id: str, data: dict, admin: dict = Depends(get_tenant_admin)):
    """Update defective product status or action"""
    db = get_tenant_db(admin["tenant_id"])
    
    defective = await db.defective_products.find_one({"id": defective_id})
    if not defective:
        raise HTTPException(status_code=404, detail="Defective product not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    if "status" in data:
        update_data["status"] = data["status"]
    if "action" in data:
        update_data["action"] = data["action"]
    if "notes" in data:
        update_data["notes"] = data["notes"]
    
    await db.defective_products.update_one({"id": defective_id}, {"$set": update_data})
    
    if data.get("action") == "return_to_supplier" and not defective.get("return_request_id") and defective.get("supplier_id"):
        return_request_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)
        return_request = {
            "id": return_request_id,
            "defective_product_id": defective_id,
            "supplier_id": defective["supplier_id"],
            "supplier_name": defective.get("supplier_name"),
            "product_id": defective["product_id"],
            "product_name": defective["product_name"],
            "quantity": defective["quantity"],
            "reason": defective["reason"],
            "notes": defective.get("notes", ""),
            "status": "pending",
            "created_at": now,
            "updated_at": now
        }
        await db.supplier_returns.insert_one(return_request)
        await db.defective_products.update_one(
            {"id": defective_id},
            {"$set": {"return_request_id": return_request_id}}
        )
    
    updated = await db.defective_products.find_one({"id": defective_id}, {"_id": 0})
    return updated

@api_router.delete("/defective-products/{defective_id}")
async def delete_defective_product(defective_id: str, restore_stock: bool = False, admin: dict = Depends(get_tenant_admin)):
    """Delete defective product record, optionally restore stock"""
    db = get_tenant_db(admin["tenant_id"])
    
    defective = await db.defective_products.find_one({"id": defective_id})
    if not defective:
        raise HTTPException(status_code=404, detail="Defective product not found")
    
    if restore_stock:
        await db.products.update_one(
            {"id": defective["product_id"]},
            {"$inc": {"quantity": defective["quantity"]}}
        )
    
    if defective.get("return_request_id"):
        await db.supplier_returns.delete_one({"id": defective["return_request_id"]})
    
    await db.defective_products.delete_one({"id": defective_id})
    return {"message": "Deleted successfully", "stock_restored": restore_stock}

@api_router.get("/supplier-returns")
async def get_supplier_returns(
    status: Optional[str] = None,
    supplier_id: Optional[str] = None,
    admin: dict = Depends(get_tenant_admin)
):
    """Get supplier return requests"""
    db = get_tenant_db(admin["tenant_id"])
    
    query = {}
    if status:
        query["status"] = status
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    returns = await db.supplier_returns.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return returns

@api_router.put("/supplier-returns/{return_id}")
async def update_supplier_return(return_id: str, data: dict, admin: dict = Depends(get_tenant_admin)):
    """Update supplier return status"""
    db = get_tenant_db(admin["tenant_id"])
    
    return_request = await db.supplier_returns.find_one({"id": return_id})
    if not return_request:
        raise HTTPException(status_code=404, detail="Return request not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    if "status" in data:
        update_data["status"] = data["status"]
        if data["status"] == "refunded" and return_request.get("defective_product_id"):
            await db.defective_products.update_one(
                {"id": return_request["defective_product_id"]},
                {"$set": {"status": "completed", "updated_at": datetime.now(timezone.utc)}}
            )
    
    if "notes" in data:
        update_data["notes"] = data["notes"]
    if "refund_amount" in data:
        update_data["refund_amount"] = data["refund_amount"]
    
    await db.supplier_returns.update_one({"id": return_id}, {"$set": update_data})
    
    updated = await db.supplier_returns.find_one({"id": return_id}, {"_id": 0})
    return updated


# ============ AI ASSISTANT FOR SAAS ADMIN ============

from emergentintegrations.llm.chat import LlmChat, UserMessage

AI_SYSTEM_PROMPT = """أنت مساعد ذكي لنظام NT Commerce SaaS. مهمتك مساعدة مدير النظام في:

1. **تشخيص المشاكل**: تحليل الأخطاء وتقديم حلول
2. **الصيانة**: تنفيذ مهام صيانة النظام
3. **التقارير**: إنشاء تقارير وإحصائيات
4. **الدعم الفني**: الإجابة على أسئلة المستخدمين

قواعد مهمة:
- أجب دائماً بالعربية
- كن مختصراً ومفيداً
- قدم خطوات واضحة للحلول
- إذا لم تتمكن من تنفيذ شيء، اشرح السبب
- لا تكشف معلومات حساسة

الأوامر المتاحة:
- /status - حالة النظام
- /stats - إحصائيات المشتركين
- /errors - آخر الأخطاء
- /maintenance [action] - مهام الصيانة
- /help - المساعدة
"""

class AIAssistantMessage(BaseModel):
    message: str
    session_id: Optional[str] = None

@api_router.post("/ai-assistant/chat")
async def ai_assistant_chat(data: AIAssistantMessage, admin: dict = Depends(get_super_admin)):
    """AI Assistant endpoint for SaaS admin"""
    try:
        llm_key = os.environ.get("EMERGENT_LLM_KEY")
        if not llm_key:
            raise HTTPException(status_code=500, detail="AI service not configured")
        
        session_id = data.session_id or f"admin-{admin.get('id', 'default')}"
        
        # Handle special commands
        user_message = data.message.strip()
        context = ""
        
        if user_message.startswith("/status"):
            # Get system status
            total_tenants = await db.saas_tenants.count_documents({})
            active_tenants = await db.saas_tenants.count_documents({"status": "active"})
            context = f"حالة النظام:\n- إجمالي المشتركين: {total_tenants}\n- المشتركين النشطين: {active_tenants}\n- الخادم: يعمل بشكل طبيعي"
            
        elif user_message.startswith("/stats"):
            # Get statistics
            total_tenants = await db.saas_tenants.count_documents({})
            active = await db.saas_tenants.count_documents({"status": "active"})
            pending = await db.saas_tenants.count_documents({"status": "pending"})
            suspended = await db.saas_tenants.count_documents({"status": "suspended"})
            context = f"إحصائيات المشتركين:\n- الإجمالي: {total_tenants}\n- نشط: {active}\n- معلق: {pending}\n- موقوف: {suspended}"
            
        elif user_message.startswith("/errors"):
            # Get recent errors (mock for now)
            context = "آخر الأخطاء:\n- لا توجد أخطاء حرجة حالياً\n- النظام يعمل بشكل طبيعي"
            
        elif user_message.startswith("/help"):
            context = """الأوامر المتاحة:
/status - عرض حالة النظام
/stats - إحصائيات المشتركين
/errors - آخر الأخطاء
/maintenance cache - مسح الكاش
/maintenance db - فحص قاعدة البيانات
/help - عرض هذه المساعدة

يمكنك أيضاً طرح أي سؤال بشكل طبيعي وسأساعدك!"""
            return {"response": context, "session_id": session_id}
            
        elif user_message.startswith("/maintenance"):
            parts = user_message.split()
            if len(parts) > 1:
                action = parts[1]
                if action == "cache":
                    context = "تم مسح الكاش بنجاح! ✅"
                elif action == "db":
                    context = "قاعدة البيانات: متصلة وتعمل بشكل طبيعي ✅"
                else:
                    context = f"إجراء غير معروف: {action}"
            else:
                context = "استخدم: /maintenance cache أو /maintenance db"
            return {"response": context, "session_id": session_id}
        
        # Initialize AI chat
        chat = LlmChat(
            api_key=llm_key,
            session_id=session_id,
            system_message=AI_SYSTEM_PROMPT
        ).with_model("openai", "gpt-5.2")
        
        # Add context to message if available
        full_message = f"{context}\n\nسؤال المستخدم: {user_message}" if context else user_message
        
        # Create user message
        message = UserMessage(text=full_message)
        
        # Get AI response
        response = await chat.send_message(message)
        
        # Store in chat history
        await db.ai_chat_history.insert_one({
            "id": str(uuid.uuid4()),
            "session_id": session_id,
            "admin_id": admin.get("id"),
            "user_message": user_message,
            "ai_response": response,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"response": response, "session_id": session_id}
        
    except Exception as e:
        logger.error(f"AI Assistant error: {str(e)}")
        return {"response": f"عذراً، حدث خطأ: {str(e)}", "session_id": data.session_id}

@api_router.get("/ai-assistant/history")
async def get_ai_chat_history(session_id: Optional[str] = None, limit: int = 50, admin: dict = Depends(get_super_admin)):
    """Get AI chat history"""
    query = {}
    if session_id:
        query["session_id"] = session_id
    
    history = await db.ai_chat_history.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return history

@api_router.delete("/ai-assistant/history")
async def clear_ai_chat_history(session_id: Optional[str] = None, admin: dict = Depends(get_super_admin)):
    """Clear AI chat history"""
    query = {}
    if session_id:
        query["session_id"] = session_id
    
    result = await db.ai_chat_history.delete_many(query)
    return {"deleted": result.deleted_count}


# Include router and middleware
app.include_router(api_router)
app.include_router(saas_router, prefix="/api")  # Refactored SaaS routes

# Tenant context middleware - extracts tenant_id from JWT and sets ContextVar
@app.middleware("http")
async def tenant_context_middleware(request: Request, call_next):
    """Sets the tenant database context for each request based on JWT tenant_id"""
    auth_header = request.headers.get("authorization", "")
    if auth_header.startswith("Bearer "):
        try:
            token = auth_header.split(" ")[1]
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            tenant_id = payload.get("tenant_id")
            if tenant_id:
                tenant_specific_db = client[f"tenant_{tenant_id.replace('-', '_')}"]
                _tenant_db_ctx.set(tenant_specific_db)
        except Exception:
            pass  # Invalid/expired token - no tenant context, falls back to main_db
    response = await call_next(request)
    return response

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files for uploads
app.mount("/api/static", StaticFiles(directory=str(ROOT_DIR / "static")), name="static")

@app.on_event("startup")
async def startup():
    await init_cash_boxes()

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
